views_code = r'''

# ==============================================================================
# TENTATIVE LEAVE PLAN
# ==============================================================================

@login_required
def plan_my_plan(request):
    from .models import TentativeLeavePlan
    emp = request.user.employee
    year = int(request.GET.get("year", timezone.now().year))
    entries = TentativeLeavePlan.objects.filter(employee=emp, year=year).select_related("leave_type")
    leave_types = LeaveType.objects.filter(is_active=True)
    years = range(timezone.now().year, timezone.now().year + 3)

    if request.method == "POST":
        action = request.POST.get("action")
        if action == "add":
            start = request.POST.get("planned_start")
            end   = request.POST.get("planned_end")
            lt_id = request.POST.get("leave_type")
            notes = request.POST.get("notes", "").strip()
            if not start or not end:
                messages.error(request, "Please provide both start and end dates.")
            elif start > end:
                messages.error(request, "End date must be after start date.")
            else:
                TentativeLeavePlan.objects.create(
                    employee=emp, year=year,
                    planned_start=start, planned_end=end,
                    leave_type_id=lt_id if lt_id else None,
                    notes=notes, status="draft",
                )
                messages.success(request, "Period added to your plan.")
        elif action == "delete":
            entry_id = request.POST.get("entry_id")
            TentativeLeavePlan.objects.filter(pk=entry_id, employee=emp, status__in=["draft","rejected"]).delete()
            messages.success(request, "Entry removed.")
        elif action == "submit":
            drafts = TentativeLeavePlan.objects.filter(employee=emp, year=year, status__in=["draft","rejected"])
            if not drafts.exists():
                messages.warning(request, "No draft entries to submit.")
            else:
                count = drafts.count()
                drafts.update(status="submitted", submitted_at=timezone.now())
                messages.success(request, f"{count} plan entries submitted to your Line Manager.")
        return redirect(request.path + "?year=" + str(year))

    return render(request, "leaves/plan_my_plan.html", {
        "entries": entries,
        "year": year,
        "years": years,
        "leave_types": leave_types,
        "has_draft": entries.filter(status__in=["draft","rejected"]).exists(),
    })


@login_required
def plan_manager_review(request):
    from .models import TentativeLeavePlan
    emp = request.user.employee
    if not (emp.is_manager() or request.user.is_superuser):
        messages.error(request, "Access denied.")
        return redirect("dashboard:home")
    year = int(request.GET.get("year", timezone.now().year))
    team = emp.subordinates.filter(is_active=True)
    base_qs = TentativeLeavePlan.objects.filter(
        employee__in=team, year=year
    ).select_related("employee__user", "employee__department", "leave_type").order_by("employee__user__last_name", "planned_start")

    if request.method == "POST":
        action   = request.POST.get("action")
        entry_id = request.POST.get("entry_id")
        note     = request.POST.get("manager_notes", "").strip()
        try:
            entry = TentativeLeavePlan.objects.get(pk=entry_id, employee__in=team)
        except TentativeLeavePlan.DoesNotExist:
            messages.error(request, "Entry not found.")
            return redirect(request.path + "?year=" + str(year))
        if action == "confirm":
            entry.status = "confirmed"
            entry.manager_confirmed_by = request.user
            entry.manager_confirmed_at = timezone.now()
            entry.manager_notes = note
            entry.save()
            messages.success(request, f"Plan confirmed for {entry.employee.get_full_name()}.")
        elif action == "reject":
            entry.status = "rejected"
            entry.manager_confirmed_by = request.user
            entry.manager_confirmed_at = timezone.now()
            entry.manager_notes = note
            entry.save()
            messages.warning(request, f"Plan rejected for {entry.employee.get_full_name()}.")
        return redirect(request.path + "?year=" + str(year))

    years = range(timezone.now().year, timezone.now().year + 3)
    return render(request, "leaves/plan_manager.html", {
        "pending":   base_qs.filter(status="submitted"),
        "confirmed": base_qs.filter(status="confirmed"),
        "rejected":  base_qs.filter(status="rejected"),
        "year": year, "years": years,
    })


@login_required
def plan_hr_overview(request):
    from .models import TentativeLeavePlan
    from accounts.models import Department
    emp = request.user.employee
    if not (emp.is_hr() or request.user.is_superuser):
        messages.error(request, "Access denied.")
        return redirect("dashboard:home")
    year    = int(request.GET.get("year", timezone.now().year))
    dept_id = request.GET.get("dept", "")
    status_f = request.GET.get("status", "confirmed")
    qs = TentativeLeavePlan.objects.filter(year=year).select_related(
        "employee__user", "employee__department", "leave_type", "manager_confirmed_by"
    ).order_by("employee__department__name", "employee__user__last_name", "planned_start")
    if dept_id:
        qs = qs.filter(employee__department_id=dept_id)
    if status_f:
        qs = qs.filter(status=status_f)
    departments = Department.objects.all().order_by("name")
    years = range(timezone.now().year, timezone.now().year + 3)

    if request.GET.get("export") == "excel":
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse as _HR
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Leave Plan {year}"
        hfont  = Font(bold=True, color="FFFFFF", size=10)
        hfill  = PatternFill("solid", fgColor="0A4D68")
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
        thin   = Side(style="thin", color="CCCCCC")
        bdr    = Border(left=thin, right=thin, top=thin, bottom=thin)
        headers = ["Department","Employee","Leave Type","Start Date","End Date","Days","Notes","Status","Confirmed By"]
        widths  = [22, 28, 18, 14, 14, 8, 30, 14, 22]
        ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
        ws["A1"].value = f"Tentative Leave Plan {year}"
        ws["A1"].font  = Font(bold=True, size=13, color="0A4D68")
        ws["A1"].alignment = center
        ws.row_dimensions[1].height = 26
        for col, (h, w) in enumerate(zip(headers, widths), 1):
            c = ws.cell(row=2, column=col, value=h)
            c.font = hfont; c.fill = hfill; c.alignment = center; c.border = bdr
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.row_dimensions[2].height = 20
        entries_list = list(qs)
        for ri, e in enumerate(entries_list, 3):
            vals = [
                str(e.employee.department) if e.employee.department else "---",
                e.employee.get_full_name(),
                str(e.leave_type) if e.leave_type else "---",
                e.planned_start.strftime("%d/%m/%Y"),
                e.planned_end.strftime("%d/%m/%Y"),
                e.total_days,
                e.notes or "",
                e.get_status_display(),
                e.manager_confirmed_by.get_full_name() if e.manager_confirmed_by else "---",
            ]
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=ri, column=col, value=val)
                c.border = bdr
                c.alignment = center if col in (4,5,6,8) else left
            ws.row_dimensions[ri].height = 16
        ws.freeze_panes = "A3"
        resp = _HR(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp["Content-Disposition"] = f"attachment; filename=tentative_plan_{year}.xlsx"
        wb.save(resp)
        return resp

    entries_list = list(qs)
    return render(request, "leaves/plan_hr.html", {
        "entries":       entries_list,
        "year":          year,
        "years":         years,
        "departments":   departments,
        "dept_id":       dept_id,
        "status_filter": status_f,
        "total":         len(entries_list),
        "total_days":    sum(e.total_days for e in entries_list),
    })
'''

with open(r'c:\Users\GROD\Desktop\leave_system\leaves\views.py', 'a', encoding='utf-8') as f:
    f.write(views_code)
print('Views appended successfully.')
