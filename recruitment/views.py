import json
import os
import re
import threading
import uuid
import requests as http_requests
from django.core.files.storage import default_storage
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.views.decorators.cache import never_cache
from django.contrib import messages
from django.core.mail import send_mail
from django.conf import settings
from django.db import models, transaction
from django.db.models import Max
from django.urls import reverse
from django.utils import timezone
from django.core.cache import cache
from django.http import FileResponse, Http404, HttpResponse, JsonResponse

from accounts.models import Department, Employee
from notifications.utils import notify
from .models import (
    JobPosting, FormFieldConfig, ScoringCriterion,
    Application, ApplicationAnswer,
    FIELD_TYPE_TEXT, FIELD_TYPE_CHOICES, APPLICATION_STATUS_CHOICES,
)


# ─── Security helpers ──────────────────────────────────────────────────────────

def _get_client_ip(request):
    x_forwarded = request.META.get('HTTP_X_FORWARDED_FOR')
    return x_forwarded.split(',')[0].strip() if x_forwarded else request.META.get('REMOTE_ADDR', '127.0.0.1')

def _check_rate_limit(key, max_calls, period_seconds):
    """Return True if allowed, False if rate-limited."""
    count = cache.get(key, 0)
    if count >= max_calls:
        return False
    cache.set(key, count + 1, period_seconds)
    return True


# ─── Applicant email helper ────────────────────────────────────────────────────

_STATUS_CONFIG = {
    'submitted': {
        'subject': 'Application Received — {title} | MICEI',
        'headline': 'We Received Your Application!',
        'color': '#0A4D68',
        'icon': '✅',
        'body': (
            'Thank you for applying for the <strong>{title}</strong> position at '
            'Magrabi ICO Cameroon Eye Institute (MICEI).<br><br>'
            'We have received your application and our HR team will begin reviewing it shortly. '
            'You will be notified by email as your application progresses through our selection process.'
        ),
        'next': 'Our HR team will carefully review your application and reach out to you regarding the next steps.',
    },
    'under_review': {
        'subject': 'Your Application is Under Review — {title}',
        'headline': 'Application Under Review',
        'color': '#088395',
        'icon': '🔍',
        'body': (
            'Thank you for applying for the <strong>{title}</strong> position at '
            'Magrabi ICO Cameroon Eye Institute (MICEI).<br><br>'
            'We are pleased to inform you that your application is currently under review '
            'by our HR team. We will be in touch with further updates.'
        ),
        'next': 'Our team will review your application and notify you of the next steps.',
    },
    'shortlisted': {
        'subject': "You've Been Shortlisted — {title}",
        'headline': "You've Been Shortlisted! 🎉",
        'color': '#0A4D68',
        'icon': '⭐',
        'body': (
            'Congratulations! We are pleased to inform you that your application for '
            'the <strong>{title}</strong> position has been shortlisted.<br><br>'
            'This means your profile stood out among all candidates. Well done!'
        ),
        'next': 'Our HR team will contact you shortly with the next steps in our selection process.',
    },
    'interview': {
        'subject': 'Interview Invitation — {title}',
        'headline': 'You Are Invited for an Interview!',
        'color': '#6366f1',
        'icon': '📅',
        'body': (
            'Congratulations! You have been selected for an interview for the '
            '<strong>{title}</strong> position at MICEI.<br><br>'
            '{interview_block}'
            'Please confirm your availability by replying to this email or contacting our HR department.'
        ),
        'next': 'Prepare for your interview — review the job description and research MICEI at micei.org.',
    },
    'offered': {
        'subject': 'Job Offer — {title} | MICEI',
        'headline': 'Congratulations — You Have Received an Offer!',
        'color': '#059669',
        'icon': '🏆',
        'body': (
            'We are delighted to extend a job offer for the <strong>{title}</strong> position '
            'at Magrabi ICO Cameroon Eye Institute.<br><br>'
            'Our HR team will contact you shortly with your formal offer letter and all the details '
            'regarding your compensation, start date, and onboarding process.'
        ),
        'next': 'Please wait for your formal offer letter from our HR team.',
    },
    'hired': {
        'subject': 'Welcome to MICEI — {title}',
        'headline': 'Welcome to the MICEI Family! 🎊',
        'color': '#059669',
        'icon': '🎊',
        'body': (
            'We are thrilled to confirm that you have been selected for the '
            '<strong>{title}</strong> position at Magrabi ICO Cameroon Eye Institute.<br><br>'
            'Welcome aboard! You are joining a team dedicated to restoring vision and improving '
            'lives across Central Africa.'
        ),
        'next': 'Our HR team will reach out with your onboarding details and start date confirmation.',
    },
    'rejected': {
        'subject': 'Application Update — {title}',
        'headline': 'Thank You for Applying',
        'color': '#64748b',
        'icon': '📝',
        'body': (
            'Thank you for your interest in the <strong>{title}</strong> position at MICEI and '
            'for the time you invested in your application.<br><br>'
            'After careful consideration, we regret to inform you that we will not be moving '
            'forward with your application at this time.{rejection_block}'
            '<br><br>We encourage you to apply for future openings that match your qualifications.'
        ),
        'next': 'We wish you all the best in your career journey. Keep an eye on our careers page for future opportunities.',
    },
}

_EMAIL_HTML = """<!DOCTYPE html>
<html>
<head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"></head>
<body style="margin:0;padding:0;background:#f4f6f8;font-family:Arial,Helvetica,sans-serif;">
<table width="100%" cellpadding="0" cellspacing="0" style="background:#f4f6f8;padding:32px 16px;">
<tr><td align="center">
<table width="580" cellpadding="0" cellspacing="0" style="background:#ffffff;border-top:4px solid #0A4D68;">
  <!-- Logo -->
  <tr><td style="padding:28px 40px 20px;border-bottom:1px solid #e8edf2;">
    <img src="{logo_url}" alt="MICEI" style="height:44px;width:auto;display:block;">
  </td></tr>
  <!-- Body -->
  <tr><td style="padding:32px 40px;">
    <p style="margin:0 0 20px;color:#1a2b3c;font-size:1rem;">Dear <strong>{name}</strong>,</p>
    <p style="margin:0 0 20px;color:#374151;font-size:.93rem;line-height:1.75;">{body}</p>
    <p style="margin:0 0 6px;color:#374151;font-size:.88rem;font-weight:600;">What happens next?</p>
    <p style="margin:0 0 20px;color:#4b5563;font-size:.88rem;line-height:1.65;">{next}</p>
    <p style="margin:0 0 28px;color:#6b7280;font-size:.83rem;line-height:1.6;font-style:italic;">Kindly note that due to the high volume of applications received, only shortlisted candidates who progress to the next stage will be contacted.</p>
    <table cellpadding="0" cellspacing="0"><tr><td style="background:#0A4D68;border-radius:4px;">
      <a href="{status_url}" style="display:inline-block;padding:11px 28px;color:#ffffff;font-size:.88rem;font-weight:700;text-decoration:none;">Check My Application Status</a>
    </td></tr></table>
  </td></tr>
  <!-- Footer -->
  <tr><td style="padding:20px 40px;border-top:1px solid #e8edf2;">
    <p style="margin:0 0 6px;font-size:.78rem;color:#6b7280;font-weight:600;">Magrabi ICO Cameroon Eye Institute (MICEI)</p>
    <p style="margin:0;font-size:.72rem;color:#b0b8c4;">This is an automated message — please do not reply directly to this email.</p>
  </td></tr>
</table>
</td></tr>
</table>
</body></html>"""


def _email_applicant(applicant_name, applicant_email, status, posting_title,
                     interview_date=None, rejection_reason='', status_url=''):
    """Send a branded HTML status-update email to an applicant."""
    if not getattr(settings, 'EMAIL_NOTIFICATIONS_ENABLED', False):
        return
    if not applicant_email:
        return
    cfg = _STATUS_CONFIG.get(status)
    if not cfg:
        return

    # Build dynamic blocks
    interview_block = ''
    if status == 'interview' and interview_date:
        fmt = interview_date.strftime('%A, %d %B %Y at %H:%M')
        interview_block = (
            f'<div style="background:#f0f0ff;border-radius:8px;padding:14px 18px;margin:16px 0;">'
            f'<strong style="color:#6366f1;">📅 Interview Scheduled:</strong><br>'
            f'<span style="font-size:1.05rem;color:#1a2b3c;font-weight:700;">{fmt}</span>'
            f'</div>'
        )

    rejection_block = ''
    if status == 'rejected' and rejection_reason:
        rejection_block = (
            f'<br><br><div style="background:#fef2f2;border-left:4px solid #dc2626;border-radius:6px;'
            f'padding:12px 16px;margin-top:8px;">'
            f'<strong style="color:#991b1b;font-size:.85rem;">Feedback:</strong><br>'
            f'<span style="color:#374151;font-size:.88rem;">{rejection_reason}</span>'
            f'</div>'
        )

    body_html = cfg['body'].format(
        title=posting_title,
        interview_block=interview_block,
        rejection_block=rejection_block,
    )

    site_url = getattr(settings, 'SITE_URL', 'https://hr.micei.org')
    logo_url = f'{site_url}/static/LOGO.png'
    html = _EMAIL_HTML.format(
        headline=cfg['headline'],
        name=applicant_name,
        body=body_html,
        next=cfg['next'],
        status_url=status_url or f'{site_url}/recruitment/jobs/my-application/',
        logo_url=logo_url,
    )

    plain = (
        f"Dear {applicant_name},\n\n"
        f"Position: {posting_title}\n\n"
        + (f"Interview: {interview_date.strftime('%A, %d %B %Y at %H:%M')}\n\n" if interview_date else '')
        + (f"Feedback: {rejection_reason}\n\n" if rejection_reason else '')
        + f"{cfg['next']}\n\n"
        f"Check your application status: {status_url}\n\n"
        f"Best regards,\nMICEI HR Team"
    )

    try:
        send_mail(
            subject=cfg['subject'].format(title=posting_title),
            message=plain,
            from_email=getattr(settings, 'DEFAULT_FROM_EMAIL', 'MICEI Careers <careers@micei.org>'),
            recipient_list=[applicant_email],
            html_message=html,
            fail_silently=True,
        )
    except Exception:
        pass


def _notify_hr_email(applicant_name, applicant_email, posting_title, detail_url):
    """Send a notification email to HR when a new application is submitted."""
    if not getattr(settings, 'EMAIL_NOTIFICATIONS_ENABLED', False):
        return
    site_url = getattr(settings, 'SITE_URL', 'https://hr.micei.org')
    logo_url = f'{site_url}/static/LOGO.png'
    subject = f'New Application — {posting_title}'
    html = f"""<!DOCTYPE html>
<html>
<head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"></head>
<body style="margin:0;padding:0;background:#f4f6f8;font-family:Arial,Helvetica,sans-serif;">
<table width="100%" cellpadding="0" cellspacing="0" style="background:#f4f6f8;padding:32px 16px;">
<tr><td align="center">
<table width="580" cellpadding="0" cellspacing="0" style="background:#ffffff;border-top:4px solid #0A4D68;">
  <tr><td style="padding:28px 40px 20px;border-bottom:1px solid #e8edf2;">
    <img src="{logo_url}" alt="MICEI" style="height:44px;width:auto;display:block;">
  </td></tr>
  <tr><td style="padding:32px 40px;">
    <p style="margin:0 0 20px;color:#1a2b3c;font-size:1rem;">Hello HR Team,</p>
    <p style="margin:0 0 20px;color:#374151;font-size:.93rem;line-height:1.75;">
      A new application has been submitted for the <strong>{posting_title}</strong> position. Please review it at your earliest convenience.
    </p>
    <table cellpadding="0" cellspacing="0" style="width:100%;border:1px solid #e8edf2;margin-bottom:24px;">
      <tr style="background:#f8fafc;"><td style="padding:10px 16px;font-size:.82rem;color:#6b7280;font-weight:700;width:110px;">Applicant</td><td style="padding:10px 16px;font-size:.88rem;color:#1a2b3c;">{applicant_name}</td></tr>
      <tr><td style="padding:10px 16px;font-size:.82rem;color:#6b7280;font-weight:700;border-top:1px solid #f1f5f9;">Email</td><td style="padding:10px 16px;font-size:.88rem;color:#1a2b3c;border-top:1px solid #f1f5f9;">{applicant_email}</td></tr>
      <tr style="background:#f8fafc;"><td style="padding:10px 16px;font-size:.82rem;color:#6b7280;font-weight:700;border-top:1px solid #f1f5f9;">Position</td><td style="padding:10px 16px;font-size:.88rem;color:#1a2b3c;border-top:1px solid #f1f5f9;">{posting_title}</td></tr>
    </table>
    <table cellpadding="0" cellspacing="0"><tr><td style="background:#0A4D68;border-radius:4px;">
      <a href="{detail_url}" style="display:inline-block;padding:11px 28px;color:#ffffff;font-size:.88rem;font-weight:700;text-decoration:none;">Review Application</a>
    </td></tr></table>
  </td></tr>
  <tr><td style="padding:20px 40px;border-top:1px solid #e8edf2;">
    <p style="margin:0 0 4px;font-size:.78rem;color:#6b7280;font-weight:600;">Magrabi ICO Cameroon Eye Institute (MICEI) — HRM System</p>
    <p style="margin:0;font-size:.72rem;color:#b0b8c4;">This is an automated notification — please do not reply directly to this email.</p>
  </td></tr>
</table>
</td></tr>
</table>
</body></html>"""
    plain = f"New application received.\nApplicant: {applicant_name}\nEmail: {applicant_email}\nPosition: {posting_title}\nReview: {detail_url}"
    try:
        from_email = getattr(settings, 'DEFAULT_FROM_EMAIL', 'MICEI HRM <pm@hr.micei.org>')
        hr_email = getattr(settings, 'HR_EMAIL', 'hr@micei.org')
        send_mail(subject, plain, from_email, [hr_email], html_message=html, fail_silently=True)
    except Exception:
        pass


# ─── Permission helpers ───────────────────────────────────────────────────────

def _is_hr_or_admin(user):
    if user.is_superuser:
        return True
    try:
        emp = user.employee
        return emp.is_hr() or emp.is_director() or emp.is_ceo()
    except Exception:
        return False


# ─── Public views (no login required) ─────────────────────────────────────────

@never_cache
def job_board(request):
    """Public job board: list all open postings that haven't passed their deadline."""
    from django.utils import timezone as _tz
    today = _tz.localdate()
    postings = list(
        JobPosting.objects
        .filter(status=JobPosting.STATUS_OPEN)
        .filter(models.Q(deadline__isnull=True) | models.Q(deadline__gte=today))
        .select_related('department')
    )
    dept_counts = {}
    for p in postings:
        key = str(p.department) if p.department else 'General'
        dept_counts[key] = dept_counts.get(key, 0) + 1
    departments = [{'name': k, 'count': v} for k, v in sorted(dept_counts.items())]
    return render(request, 'recruitment/job_board.html', {
        'postings': postings,
        'departments': departments,
    })


@never_cache
def job_detail(request, pk):
    """Public: view job description."""
    posting = get_object_or_404(JobPosting, pk=pk, status=JobPosting.STATUS_OPEN)
    return render(request, 'recruitment/job_detail.html', {'posting': posting})


@never_cache
def apply(request, pk):
    """Public: submit an application for a job posting."""
    from django.utils import timezone as _tz
    posting = get_object_or_404(JobPosting, pk=pk, status=JobPosting.STATUS_OPEN)
    # Block if deadline has passed
    if posting.deadline and posting.deadline < _tz.localdate():
        messages.error(request, 'The application deadline for this position has passed.')
        return redirect('recruitment:job_detail', pk=pk)
    fields = posting.form_fields.filter(is_enabled=True).order_by('field_order', 'pk')

    if request.method == 'POST':
        # Honeypot: bots fill this hidden field, humans leave it empty
        if request.POST.get('_url_confirm', ''):
            return redirect('recruitment:apply_success', pk=posting.pk)

        # Rate limit: max 5 submissions per IP per hour
        ip = _get_client_ip(request)
        if not _check_rate_limit(f'apply_{ip}', 5, 3600):
            messages.error(request, 'Too many submissions from your connection. Please try again later.')
            return render(request, 'recruitment/apply.html', {
                'posting': posting, 'fields': fields,
                'applicant_name': '', 'applicant_email': '',
            })

        name  = request.POST.get('applicant_name', '').strip()
        email = request.POST.get('applicant_email', '').strip()
        cv    = request.FILES.get('cv_file')

        import os as _os
        _ALLOWED_CV_EXTS = {'.pdf', '.doc', '.docx'}
        _MAX_CV_BYTES = 10 * 1024 * 1024  # 10 MB

        from django.core.validators import validate_email
        from django.core.exceptions import ValidationError as _VE
        errors = []
        if not name:
            errors.append('Full name is required.')
        if not email:
            errors.append('Email address is required.')
        else:
            try:
                validate_email(email)
            except _VE:
                errors.append('Please enter a valid email address.')
        if not cv:
            errors.append('Please upload your CV / résumé.')
        else:
            ext = _os.path.splitext(cv.name)[1].lower()
            if ext not in _ALLOWED_CV_EXTS:
                errors.append('CV must be a PDF, DOC, or DOCX file.')
            elif cv.size > _MAX_CV_BYTES:
                errors.append('CV file size must not exceed 10 MB.')

        # Additional documents are required
        if not request.FILES.getlist('extra_docs'):
            errors.append('Please upload at least one additional document (certificate, diploma, licence, etc.).')

        # Validate enabled required fields
        for field in fields:
            if field.is_required:
                if field.field_type == 'file':
                    val = request.FILES.get(field.field_name)
                elif field.field_type == 'file_multi':
                    val = request.FILES.getlist(field.field_name)
                else:
                    val = request.POST.get(field.field_name, '').strip()
                if not val:
                    errors.append(f'{field.label} is required.')

        if errors:
            for e in errors:
                messages.error(request, e)
            return render(request, 'recruitment/apply.html', {
                'posting':         posting,
                'fields':          fields,
                'applicant_name':  name,
                'applicant_email': email,
            })

        # Duplicate check: same email cannot apply to the same posting twice
        if Application.objects.filter(posting=posting, applicant_email__iexact=email).exists():
            messages.error(request, 'You have already submitted an application for this position. Check your status using the "My Application" button.')
            return render(request, 'recruitment/apply.html', {
                'posting':         posting,
                'fields':          fields,
                'applicant_name':  name,
                'applicant_email': email,
            })

        with transaction.atomic():
            # Save CV with UUID filename (same as other uploads)
            import os as _os
            _cv_ext = _os.path.splitext(cv.name)[1].lower()
            _cv_safe = uuid.uuid4().hex + _cv_ext
            _cv_path = os.path.join('recruitment', 'cvs', _tz.now().strftime('%Y'), _tz.now().strftime('%m'), _cv_safe)
            _saved_cv = default_storage.save(_cv_path, cv)
            app = Application.objects.create(
                posting=posting,
                applicant_name=name,
                applicant_email=email,
                cv_file=_saved_cv,
            )
            # Save answers for all enabled fields
            answers_to_create = []
            _ALLOWED_DOC_EXTS = {'.pdf', '.doc', '.docx', '.jpg', '.jpeg', '.png'}
            _MAX_CV_BYTES = 10 * 1024 * 1024
            for field in fields:
                if field.field_type == 'file':
                    uploaded_file = request.FILES.get(field.field_name)
                    if uploaded_file:
                        ext = os.path.splitext(uploaded_file.name)[1].lower()
                        safe_name = uuid.uuid4().hex + ext
                        save_path = os.path.join('recruitment', 'field_uploads', safe_name)
                        saved = default_storage.save(save_path, uploaded_file)
                        val = saved
                    else:
                        val = ''
                    answers_to_create.append(ApplicationAnswer(
                        application=app, field_name=field.field_name, value=val,
                    ))
                elif field.field_type == 'file_multi':
                    # Save each file as fieldname_0, fieldname_1, …
                    multi_files = request.FILES.getlist(field.field_name)
                    for i, mf in enumerate(multi_files[:10]):
                        ext = _os.path.splitext(mf.name)[1].lower()
                        if ext in _ALLOWED_DOC_EXTS and mf.size <= _MAX_CV_BYTES:
                            safe_name = uuid.uuid4().hex + ext
                            save_path = os.path.join('recruitment', 'field_uploads', safe_name)
                            saved = default_storage.save(save_path, mf)
                            answers_to_create.append(ApplicationAnswer(
                                application=app,
                                field_name=f'{field.field_name}_{i}',
                                value=saved,
                            ))
                else:
                    val = request.POST.get(field.field_name, '').strip()
                    answers_to_create.append(ApplicationAnswer(
                        application=app, field_name=field.field_name, value=val,
                    ))
            ApplicationAnswer.objects.bulk_create(answers_to_create)

            # Save extra/additional documents (multiple files)
            _ALLOWED_DOC_EXTS = {'.pdf', '.doc', '.docx', '.jpg', '.jpeg', '.png'}
            extra_files = request.FILES.getlist('extra_docs')
            extra_answers = []
            for i, doc in enumerate(extra_files[:10]):  # cap at 10 extra docs
                ext = _os.path.splitext(doc.name)[1].lower()
                if ext in _ALLOWED_DOC_EXTS and doc.size <= _MAX_CV_BYTES:
                    safe_name = uuid.uuid4().hex + ext
                    save_path = os.path.join('recruitment', 'extra_docs', safe_name)
                    saved = default_storage.save(save_path, doc)
                    extra_answers.append(ApplicationAnswer(
                        application=app,
                        field_name=f'extra_doc_{i}',
                        value=saved,
                    ))
            if extra_answers:
                ApplicationAnswer.objects.bulk_create(extra_answers)

            # Auto-score (rule-based)
            app.compute_score()

        # AI analysis in background — runs after the transaction commits
        if getattr(settings, 'GEMINI_API_KEY', None):
            threading.Thread(
                target=_ai_analyse_background,
                args=(posting, app),
                daemon=True,
            ).start()

        # Notify all HR users of the new application
        detail_url = reverse('recruitment:applicant_detail', kwargs={'posting_pk': posting.pk, 'pk': app.pk})
        for hr_emp in Employee.objects.filter(role='hr', is_active=True).select_related('user'):
            notify(
                hr_emp.user,
                f'New Application — {posting.title}',
                f'{name} has submitted an application for the {posting.title} position.\n\n'
                f'Score: {int(app.score)} pts  |  Email: {email}',
                notification_type='system',
                url=detail_url,
            )

        # Confirmation email to applicant
        status_url = request.build_absolute_uri(reverse('recruitment:check_status'))
        threading.Thread(
            target=_email_applicant,
            args=(name, email, 'submitted', posting.title),
            kwargs={'status_url': status_url},
            daemon=True,
        ).start()

        # Email notification to HR
        abs_detail_url = request.build_absolute_uri(detail_url)
        threading.Thread(
            target=_notify_hr_email,
            args=(name, email, posting.title, abs_detail_url),
            daemon=True,
        ).start()

        return redirect('recruitment:apply_success', pk=posting.pk)

    return render(request, 'recruitment/apply.html', {
        'posting':         posting,
        'fields':          fields,
        'applicant_name':  '',
        'applicant_email': '',
    })


@never_cache
def apply_success(request, pk):
    posting = get_object_or_404(JobPosting, pk=pk)
    return render(request, 'recruitment/apply_success.html', {'posting': posting})


@never_cache
def check_status(request):
    """Public: applicant can look up their application status by email."""
    applications = []
    searched = False
    email = ''
    if request.method == 'POST':
        # Honeypot
        if request.POST.get('_url_confirm', ''):
            return render(request, 'recruitment/check_status.html', {
                'applications': [], 'searched': True, 'email': '',
            })
        # Rate limit: 15 lookups per IP per 10 minutes
        ip = _get_client_ip(request)
        if not _check_rate_limit(f'status_{ip}', 15, 600):
            messages.error(request, 'Too many requests. Please wait a few minutes.')
            return render(request, 'recruitment/check_status.html', {
                'applications': [], 'searched': False, 'email': '',
            })
        email = request.POST.get('email', '').strip().lower()
        searched = True
        if email:
            applications = list(
                Application.objects
                .filter(applicant_email__iexact=email)
                .select_related('posting')
                .order_by('-submitted_at')
            )
    return render(request, 'recruitment/check_status.html', {
        'applications': applications,
        'searched': searched,
        'email': email,
    })


# ─── Protected file serving (HR only) ─────────────────────────────────────────

@login_required
def serve_recruitment_file(request, filepath):
    """Serve recruitment uploads only to logged-in HR/admin users."""
    if not _is_hr_or_admin(request.user):
        raise Http404
    # Normalize path and ensure it stays inside MEDIA_ROOT/recruitment/
    full_path = os.path.realpath(os.path.join(settings.MEDIA_ROOT, filepath))
    allowed_root = os.path.realpath(os.path.join(settings.MEDIA_ROOT, 'recruitment'))
    if not full_path.startswith(allowed_root + os.sep):
        raise Http404
    if not os.path.isfile(full_path):
        raise Http404
    return FileResponse(open(full_path, 'rb'))


# ─── HR views (login required) ────────────────────────────────────────────────

@login_required
def posting_list(request):
    if not _is_hr_or_admin(request.user):
        messages.error(request, 'Only HR and above can access recruitment management.')
        return redirect('dashboard:home')

    postings = JobPosting.objects.select_related('department', 'created_by').all()
    status_filter = request.GET.get('status', '')
    if status_filter:
        postings = postings.filter(status=status_filter)

    return render(request, 'recruitment/posting_list.html', {
        'postings':      postings,
        'status_filter': status_filter,
        'STATUS_CHOICES': JobPosting.STATUS_CHOICES,
    })


@login_required
def posting_create(request):
    if not _is_hr_or_admin(request.user):
        messages.error(request, 'Only HR and above can create job postings.')
        return redirect('recruitment:list')

    departments = Department.objects.all()

    if request.method == 'POST':
        title  = request.POST.get('title', '').strip()
        desc   = request.POST.get('description', '').strip()
        if not title or not desc:
            messages.error(request, 'Title and description are required.')
            return render(request, 'recruitment/posting_form.html', {
                'departments': departments,
                'type_choices': JobPosting.TYPE_CHOICES,
                'post': request.POST,
            })
        dept_pk = request.POST.get('department')
        dept = Department.objects.filter(pk=dept_pk).first() if dept_pk else None

        dl_str = request.POST.get('deadline', '').strip()
        deadline = None
        if dl_str:
            from datetime import date
            try:
                deadline = date.fromisoformat(dl_str)
            except ValueError:
                pass

        posting = JobPosting.objects.create(
            title=title,
            department=dept,
            location=request.POST.get('location', '').strip(),
            employment_type=request.POST.get('employment_type', JobPosting.TYPE_FULL_TIME),
            about=request.POST.get('about', '').strip(),
            description=desc,
            requirements=request.POST.get('requirements', '').strip(),
            status=request.POST.get('status', JobPosting.STATUS_DRAFT),
            deadline=deadline,
            created_by=request.user,
        )
        posting.create_default_fields()
        messages.success(request, f'Job posting "{posting.title}" created. Configure the application form below.')
        return redirect('recruitment:form_config', pk=posting.pk)

    return render(request, 'recruitment/posting_form.html', {
        'departments':  departments,
        'type_choices': JobPosting.TYPE_CHOICES,
        'post': {},
    })


@login_required
def posting_edit(request, pk):
    if not _is_hr_or_admin(request.user):
        messages.error(request, 'Only HR and above can edit job postings.')
        return redirect('recruitment:list')

    posting = get_object_or_404(JobPosting, pk=pk)
    departments = Department.objects.all()

    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        desc  = request.POST.get('description', '').strip()
        if not title or not desc:
            messages.error(request, 'Title and description are required.')
        else:
            dept_pk = request.POST.get('department')
            dept = Department.objects.filter(pk=dept_pk).first() if dept_pk else None
            dl_str = request.POST.get('deadline', '').strip()
            deadline = None
            if dl_str:
                from datetime import date
                try:
                    deadline = date.fromisoformat(dl_str)
                except ValueError:
                    pass
            posting.title           = title
            posting.department      = dept
            posting.location        = request.POST.get('location', '').strip()
            posting.employment_type = request.POST.get('employment_type', JobPosting.TYPE_FULL_TIME)
            posting.about           = request.POST.get('about', '').strip()
            posting.description     = desc
            posting.requirements    = request.POST.get('requirements', '').strip()
            posting.status          = request.POST.get('status', posting.status)
            posting.deadline        = deadline
            posting.save()
            messages.success(request, 'Job posting updated.')
            return redirect('recruitment:list')

    return render(request, 'recruitment/posting_form.html', {
        'posting':      posting,
        'departments':  departments,
        'type_choices': JobPosting.TYPE_CHOICES,
        'post': posting,
    })


@login_required
def posting_delete(request, pk):
    if not _is_hr_or_admin(request.user):
        messages.error(request, 'Only HR and above can delete job postings.')
        return redirect('recruitment:list')
    posting = get_object_or_404(JobPosting, pk=pk)
    if request.method == 'POST':
        title = posting.title
        posting.delete()
        messages.success(request, f'Job posting "{title}" deleted.')
    return redirect('recruitment:list')


@login_required
def form_config(request, pk):
    """Configure which fields appear on the application form for this posting."""
    if not _is_hr_or_admin(request.user):
        messages.error(request, 'Only HR and above can configure application forms.')
        return redirect('recruitment:list')

    posting = get_object_or_404(JobPosting, pk=pk)

    if request.method == 'POST':
        action = request.POST.get('action', '')

        if action == 'save_fields':
            valid_types = [c[0] for c in FIELD_TYPE_CHOICES]
            fields = posting.form_fields.all()
            for field in fields:
                enabled   = f'enabled_{field.pk}' in request.POST
                required  = f'required_{field.pk}' in request.POST
                label     = request.POST.get(f'label_{field.pk}', field.label).strip() or field.label
                order     = request.POST.get(f'order_{field.pk}', str(field.field_order))
                ph        = request.POST.get(f'placeholder_{field.pk}', '').strip()
                options   = request.POST.get(f'options_{field.pk}', field.options).strip()
                new_type  = request.POST.get(f'type_{field.pk}', field.field_type)
                try:
                    order = int(order)
                except ValueError:
                    order = field.field_order
                field.is_enabled   = enabled
                field.is_required  = required and enabled
                field.label        = label
                field.field_order  = order
                field.placeholder  = ph
                field.options      = options
                if new_type in valid_types:
                    field.field_type = new_type
                field.save(update_fields=['is_enabled', 'is_required', 'label', 'field_order', 'placeholder', 'options', 'field_type'])
            messages.success(request, 'Form configuration saved.')
            return redirect('recruitment:form_config', pk=pk)

        if action == 'add_custom':
            label      = request.POST.get('new_label', '').strip()
            field_type = request.POST.get('new_field_type', FIELD_TYPE_TEXT)
            required   = 'new_required' in request.POST
            options    = request.POST.get('new_options', '').strip()
            ph         = request.POST.get('new_placeholder', '').strip()
            if not label:
                messages.error(request, 'A label is required for the custom field.')
            else:
                # Generate a safe field_name from label
                field_name = re.sub(r'[^a-z0-9_]', '_', label.lower())[:60]
                field_name = f'custom_{field_name}'
                # Ensure uniqueness within this posting
                base = field_name
                i = 2
                while posting.form_fields.filter(field_name=field_name).exists():
                    field_name = f'{base}_{i}'
                    i += 1
                max_order = posting.form_fields.aggregate(m=Max('field_order'))['m'] or 0
                FormFieldConfig.objects.create(
                    posting=posting,
                    field_name=field_name,
                    label=label,
                    field_type=field_type,
                    is_enabled=True,
                    is_required=required,
                    field_order=max_order + 1,
                    options=options,
                    placeholder=ph,
                    is_custom=True,
                )
                messages.success(request, f'Custom field "{label}" added.')
            return redirect('recruitment:form_config', pk=pk)

        if action == 'delete_field':
            field_pk = request.POST.get('field_pk')
            field = posting.form_fields.filter(pk=field_pk).first()
            if field:
                field.delete()
                messages.success(request, 'Custom field deleted.')
            return redirect('recruitment:form_config', pk=pk)

    fields = posting.form_fields.all()
    return render(request, 'recruitment/form_config.html', {
        'posting':        posting,
        'fields':         fields,
        'FIELD_TYPE_CHOICES': FIELD_TYPE_CHOICES,
    })


@login_required
def scoring_config(request, pk):
    """Configure auto-scoring criteria for a posting."""
    if not _is_hr_or_admin(request.user):
        messages.error(request, 'Only HR and above can configure scoring criteria.')
        return redirect('recruitment:list')

    posting = get_object_or_404(JobPosting, pk=pk)
    fields  = posting.form_fields.filter(is_enabled=True).order_by('field_order')

    if request.method == 'POST':
        action = request.POST.get('action', '')

        if action == 'add_criterion':
            field_name = request.POST.get('field_name', '').strip()
            label      = request.POST.get('label', '').strip()
            condition  = request.POST.get('condition', '')
            value      = request.POST.get('value', '').strip()
            points_str = request.POST.get('points', '0')
            try:
                points = int(points_str)
            except ValueError:
                points = 0
            if not field_name or not label or not condition:
                messages.error(request, 'Field, label, and condition are required.')
            else:
                ScoringCriterion.objects.create(
                    posting=posting,
                    field_name=field_name,
                    label=label,
                    condition=condition,
                    value=value,
                    points=points,
                )
                messages.success(request, f'Scoring rule "{label}" added.')
            return redirect('recruitment:scoring_config', pk=pk)

        if action == 'delete_criterion':
            crit_pk = request.POST.get('criterion_pk')
            ScoringCriterion.objects.filter(pk=crit_pk, posting=posting).delete()
            messages.success(request, 'Scoring rule deleted.')
            return redirect('recruitment:scoring_config', pk=pk)

        if action == 'rescore_all':
            count = 0
            for app in posting.applications.prefetch_related('answers'):
                app.compute_score()
                count += 1
            messages.success(request, f'Re-scored {count} application(s).')
            return redirect('recruitment:scoring_config', pk=pk)

    criteria = posting.scoring_criteria.all()
    total_max_score = sum(c.points for c in criteria if c.points > 0)
    return render(request, 'recruitment/scoring_config.html', {
        'posting':         posting,
        'criteria':        criteria,
        'fields':          fields,
        'COND_CHOICES':    ScoringCriterion.COND_CHOICES,
        'total_max_score': total_max_score,
    })


@login_required
def applicant_list(request, pk):
    """HR: ranked list of applicants for a posting."""
    if not _is_hr_or_admin(request.user):
        return redirect('recruitment:list')

    posting = get_object_or_404(JobPosting, pk=pk)
    apps    = posting.applications.all().order_by('-ai_score', '-score')

    status_filter = request.GET.get('status', '')
    ai_rec_filter = request.GET.get('ai_rec', '')

    if status_filter:
        apps = apps.filter(status=status_filter)
    if ai_rec_filter:
        apps = apps.filter(ai_recommendation=ai_rec_filter)

    return render(request, 'recruitment/applicant_list.html', {
        'posting':        posting,
        'applications':   apps,
        'status_filter':  status_filter,
        'ai_rec_filter':  ai_rec_filter,
        'STATUS_CHOICES': APPLICATION_STATUS_CHOICES,
    })


@login_required
def applicant_export_excel(request, pk):
    """Export ranked applicant list to Excel (.xlsx)."""
    if not _is_hr_or_admin(request.user):
        return redirect('recruitment:list')

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    posting = get_object_or_404(JobPosting, pk=pk)
    apps = posting.applications.all().order_by('-ai_score', '-score')

    ai_rec_filter = request.GET.get('ai_rec', '')
    if ai_rec_filter in ('invite', 'hold', 'reject'):
        apps = apps.filter(ai_recommendation=ai_rec_filter)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Applicants'

    # ── Styles ──
    header_font    = Font(bold=True, color='FFFFFF', size=10)
    header_fill    = PatternFill('solid', fgColor='0A4D68')
    subheader_fill = PatternFill('solid', fgColor='2db4c3')
    center         = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left           = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    thin           = Side(style='thin', color='CCCCCC')
    border         = Border(left=thin, right=thin, top=thin, bottom=thin)

    rec_fills = {
        'invite': PatternFill('solid', fgColor='D1FAE5'),
        'hold':   PatternFill('solid', fgColor='FEF3C7'),
        'reject': PatternFill('solid', fgColor='FEE2E2'),
    }

    # ── Get enabled form fields (exclude file uploads) ──
    file_types = ('file', 'file_multi')
    form_fields = list(
        posting.form_fields.filter(is_enabled=True)
        .exclude(field_type__in=file_types)
        .order_by('field_order', 'pk')
    )

    # ── Pre-fetch all answers keyed by (app_id, field_name) ──
    from recruitment.models import ApplicationAnswer
    all_answers = ApplicationAnswer.objects.filter(
        application__in=apps
    ).values_list('application_id', 'field_name', 'value')
    answers_map = {}
    for app_id, fn, val in all_answers:
        answers_map[(app_id, fn)] = val

    filter_label = {'invite': ' — Invite Only', 'hold': ' — Hold Only', 'reject': ' — Rejected Only'}.get(ai_rec_filter, '')

    # ── Fixed columns ──
    fixed_headers = [
        'Rank', 'Name', 'Email', 'Applied Date', 'Applied Time', 'Status',
        'Rule Score', 'AI Score (/100)', 'AI Recommendation',
        'AI Summary', 'Strengths', 'Gaps',
    ]
    fixed_widths = [6, 24, 30, 14, 12, 14, 11, 14, 16, 40, 35, 35]

    # ── Dynamic columns from form fields ──
    dynamic_headers = [f.label for f in form_fields]
    dynamic_widths  = [max(18, len(f.label) + 4) for f in form_fields]

    all_headers = fixed_headers + dynamic_headers
    all_widths  = fixed_widths  + dynamic_widths
    total_cols  = len(all_headers)

    # ── Title row ──
    ws.merge_cells(f'A1:{get_column_letter(total_cols)}1')
    title_cell = ws['A1']
    title_cell.value     = f'Applicants — {posting.title}{filter_label}'
    title_cell.font      = Font(bold=True, size=13, color='0A4D68')
    title_cell.alignment = center
    ws.row_dimensions[1].height = 28

    # ── Header row ──
    for col, (h, w) in enumerate(zip(all_headers, all_widths), 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill if col <= len(fixed_headers) else subheader_fill
        cell.alignment = center
        cell.border    = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 22

    # ── Data rows ──
    for rank, app in enumerate(apps, 1):
        row = rank + 2
        rec = (app.ai_recommendation or '').lower()
        row_fill = rec_fills.get(rec)

        fixed_values = [
            rank,
            app.applicant_name,
            app.applicant_email,
            app.submitted_at.strftime('%d %b %Y') if app.submitted_at else '',
            app.submitted_at.strftime('%H:%M') if app.submitted_at else '',
            app.get_status_display(),
            int(app.score) if app.score is not None else '',
            int(app.ai_score) if app.ai_score is not None else 'N/A',
            (app.ai_recommendation or '').upper() or '—',
            app.ai_summary or '',
            app.ai_strengths or '',
            app.ai_gaps or '',
        ]
        dynamic_values = [
            answers_map.get((app.pk, f.field_name), '') for f in form_fields
        ]
        all_values = fixed_values + dynamic_values

        centered_cols = {1, 5, 6, 7, 8, 9}
        for col, val in enumerate(all_values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border    = border
            cell.alignment = center if col in centered_cols else left
            if row_fill:
                cell.fill = row_fill
        ws.row_dimensions[row].height = 18

    # Freeze header rows
    ws.freeze_panes = 'A3'

    # ── Write response ──
    filename = f'applicants_{posting.pk}_{timezone.now().strftime("%Y%m%d")}.xlsx'
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def applicant_detail(request, posting_pk, pk):
    """HR: full applicant profile."""
    if not _is_hr_or_admin(request.user):
        return redirect('recruitment:list')

    posting = get_object_or_404(JobPosting, pk=posting_pk)
    app        = get_object_or_404(Application, pk=pk, posting=posting)
    answer_map = {a.field_name: a.value for a in app.answers.all()}
    fields     = posting.form_fields.filter(is_enabled=True).order_by('field_order')

    # For file_multi fields, collect all saved file paths grouped by field
    field_multi_files = {}  # {field_name: [path, path, ...]}
    for key, val in answer_map.items():
        if val:
            for f in fields.filter(field_type='file_multi'):
                if key.startswith(f'{f.field_name}_'):
                    field_multi_files.setdefault(f.field_name, []).append(val)

    field_answers = []
    for f in fields:
        if f.field_type == 'file_multi':
            field_answers.append((f, field_multi_files.get(f.field_name, [])))
        else:
            field_answers.append((f, answer_map.get(f.field_name, '')))

    extra_docs = [
        (a.field_name, a.value)
        for a in app.answers.filter(field_name__startswith='extra_doc_').order_by('field_name')
        if a.value
    ]

    if request.method == 'POST':
        action = request.POST.get('action', '')

        if action == 'update_status':
            new_status = request.POST.get('status', app.status)
            valid = [s[0] for s in APPLICATION_STATUS_CHOICES]
            if new_status in valid:
                app.status = new_status
                app.reviewed_by = request.user
                if new_status == 'rejected':
                    app.rejection_reason = request.POST.get('rejection_reason', '').strip()
                if new_status == 'interview':
                    dt_str = request.POST.get('interview_date', '').strip()
                    if dt_str:
                        from datetime import datetime
                        try:
                            app.interview_date = datetime.fromisoformat(dt_str)
                        except ValueError:
                            pass
                app.save()
                # Email the applicant about their status change (background thread)
                status_url = request.build_absolute_uri(
                    reverse('recruitment:check_status')
                )
                threading.Thread(
                    target=_email_applicant,
                    args=(app.applicant_name, app.applicant_email, new_status, posting.title),
                    kwargs={
                        'interview_date': app.interview_date if new_status == 'interview' else None,
                        'rejection_reason': app.rejection_reason if new_status == 'rejected' else '',
                        'status_url': status_url,
                    },
                    daemon=True,
                ).start()
                messages.success(request, f'Status updated to {app.get_status_display()}.')
            return redirect('recruitment:applicant_detail', posting_pk=posting_pk, pk=pk)

        if action == 'save_notes':
            app.hr_notes = request.POST.get('hr_notes', '').strip()
            app.interview_notes = request.POST.get('interview_notes', '').strip()
            app.save(update_fields=['hr_notes', 'interview_notes'])
            messages.success(request, 'Notes saved.')
            return redirect('recruitment:applicant_detail', posting_pk=posting_pk, pk=pk)

    return render(request, 'recruitment/applicant_detail.html', {
        'posting':       posting,
        'app':           app,
        'field_answers': field_answers,
        'extra_docs':    extra_docs,
        'STATUS_CHOICES': APPLICATION_STATUS_CHOICES,
    })


# ─── AI Analysis ────────────────────────────────────────────────────────────────

def _ai_analyse_background(posting, app):
    """Run AI analysis in a background thread after application submission."""
    import logging
    logger = logging.getLogger(__name__)
    try:
        data = _call_gemini(posting, app)
        app.ai_score          = float(data.get('score', 0))
        app.ai_recommendation = data.get('recommendation', '').lower()
        app.ai_summary        = data.get('summary', '')
        app.ai_strengths      = data.get('strengths', '')
        app.ai_gaps           = data.get('gaps', '')
        app.ai_analysed_at    = timezone.now()
        app.save(update_fields=[
            'ai_score', 'ai_recommendation', 'ai_summary',
            'ai_strengths', 'ai_gaps', 'ai_analysed_at',
        ])
    except Exception as exc:
        logger.warning('Auto AI analysis failed for application %s: %s', app.pk, exc)


def _extract_cv_text(cv_file):
    """
    Extract plain text from the applicant's CV (PDF only).
    - First tries direct text extraction (works for digital/typed PDFs).
    - If result is too short (< 80 chars), assumes a scanned PDF and falls back
      to OCR using Tesseract (supports English + French).
    Returns empty string on failure.
    """
    try:
        import fitz  # PyMuPDF
        with fitz.open(cv_file.path) as doc:
            text = '\n'.join(page.get_text() for page in doc).strip()

        if len(text) >= 80:
            return text[:5000]

        # Scanned PDF — fall back to OCR
        try:
            import pytesseract
            from PIL import Image
            import io
            ocr_pages = []
            with fitz.open(cv_file.path) as doc:
                for page in doc:
                    # Render page at 200 DPI for good OCR accuracy
                    pix = page.get_pixmap(dpi=200)
                    img = Image.open(io.BytesIO(pix.tobytes('png')))
                    ocr_pages.append(
                        pytesseract.image_to_string(img, lang='eng+fra')
                    )
            return '\n'.join(ocr_pages)[:5000]
        except Exception:
            return text  # Return whatever we had even if short

    except Exception:
        return ''


def _call_gemini(posting, application):
    """
    Send the applicant profile to Gemini 2.5 Flash and return a parsed dict with
    keys: score (0-100), recommendation (invite|hold|reject), summary, strengths, gaps.
    Raises requests.HTTPError or ValueError on failure.
    """
    answers = {a.field_name: a.value for a in application.answers.all()}
    cv_text = _extract_cv_text(application.cv_file)

    # Extract text from any extra uploaded PDFs and append to cv_text
    extra_texts = []
    for key, path in answers.items():
        if key.startswith('extra_doc_') and path and path.lower().endswith('.pdf'):
            try:
                import fitz
                full_path = os.path.join(settings.MEDIA_ROOT, path)
                with fitz.open(full_path) as doc:
                    extra_texts.append('\n'.join(page.get_text() for page in doc)[:2000])
            except Exception:
                pass
    if extra_texts:
        cv_text = (cv_text + '\n\n--- Additional Documents ---\n' + '\n\n'.join(extra_texts))[:6000]

    # Use actual form field labels configured by HR for this posting
    field_label_map = {
        fc.field_name: fc.label
        for fc in posting.form_fields.filter(is_enabled=True).order_by('field_order', 'pk')
    }
    builtin_labels = {
        'education_level':  'Education Level',
        'years_experience': 'Years of Experience',
        'current_employer': 'Current / Last Employer',
        'expected_salary':  'Expected Salary',
        'cover_letter':     'Cover Letter / Motivation',
        'nationality':      'Nationality',
        'available_from':   'Available From',
        'source':           'How they heard about us',
    }
    answer_lines = []
    for key, val in answers.items():
        if val and str(val).strip():
            label = field_label_map.get(key) or builtin_labels.get(key) or key.replace('_', ' ').title()
            answer_lines.append(f'- {label}: {str(val).strip()}')
    applicant_block = '\n'.join(answer_lines) or '(No form answers provided.)'
    cv_block = f'\n\nCV / Resume Content:\n{cv_text}' if cv_text else '\n\n(No CV text could be extracted — penalise if CV was expected.)'

    requirements_block = posting.requirements.strip() if posting.requirements.strip() else None
    description_block  = posting.description.strip()  if posting.description.strip()  else None

    if requirements_block:
        requirements_section = (
            'JOB REQUIREMENTS — evaluate the applicant STRICTLY against each one:\n'
            + requirements_block
        )
    else:
        requirements_section = (
            'JOB REQUIREMENTS: Not explicitly listed.\n'
            'Use the job description below as the benchmark. Be conservative — '
            'if a qualification is not clearly evidenced, treat it as missing.'
        )

    description_section = (
        f'JOB DESCRIPTION (additional context on expectations):\n{description_block}'
    ) if description_block else ''

    prompt = f"""You are a strict HR screening analyst. Your role is to protect the organisation from weak hires by evaluating candidates objectively and without leniency.

POSITION: {posting.title} ({posting.get_employment_type_display()})
DEPARTMENT: {posting.department or 'Not specified'}

{requirements_section}

{description_section}

APPLICANT: {application.applicant_name}
=== Application Form Answers ===
{applicant_block}
=== End of Form ==={cv_block}

MANDATORY EVALUATION RULES:
1. Score ONLY based on clear evidence in the form answers and CV — do NOT assume or infer qualifications that are not explicitly stated.
2. If a requirement is not clearly evidenced, mark it as a GAP. Giving benefit of the doubt is not permitted.
3. A vague cover letter or missing answers for required fields must reduce the score.
4. If no CV is provided or CV text is missing, penalise heavily (deduct at least 20 points).
5. Consider professionalism, clarity, and completeness of the applicant's responses.
6. Each stated requirement that is NOT met must appear in the gaps field.

Scoring thresholds:
- 80–100 → meets all or nearly all requirements → recommend: invite
- 50–79  → meets some requirements but has notable gaps → recommend: hold
- 0–49   → does not meet the requirements → recommend: reject

Respond with ONLY valid JSON (no markdown, no extra text). For strengths and gaps, write each point as a short, clear sentence on its own line starting with "• ". Use plain language an HR officer can read at a glance — no jargon.
{{"score": <integer 0-100>, "recommendation": "<invite|hold|reject>", "summary": "<2-sentence overall fit verdict in plain language>", "strengths": "<each strength on its own line starting with • >", "gaps": "<each gap on its own line starting with • >"}}"""

    import time as _time
    payload = {
        'contents': [{'parts': [{'text': prompt}]}],
        'generationConfig': {
            'temperature': 0.1,
            'maxOutputTokens': 600,
            'responseMimeType': 'application/json',
            'thinkingConfig': {'thinkingBudget': 0},
        },
    }
    url = (
        f'https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent'
        f'?key={settings.GEMINI_API_KEY}'
    )
    for attempt in range(3):
        resp = http_requests.post(url, json=payload, timeout=40)
        if resp.status_code == 429:
            wait = 15 * (attempt + 1)   # 15s, 30s, 45s
            _time.sleep(wait)
            continue
        break
    resp.raise_for_status()
    candidate = resp.json()['candidates'][0]
    if candidate.get('finishReason') == 'MAX_TOKENS':
        raise ValueError('Gemini response was truncated (MAX_TOKENS). Try a shorter CV.')
    raw = candidate['content']['parts'][0]['text']
    import re as _re
    clean = _re.sub(r'^```[a-z]*\n?', '', raw.strip())
    clean = _re.sub(r'\n?```$', '', clean).strip()
    return json.loads(clean)


@login_required
def ai_analyse(request, pk):
    """AJAX endpoint: analyse one (app_pk in POST) or all applicants for a posting."""
    if not _is_hr_or_admin(request.user):
        return JsonResponse({'error': 'Forbidden'}, status=403)

    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    if not getattr(settings, 'GEMINI_API_KEY', None):
        return JsonResponse({'error': 'GEMINI_API_KEY is not configured. Add it to your .env file.'}, status=503)

    posting = get_object_or_404(JobPosting, pk=pk)
    app_pk  = request.POST.get('app_pk', '').strip()

    if app_pk:
        apps = [get_object_or_404(Application, pk=app_pk, posting=posting)]
    else:
        apps = list(posting.applications.all())

    import time

    results = []
    errors  = []
    for idx, app in enumerate(apps):
        if idx > 0:
            time.sleep(5)   # 5-second gap → ~12 req/min, safely under free-tier limit
        try:
            data = _call_gemini(posting, app)
            app.ai_score          = float(data.get('score', 0))
            app.ai_recommendation = data.get('recommendation', '').lower()
            app.ai_summary        = data.get('summary', '')
            app.ai_strengths      = data.get('strengths', '')
            app.ai_gaps           = data.get('gaps', '')
            app.ai_analysed_at    = timezone.now()
            app.save(update_fields=[
                'ai_score', 'ai_recommendation', 'ai_summary',
                'ai_strengths', 'ai_gaps', 'ai_analysed_at',
            ])
            results.append({
                'app_pk':            app.pk,
                'ai_score':          app.ai_score,
                'ai_recommendation': app.ai_recommendation,
                'ai_summary':        app.ai_summary,
                'ai_strengths':      app.ai_strengths,
                'ai_gaps':           app.ai_gaps,
            })
        except Exception as exc:
            errors.append({'app_pk': app.pk, 'name': app.applicant_name, 'error': str(exc)})

    return JsonResponse({'results': results, 'errors': errors})


@login_required
def shortlist_report(request, pk):
    """Generate a PDF shortlist report for the CEO."""
    if not _is_hr_or_admin(request.user):
        raise Http404

    from io import BytesIO
    from django.http import HttpResponse
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        HRFlowable, KeepTogether,
    )
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    posting = get_object_or_404(JobPosting, pk=pk)

    # Which candidates to include: filter by status or AI recommendation
    filter_by = request.GET.get('filter', 'invite')  # 'invite', 'all', 'shortlisted'
    apps = posting.applications.all().order_by('-ai_score', '-score')
    if filter_by == 'invite':
        apps = apps.filter(ai_recommendation='invite')
    elif filter_by == 'shortlisted':
        apps = apps.filter(status='shortlisted')
    apps = list(apps)

    # ── Document setup ──────────────────────────────────────────────────────────
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm,
    )

    PRIMARY   = colors.HexColor('#0A4D68')
    ACCENT    = colors.HexColor('#2db4c3')
    INVITE_BG = colors.HexColor('#d1fae5')
    HOLD_BG   = colors.HexColor('#fef3c7')
    REJECT_BG = colors.HexColor('#fee2e2')
    LIGHT_BG  = colors.HexColor('#f8f9fa')
    WHITE     = colors.white
    DARK_TEXT = colors.HexColor('#1f2937')
    MUTED     = colors.HexColor('#6b7280')

    styles = getSampleStyleSheet()
    h1  = ParagraphStyle('h1',  fontSize=20, textColor=WHITE,     leading=24, spaceAfter=4,  fontName='Helvetica-Bold', alignment=TA_CENTER)
    h2  = ParagraphStyle('h2',  fontSize=11, textColor=WHITE,     leading=14, spaceAfter=0,  fontName='Helvetica-Bold', alignment=TA_CENTER)
    h3  = ParagraphStyle('h3',  fontSize=10, textColor=PRIMARY,   leading=13, spaceBefore=4, fontName='Helvetica-Bold')
    bod = ParagraphStyle('bod', fontSize=8,  textColor=DARK_TEXT, leading=11, fontName='Helvetica')
    sm  = ParagraphStyle('sm',  fontSize=7,  textColor=MUTED,     leading=10, fontName='Helvetica')
    lbl = ParagraphStyle('lbl', fontSize=7,  textColor=MUTED,     leading=9,  fontName='Helvetica-Bold')
    ctr = ParagraphStyle('ctr', fontSize=8,  textColor=DARK_TEXT, leading=11, fontName='Helvetica', alignment=TA_CENTER)

    story = []

    # ── Header banner ───────────────────────────────────────────────────────────
    header_data = [[
        Paragraph('AEF HRM', h1),
        Paragraph('Interview Shortlist Report', h2),
        Paragraph(f'Generated {timezone.now().strftime("%d %b %Y")}', ParagraphStyle('rt', fontSize=8, textColor=WHITE, fontName='Helvetica', alignment=TA_RIGHT)),
    ]]
    header_tbl = Table(header_data, colWidths=[4*cm, 9*cm, 4*cm])
    header_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), PRIMARY),
        ('VALIGN',     (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 14),
        ('BOTTOMPADDING', (0,0), (-1,-1), 14),
        ('LEFTPADDING',  (0,0), (0,-1), 12),
        ('RIGHTPADDING', (-1,0), (-1,-1), 12),
    ]))
    story.append(header_tbl)
    story.append(Spacer(1, 0.4*cm))

    # ── Job info bar ────────────────────────────────────────────────────────────
    job_data = [[
        Paragraph(f'<b>Position:</b> {posting.title}', bod),
        Paragraph(f'<b>Status:</b> {posting.get_status_display()}', bod),
        Paragraph(f'<b>Total Applications:</b> {posting.applications.count()}', bod),
        Paragraph(f'<b>Shortlisted:</b> {len(apps)}', bod),
    ]]
    job_tbl = Table(job_data, colWidths=[5*cm, 3.5*cm, 4*cm, 4.5*cm])
    job_tbl.setStyle(TableStyle([
        ('BACKGROUND',    (0,0), (-1,-1), LIGHT_BG),
        ('TOPPADDING',    (0,0), (-1,-1), 8),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ('LEFTPADDING',   (0,0), (-1,-1), 10),
        ('BOX',           (0,0), (-1,-1), 0.5, colors.HexColor('#dee2e6')),
        ('LINEAFTER',     (0,0), (2,-1),  0.5, colors.HexColor('#dee2e6')),
    ]))
    story.append(job_tbl)
    story.append(Spacer(1, 0.5*cm))

    if not apps:
        story.append(Paragraph('No candidates matched the selected filter.', bod))
    else:
        # ── Summary table ────────────────────────────────────────────────────────
        story.append(Paragraph('Candidate Rankings', h3))
        story.append(Spacer(1, 0.2*cm))

        tbl_head = ['#', 'Candidate', 'Email', 'AI Score', 'Verdict', 'Status']
        tbl_rows = [tbl_head]
        for i, app in enumerate(apps, 1):
            verdict = (app.ai_recommendation or '—').title()
            tbl_rows.append([
                str(i),
                app.applicant_name,
                app.applicant_email,
                f'{int(app.ai_score)}/100' if app.ai_score is not None else '—',
                verdict,
                app.get_status_display(),
            ])
        summary_tbl = Table(tbl_rows, colWidths=[0.8*cm, 4.5*cm, 4.5*cm, 1.8*cm, 2*cm, 2.4*cm])
        ts = TableStyle([
            ('BACKGROUND',    (0,0), (-1,0),  PRIMARY),
            ('TEXTCOLOR',     (0,0), (-1,0),  WHITE),
            ('FONTNAME',      (0,0), (-1,0),  'Helvetica-Bold'),
            ('FONTSIZE',      (0,0), (-1,-1), 7.5),
            ('TOPPADDING',    (0,0), (-1,-1), 5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('LEFTPADDING',   (0,0), (-1,-1), 6),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [WHITE, LIGHT_BG]),
            ('GRID',          (0,0), (-1,-1), 0.3, colors.HexColor('#dee2e6')),
            ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
            ('ALIGN',         (3,0), (3,-1),  'CENTER'),
            ('ALIGN',         (4,0), (4,-1),  'CENTER'),
        ])
        # Colour AI score cells
        for i, app in enumerate(apps, 1):
            score = app.ai_score or 0
            if score >= 70:
                ts.add('TEXTCOLOR', (3,i), (3,i), colors.HexColor('#065f46'))
                ts.add('BACKGROUND', (3,i), (3,i), INVITE_BG)
            elif score >= 50:
                ts.add('TEXTCOLOR', (3,i), (3,i), colors.HexColor('#92400e'))
                ts.add('BACKGROUND', (3,i), (3,i), HOLD_BG)
            else:
                ts.add('TEXTCOLOR', (3,i), (3,i), colors.HexColor('#991b1b'))
                ts.add('BACKGROUND', (3,i), (3,i), REJECT_BG)
        summary_tbl.setStyle(ts)
        story.append(summary_tbl)
        story.append(Spacer(1, 0.6*cm))

        # ── Per-candidate detail cards ───────────────────────────────────────────
        story.append(HRFlowable(width='100%', thickness=1, color=ACCENT, spaceAfter=8))
        story.append(Paragraph('Candidate Profiles', h3))
        story.append(Spacer(1, 0.3*cm))

        REC_COLORS = {
            'invite': (colors.HexColor('#065f46'), INVITE_BG),
            'hold':   (colors.HexColor('#92400e'), HOLD_BG),
            'reject': (colors.HexColor('#991b1b'), REJECT_BG),
        }

        for rank, app in enumerate(apps, 1):
            rec = (app.ai_recommendation or '').lower()
            txt_c, bg_c = REC_COLORS.get(rec, (DARK_TEXT, LIGHT_BG))
            score_str = f'{int(app.ai_score)}/100' if app.ai_score is not None else 'N/A'

            # Card header row
            card_head = Table([[
                Paragraph(f'<b>#{rank}  {app.applicant_name}</b>', ParagraphStyle('ch', fontSize=10, textColor=WHITE, fontName='Helvetica-Bold')),
                Paragraph(f'AI Score: <b>{score_str}</b>', ParagraphStyle('cs', fontSize=9, textColor=WHITE, fontName='Helvetica', alignment=TA_RIGHT)),
            ]], colWidths=[11*cm, 6*cm])
            card_head.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,-1), PRIMARY),
                ('TOPPADDING', (0,0), (-1,-1), 7), ('BOTTOMPADDING', (0,0), (-1,-1), 7),
                ('LEFTPADDING', (0,0), (0,-1), 10), ('RIGHTPADDING', (-1,0), (-1,-1), 10),
            ]))

            # Contact + verdict row
            contact_verdict = Table([[
                Paragraph(f'{app.applicant_email}  ·  Applied {app.submitted_at.strftime("%d %b %Y")}', sm),
                Paragraph(f'<b>{rec.upper() or "—"}</b>', ParagraphStyle('vt', fontSize=8, textColor=txt_c, fontName='Helvetica-Bold', alignment=TA_RIGHT)),
            ]], colWidths=[11*cm, 6*cm])
            contact_verdict.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,-1), bg_c),
                ('TOPPADDING', (0,0), (-1,-1), 5), ('BOTTOMPADDING', (0,0), (-1,-1), 5),
                ('LEFTPADDING', (0,0), (-1,-1), 10), ('RIGHTPADDING', (-1,0), (-1,-1), 10),
            ]))

            # Analysis body
            def ai_row(label, text):
                return [Paragraph(label, lbl), Paragraph(text or '—', bod)]

            analysis_rows = []
            if app.ai_summary:
                analysis_rows.append(ai_row('SUMMARY', app.ai_summary))
            if app.ai_strengths:
                analysis_rows.append(ai_row('STRENGTHS', app.ai_strengths))
            if app.ai_gaps:
                analysis_rows.append(ai_row('GAPS / CONCERNS', app.ai_gaps))

            if analysis_rows:
                body_tbl = Table(analysis_rows, colWidths=[3*cm, 14*cm])
                body_tbl.setStyle(TableStyle([
                    ('BACKGROUND',    (0,0), (-1,-1), WHITE),
                    ('TOPPADDING',    (0,0), (-1,-1), 5),
                    ('BOTTOMPADDING', (0,0), (-1,-1), 5),
                    ('LEFTPADDING',   (0,0), (-1,-1), 10),
                    ('LINEBELOW',     (0,0), (-1,-2), 0.3, colors.HexColor('#e5e7eb')),
                    ('BOX',           (0,0), (-1,-1), 0.3, colors.HexColor('#dee2e6')),
                    ('VALIGN',        (0,0), (-1,-1), 'TOP'),
                ]))
                card_elements = [card_head, contact_verdict, body_tbl]
            else:
                no_ai = Table([[Paragraph('No AI analysis yet. Click "Analyse with AI" on the applicant list.', sm)]],
                              colWidths=[17*cm])
                no_ai.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,-1), WHITE),
                    ('TOPPADDING', (0,0), (-1,-1), 8), ('BOTTOMPADDING', (0,0), (-1,-1), 8),
                    ('LEFTPADDING', (0,0), (-1,-1), 10),
                    ('BOX', (0,0), (-1,-1), 0.3, colors.HexColor('#dee2e6')),
                ]))
                card_elements = [card_head, contact_verdict, no_ai]

            story.append(KeepTogether(card_elements))
            story.append(Spacer(1, 0.4*cm))

    # ── Footer note ─────────────────────────────────────────────────────────────
    story.append(HRFlowable(width='100%', thickness=0.5, color=colors.HexColor('#dee2e6'), spaceBefore=8))
    story.append(Paragraph(
        f'Confidential — AEF HRM · Generated by {request.user.get_full_name() or request.user.username} · {timezone.now().strftime("%d %b %Y %H:%M")}',
        ParagraphStyle('ft', fontSize=7, textColor=MUTED, fontName='Helvetica', alignment=TA_CENTER, spaceBefore=4),
    ))

    doc.build(story)
    buf.seek(0)
    filename = f'shortlist_{posting.pk}_{timezone.now().strftime("%Y%m%d")}.pdf'
    return HttpResponse(buf, content_type='application/pdf',
                        headers={'Content-Disposition': f'attachment; filename="{filename}"'})
