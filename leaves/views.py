import json
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from django.db.models import Q
from django.urls import reverse
from datetime import date
from .models import LeaveRequest, LeaveBalance, LeaveType, LeaveConsultation
from .forms import LeaveRequestForm, ApprovalForm
from .seniority import seniority_entitlement
from accounts.models import Employee
from notifications.utils import notify


def get_employee(request):
    try:
        return request.user.employee
    except Employee.DoesNotExist:
        return None


def _save_drawn_signature(employee, b64_data):
    """Decode a base64 PNG from the signature pad and save to employee.signature + signature_b64."""
    import base64
    from django.core.files.base import ContentFile
    if not b64_data or not b64_data.startswith('data:image/png;base64,'):
        return
    # Always save b64 to DB first — this works even if the filesystem write fails
    employee.signature_b64 = b64_data
    save_fields = ['signature_b64']
    try:
        raw = base64.b64decode(b64_data.split(',', 1)[1])
        if employee.signature:
            try:
                employee.signature.delete(save=False)
            except Exception:
                pass
        fname = f"{employee.employee_id}_sig.png"
        employee.signature.save(fname, ContentFile(raw), save=False)
        save_fields.append('signature')
    except Exception:
        pass
    employee.save(update_fields=save_fields)


@login_required
def submit_leave(request):
    employee = get_employee(request)
    if not employee:
        messages.error(request, "Employee profile not found.")
        return redirect('dashboard:home')

    # Block suspended employees from applying for leave
    from datetime import date as _date
    from discipline.models import DisciplineRecord
    today = _date.today()
    active_suspension = DisciplineRecord.objects.filter(
        employee=employee,
        action_type='suspension',
        suspension_start__lte=today,
        suspension_end__gte=today,
    ).first()
    if active_suspension:
        messages.error(request, f"You cannot apply for leave while suspended (until {active_suspension.suspension_end.strftime('%d %B %Y')}).")
        return redirect('dashboard:home')

    # Get or create balance
    balance, _ = LeaveBalance.objects.get_or_create(
        employee=employee, year=date.today().year,
        defaults={'total_entitlement': seniority_entitlement(employee)}
    )

    # Build backup employee list: ALL active staff (not just same dept), excluding self
    today_date = date.today()
    all_employees = Employee.objects.filter(
        is_active=True,
    ).exclude(pk=employee.pk).select_related('user', 'department')

    on_leave_pks = set(
        LeaveRequest.objects.filter(
            status='approved',
            employee__in=all_employees,
            start_date__lte=today_date,
            end_date__gte=today_date,
        ).values_list('employee_id', flat=True)
    )

    backup_choices = [
        {'id': e.pk, 'name': e.get_full_name(), 'dept': str(e.department or ''), 'unavailable': e.pk in on_leave_pks}
        for e in all_employees.order_by('user__last_name', 'user__first_name')
    ]

    form = LeaveRequestForm(request.POST or None, request.FILES or None)
    if request.method == 'POST' and form.is_valid():
        leave = form.save(commit=False)
        leave.employee = employee

        # Backup employee — required
        backup_id = request.POST.get('backup_employee')
        if not backup_id:
            messages.error(request, "Please select a back-up employee to cover you during your leave.")
            return render(request, 'leaves/request_form.html', {
                'form': form, 'balance': balance, 'backup_choices_json': json.dumps(backup_choices),
                'employee': employee, 'current_sig_b64': employee.signature_b64 or '',
            })
        try:
            leave.backup_employee = Employee.objects.get(pk=backup_id)
        except Employee.DoesNotExist:
            messages.error(request, "Selected back-up employee not found.")
            return render(request, 'leaves/request_form.html', {
                'form': form, 'balance': balance, 'backup_choices_json': json.dumps(backup_choices),
                'employee': employee, 'current_sig_b64': employee.signature_b64 or '',
            })

        # Validate supporting document if provided
        supporting_doc = request.FILES.get('supporting_document')
        if supporting_doc:
            from leave_system.file_utils import validate_upload
            ok, err = validate_upload(supporting_doc)
            if not ok:
                messages.error(request, err)
                return render(request, 'leaves/request_form.html', {
                    'form': form, 'balance': balance, 'backup_choices_json': json.dumps(backup_choices),
                    'employee': employee, 'current_sig_b64': employee.signature_b64 or '',
                })

        # Validate balance (only for deductible leave types)
        if leave.leave_type.is_deductible and leave.total_days > balance.remaining_days:
            messages.error(request, f"Insufficient leave balance. You have {balance.remaining_days} days remaining.")
        else:
            # Check for overlapping requests
            overlapping = LeaveRequest.objects.filter(
                employee=employee,
                status__in=['pending', 'unit_head_approved', 'manager_approved', 'hr_approved', 'approved'],
                start_date__lte=leave.end_date,
                end_date__gte=leave.start_date,
            )
            if overlapping.exists():
                messages.error(request, "You already have a leave request for overlapping dates.")
            else:
                sig_b64 = request.POST.get('signature_data', '')
                _save_drawn_signature(employee, sig_b64)
                if sig_b64 and sig_b64.startswith('data:image/'):
                    leave.employee_sig_b64 = sig_b64
                elif employee.signature_b64:
                    leave.employee_sig_b64 = employee.signature_b64
                leave.save()
                if employee.is_intern() or employee.is_wacs_resident():
                    # Interns and WACS residents skip manager — go directly to HR
                    role_label = 'Intern' if employee.is_intern() else 'WACS Resident'
                    leave.status = LeaveRequest.STATUS_MANAGER_APPROVED
                    leave.save(update_fields=['status'])
                    messages.success(request, f"Leave request submitted for {leave.total_days} day(s). Sent directly to HR for review.")
                    for hr_emp in Employee.objects.filter(role='hr', is_active=True).select_related('user'):
                        notify(
                            hr_emp.user,
                            f'{role_label} Leave Request — {employee.get_full_name()}',
                            f'{employee.get_full_name()} ({role_label}) has submitted a {leave.leave_type} request '
                            f'for {leave.total_days} day(s) ({leave.start_date} → {leave.end_date}). '
                            f'No manager approval required — awaiting your HR review.',
                            notification_type='leave_submitted',
                            url=reverse('leaves:detail', kwargs={'pk': leave.pk}),
                        )
                elif employee.role in ('admin_director', 'medical_director') or employee.reports_to_ceo:
                    # Admin Director, Medical Director, or any staff flagged reports_to_ceo — CEO sole approver
                    role_labels = {'admin_director': 'Administration Director', 'medical_director': 'Medical Director'}
                    role_label = role_labels.get(employee.role, 'Staff')
                    leave.status = LeaveRequest.STATUS_HR_APPROVED
                    leave.save(update_fields=['status'])
                    messages.success(request, f"Leave request submitted for {leave.total_days} day(s). Sent directly to the CEO for approval.")
                    for ceo_emp in Employee.objects.filter(role='ceo', is_active=True).select_related('user'):
                        notify(
                            ceo_emp.user,
                            f'{role_label} Leave — Awaiting Your Approval',
                            f'{employee.get_full_name()} ({role_label}) has submitted a {leave.leave_type} request '
                            f'for {leave.total_days} day(s) ({leave.start_date} → {leave.end_date}). Awaiting your approval.',
                            notification_type='leave_submitted',
                            url=reverse('leaves:ceo_action', kwargs={'pk': leave.pk}),
                        )
                elif employee.is_hr():
                    # HR staff — skip all intermediate steps; Admin Director approves directly
                    leave.status = LeaveRequest.STATUS_HR_APPROVED
                    leave.save(update_fields=['status'])
                    messages.success(request, f"Leave request submitted for {leave.total_days} day(s). Sent directly to Administration Director for approval.")
                    for dir_emp in Employee.objects.filter(role__in=('admin_director', 'finance_director'), is_active=True).select_related('user'):
                        notify(
                            dir_emp.user,
                            f'HR Staff Leave — Awaiting Your Approval — {employee.get_full_name()}',
                            f'{employee.get_full_name()} (HR) has submitted a {leave.leave_type} request '
                            f'for {leave.total_days} day(s) ({leave.start_date} → {leave.end_date}). '
                            f'No intermediate approvals required — awaiting your decision.',
                            notification_type='leave_submitted',
                            url=reverse('leaves:director_action', kwargs={'pk': leave.pk}),
                        )
                elif employee.reports_to_hr:
                    # Reports directly to HR — HR is the FINAL approver, no Director step
                    leave.status = LeaveRequest.STATUS_MANAGER_APPROVED
                    leave.save(update_fields=['status'])
                    messages.success(request, f"Leave request submitted for {leave.total_days} day(s). Sent to HR for final approval.")
                    for hr_emp in Employee.objects.filter(role='hr', is_active=True).select_related('user'):
                        notify(
                            hr_emp.user,
                            f'Leave Request (HR Final) — {employee.get_full_name()}',
                            f'{employee.get_full_name()} has submitted a {leave.leave_type} request '
                            f'for {leave.total_days} day(s) ({leave.start_date} → {leave.end_date}). '
                            f'You are the final approver — no Director step required.',
                            notification_type='leave_submitted',
                            url=reverse('leaves:hr_action', kwargs={'pk': leave.pk}),
                        )
                elif employee.reports_to_director:
                    # Reports directly to Director — skip unit head and manager; goes to HR then Director
                    leave.status = LeaveRequest.STATUS_MANAGER_APPROVED
                    leave.save(update_fields=['status'])
                    messages.success(request, f"Leave request submitted for {leave.total_days} day(s). Sent directly to HR for review.")
                    for hr_emp in Employee.objects.filter(role='hr', is_active=True).select_related('user'):
                        notify(
                            hr_emp.user,
                            f'Leave Request (Direct Report) — {employee.get_full_name()}',
                            f'{employee.get_full_name()} has submitted a {leave.leave_type} request '
                            f'for {leave.total_days} day(s) ({leave.start_date} → {leave.end_date}). '
                            f'No unit head or manager approval required — awaiting your HR review.',
                            notification_type='leave_submitted',
                            url=reverse('leaves:hr_action', kwargs={'pk': leave.pk}),
                        )
                elif employee.unit_head:
                    # Has unit head — notify unit head first
                    messages.success(request, f"Leave request submitted for {leave.total_days} day(s). Awaiting Unit Head approval.")
                    notify(
                        employee.unit_head.user,
                        f'New Leave Request — {employee.get_full_name()}',
                        f'{employee.get_full_name()} has submitted a {leave.leave_type} request '
                        f'for {leave.total_days} day(s) ({leave.start_date} → {leave.end_date}). Awaiting your Unit Head approval.',
                        notification_type='leave_submitted',
                        url=reverse('leaves:unit_head_action', kwargs={'pk': leave.pk}),
                    )
                else:
                    # No unit head — notify supervisor (line manager) directly
                    messages.success(request, f"Leave request submitted successfully for {leave.total_days} day(s). Awaiting manager approval.")
                    if employee.supervisor:
                        notify(
                            employee.supervisor.user,
                            f'New Leave Request — {employee.get_full_name()}',
                            f'{employee.get_full_name()} has submitted a {leave.leave_type} request '
                            f'for {leave.total_days} day(s) ({leave.start_date} → {leave.end_date}). Awaiting your approval.',
                            notification_type='leave_submitted',
                            url=reverse('leaves:manager_action', kwargs={'pk': leave.pk}),
                        )
                try:
                    from dashboard.models import AuditLog
                    AuditLog.log(
                        request, AuditLog.ACTION_LEAVE_SUBMIT,
                        f"Submitted {leave.leave_type} leave request for {leave.total_days} day(s) "
                        f"({leave.start_date} → {leave.end_date})",
                    )
                except Exception:
                    pass
                return redirect('leaves:my_requests')

    leave_types = LeaveType.objects.filter(is_active=True).values('id', 'is_deductible')
    deductible_map = json.dumps({str(lt['id']): lt['is_deductible'] for lt in leave_types})

    return render(request, 'leaves/request_form.html', {
        'form': form,
        'balance': balance,
        'deductible_map': deductible_map,
        'backup_choices_json': json.dumps(backup_choices),
        'employee': employee,
        'current_sig_b64': employee.signature_b64 or '',
    })


@login_required
def my_requests(request):
    employee = get_employee(request)
    if not employee:
        return redirect('dashboard:home')

    current_year = date.today().year
    try:
        year_filter = int(request.GET.get('year', current_year))
    except (ValueError, TypeError):
        year_filter = current_year
    status_filter = request.GET.get('status', '')

    qs = LeaveRequest.objects.filter(
        employee=employee,
        start_date__year=year_filter,
    ).select_related('leave_type', 'manager_action_by__user', 'hr_action_by__user')

    if status_filter:
        qs = qs.filter(status=status_filter)

    balance, _ = LeaveBalance.objects.get_or_create(
        employee=employee, year=current_year,
        defaults={'total_entitlement': seniority_entitlement(employee, current_year)}
    )

    return render(request, 'leaves/my_requests.html', {
        'leave_requests': qs,
        'balance': balance,
        'year_filter': year_filter,
        'status_filter': status_filter,
        'status_choices': LeaveRequest.STATUS_CHOICES,
        'years': range(2000, current_year + 11),
    })


@login_required
def cancel_request(request, pk):
    employee = get_employee(request)
    leave = get_object_or_404(LeaveRequest, pk=pk, employee=employee)
    if leave.can_cancel():
        leave.status = LeaveRequest.STATUS_CANCELLED
        leave.save()
        messages.success(request, "Leave request cancelled.")

        # Notify unit head and/or line manager so they know not to act on it
        notif_title = f'Leave Request Cancelled — {employee.get_full_name()}'
        notif_msg = (
            f'{employee.get_full_name()} has cancelled their {leave.leave_type} leave request '
            f'({leave.start_date} → {leave.end_date}). No further action is needed.'
        )
        detail_url = reverse('leaves:detail', kwargs={'pk': leave.pk})
        recipients = set()
        if leave.employee.unit_head:
            recipients.add(leave.employee.unit_head.user)
        if leave.employee.supervisor:
            recipients.add(leave.employee.supervisor.user)
        for recipient in recipients:
            notify(recipient, notif_title, notif_msg,
                   notification_type='leave_cancelled', url=detail_url)
    else:
        messages.error(request, "This request cannot be cancelled.")
    return redirect('leaves:my_requests')


@login_required
def unit_head_approvals(request):
    employee = get_employee(request)
    if not employee or not employee.is_unit_head():
        messages.error(request, "Access denied. Unit Head role required.")
        return redirect('dashboard:home')

    pending = LeaveRequest.objects.filter(
        status=LeaveRequest.STATUS_PENDING,
        employee__unit_head=employee,
    ).select_related('employee__user', 'employee__department', 'leave_type')

    return render(request, 'leaves/unit_head_approvals.html', {
        'pending_requests': pending,
    })


@login_required
def unit_head_action(request, pk):
    employee = get_employee(request)
    if not employee or not employee.is_unit_head():
        messages.error(request, "Access denied. Unit Head role required.")
        return redirect('dashboard:home')

    leave = get_object_or_404(LeaveRequest, pk=pk)

    if leave.employee.unit_head != employee and not request.user.is_superuser:
        messages.error(request, "You are not the Unit Head for this employee.")
        return redirect('leaves:unit_head_approvals')

    if leave.status != LeaveRequest.STATUS_PENDING:
        messages.warning(request, f"Leave request #{pk} is no longer pending (status: {leave.get_status_display()}).")
        return redirect('leaves:unit_head_approvals')

    if request.method == 'POST':
        form = ApprovalForm(request.POST)
        if form.is_valid():
            action = form.cleaned_data['action']
            remarks = form.cleaned_data['remarks']
            leave.unit_head_action_by = employee
            leave.unit_head_action_date = timezone.now()
            leave.unit_head_remarks = remarks
            if action == 'approve':
                sig_b64 = request.POST.get('signature_data', '')
                _save_drawn_signature(employee, sig_b64)
                leave.unit_head_sig_b64 = sig_b64 if sig_b64.startswith('data:image/') else (employee.signature_b64 or '')
                leave.status = LeaveRequest.STATUS_UNIT_HEAD_APPROVED
                leave.save()
                messages.success(request, "Leave approved. Forwarded to Line Manager.")
                notify(
                    leave.employee.user,
                    'Leave Request — Unit Head Approved',
                    f'Your {leave.leave_type} request ({leave.start_date} → {leave.end_date}) '
                    f'has been approved by your Unit Head and is now awaiting your Line Manager.',
                    notification_type='leave_manager_approved',
                    url=reverse('leaves:detail', kwargs={'pk': leave.pk}),
                )
                if leave.employee.supervisor:
                    notify(
                        leave.employee.supervisor.user,
                        f'Leave Awaiting Your Approval — {leave.employee.get_full_name()}',
                        f'{leave.employee.get_full_name()} ({leave.employee.department or "No dept"}) '
                        f'has a {leave.leave_type} request ({leave.start_date} → {leave.end_date}, '
                        f'{leave.total_days} day(s)) approved by Unit Head, pending your review.',
                        notification_type='leave_submitted',
                        url=reverse('leaves:manager_action', kwargs={'pk': leave.pk}),
                    )
            else:
                leave.status = LeaveRequest.STATUS_REJECTED_UNIT_HEAD
                leave.save()
                messages.warning(request, f"Leave request #{pk} rejected.")
                notify(
                    leave.employee.user,
                    'Leave Request — Rejected by Unit Head',
                    f'Your {leave.leave_type} request ({leave.start_date} → {leave.end_date}) '
                    f'was rejected by your Unit Head.'
                    + (f' Remarks: {remarks}' if remarks else ''),
                    notification_type='leave_rejected',
                    url=reverse('leaves:detail', kwargs={'pk': leave.pk}),
                )
            return redirect('leaves:unit_head_approvals')
    else:
        form = ApprovalForm()

    return render(request, 'leaves/action_form.html', {
        'leave': leave,
        'form': form,
        'action_title': 'Unit Head Review',
        'action_type': 'unit_head',
        'current_sig_b64': employee.signature_b64 or '',
    })


@login_required
def manager_approvals(request):
    employee = get_employee(request)
    if not employee or not employee.is_manager():
        messages.error(request, "Access denied.")
        return redirect('dashboard:home')

    if request.user.is_superuser:
        pending = LeaveRequest.objects.filter(
            status__in=[LeaveRequest.STATUS_PENDING, LeaveRequest.STATUS_UNIT_HEAD_APPROVED]
        ).select_related('employee__user', 'employee__department', 'leave_type')
    else:
        # Show 'pending' leaves from employees WITHOUT a unit_head (direct reports)
        # and 'unit_head_approved' from employees WITH a unit_head (both supervised by this manager)
        pending = LeaveRequest.objects.filter(
            employee__supervisor=employee
        ).filter(
            Q(status=LeaveRequest.STATUS_PENDING, employee__unit_head__isnull=True) |
            Q(status=LeaveRequest.STATUS_UNIT_HEAD_APPROVED)
        ).select_related('employee__user', 'employee__department', 'leave_type')

    return render(request, 'leaves/manager_approvals.html', {
        'pending_requests': pending,
    })


@login_required
def manager_action(request, pk):
    employee = get_employee(request)
    if not employee or not employee.is_manager():
        messages.error(request, "Access denied. Manager role required.")
        return redirect('dashboard:home')

    leave = get_object_or_404(LeaveRequest, pk=pk)

    # Check supervisor authority (superuser bypasses this)
    if not request.user.is_superuser and leave.employee.supervisor != employee:
        messages.error(request, "You are not the supervisor for this employee.")
        return redirect('leaves:manager_approvals')

    # Accept both 'pending' (no unit_head employees) and 'unit_head_approved' (has unit_head)
    valid_statuses = [LeaveRequest.STATUS_PENDING, LeaveRequest.STATUS_UNIT_HEAD_APPROVED]
    if leave.status not in valid_statuses:
        messages.warning(
            request,
            f"Leave request #{pk} is no longer pending "
            f"(current status: {leave.get_status_display()}). No action taken."
        )
        return redirect('leaves:manager_approvals')

    if request.method == 'POST':
        form = ApprovalForm(request.POST)
        if form.is_valid():
            action = form.cleaned_data['action']
            remarks = form.cleaned_data['remarks']
            leave.manager_action_by = employee
            leave.manager_action_date = timezone.now()
            leave.manager_remarks = remarks
            if action == 'approve':
                sig_b64 = request.POST.get('signature_data', '')
                _save_drawn_signature(employee, sig_b64)
                leave.manager_sig_b64 = sig_b64 if sig_b64.startswith('data:image/') else (employee.signature_b64 or '')
                leave.status = LeaveRequest.STATUS_MANAGER_APPROVED
                from accounts.models import Employee as _Emp
                nurse_supt = leave.employee.nurse_superintendent if leave.employee.requires_nurse_supt else None
                if nurse_supt:
                    # Route to Nurse Superintendent next
                    messages.success(request, "Leave request approved. Forwarded to Nurse Superintendent for review.")
                    notify(
                        leave.employee.user,
                        'Leave Request — Manager Approved',
                        f'Your {leave.leave_type} request ({leave.start_date} → {leave.end_date}) '
                        f'has been approved by your manager and is now awaiting Nurse Superintendent review.',
                        notification_type='leave_manager_approved',
                        url=reverse('leaves:detail', kwargs={'pk': leave.pk}),
                    )
                    notify(
                        nurse_supt.user,
                        f'Leave Awaiting Your Review — {leave.employee.get_full_name()}',
                        f'{leave.employee.get_full_name()} has a {leave.leave_type} request '
                        f'({leave.start_date} → {leave.end_date}, {leave.total_days} day(s)) '
                        f'pending your review as Nurse Superintendent.',
                        notification_type='leave_submitted',
                        url=reverse('leaves:nurse_supt_action', kwargs={'pk': leave.pk}),
                    )
                else:
                    # Standard flow: go to HR
                    messages.success(request, "Leave request approved. Forwarded to HR for review.")
                    notify(
                        leave.employee.user,
                        'Leave Request — Manager Approved',
                        f'Your {leave.leave_type} request ({leave.start_date} → {leave.end_date}) '
                        f'has been approved by your manager and is now awaiting HR review.',
                        notification_type='leave_manager_approved',
                        url=reverse('leaves:detail', kwargs={'pk': leave.pk}),
                    )
                    for hr_emp in _Emp.objects.filter(role='hr', is_active=True).select_related('user'):
                        notify(
                            hr_emp.user,
                            f'Leave Awaiting HR Review — {leave.employee.get_full_name()}',
                            f'{leave.employee.get_full_name()} ({leave.employee.department or "No dept"}) '
                            f'has a {leave.leave_type} request ({leave.start_date} → {leave.end_date}, '
                            f'{leave.total_days} day(s)) pending your review.',
                            notification_type='leave_submitted',
                            url=reverse('leaves:hr_action', kwargs={'pk': leave.pk}),
                        )
            else:
                leave.status = LeaveRequest.STATUS_REJECTED_MANAGER
                messages.warning(request, f"Leave request #{pk} has been rejected.")
                notify(
                    leave.employee.user,
                    'Leave Request — Rejected by Manager',
                    f'Your {leave.leave_type} request ({leave.start_date} → {leave.end_date}) '
                    f'was rejected by your manager.'
                    + (f' Remarks: {remarks}' if remarks else ''),
                    notification_type='leave_rejected',
                    url=reverse('leaves:detail', kwargs={'pk': leave.pk}),
                )
            leave.save()
            try:
                from dashboard.models import AuditLog
                audit_action = AuditLog.ACTION_LEAVE_APPROVE if action == 'approve' else AuditLog.ACTION_LEAVE_REJECT
                AuditLog.log(
                    request, audit_action,
                    f"Manager {'approved' if action == 'approve' else 'rejected'} leave request "
                    f"for {leave.employee.get_full_name()} ({leave.leave_type}, {leave.start_date} → {leave.end_date})",
                    target_user=leave.employee.user,
                )
            except Exception:
                pass
            return redirect('leaves:manager_approvals')
    else:
        form = ApprovalForm()

    return render(request, 'leaves/action_form.html', {
        'leave': leave,
        'form': form,
        'action_title': 'Manager Review',
        'action_type': 'manager',
        'current_sig_b64': employee.signature_b64 or '',
    })


@login_required
def nurse_supt_approvals(request):
    """Nurse Superintendent: list of leaves awaiting their approval."""
    employee = get_employee(request)
    if not employee or not employee.is_nurse_superintendent():
        messages.error(request, "Access denied. Nurse Superintendent role required.")
        return redirect('dashboard:home')

    pending = LeaveRequest.objects.filter(
        status=LeaveRequest.STATUS_MANAGER_APPROVED,
        employee__requires_nurse_supt=True,
        employee__nurse_superintendent=employee,
    ).select_related('employee__user', 'employee__department', 'leave_type', 'manager_action_by__user')

    return render(request, 'leaves/nurse_supt_approvals.html', {
        'pending_requests': pending,
    })


@login_required
def nurse_supt_action(request, pk):
    """Nurse Superintendent approves or rejects a leave request."""
    employee = get_employee(request)
    if not employee or not employee.is_nurse_superintendent():
        messages.error(request, "Access denied. Nurse Superintendent role required.")
        return redirect('dashboard:home')

    leave = get_object_or_404(LeaveRequest, pk=pk)

    if leave.status != LeaveRequest.STATUS_MANAGER_APPROVED or not leave.employee.requires_nurse_supt:
        messages.warning(request, "This leave request is not awaiting Nurse Superintendent approval.")
        return redirect('leaves:nurse_supt_approvals')

    if not request.user.is_superuser and leave.employee.nurse_superintendent != employee:
        messages.error(request, "You are not the Nurse Superintendent for this employee.")
        return redirect('leaves:nurse_supt_approvals')

    if request.method == 'POST':
        form = ApprovalForm(request.POST)
        if form.is_valid():
            action = form.cleaned_data['action']
            remarks = form.cleaned_data['remarks']
            leave.nurse_supt_action_by = employee
            leave.nurse_supt_action_date = timezone.now()
            leave.nurse_supt_remarks = remarks
            if action == 'approve':
                sig_b64 = request.POST.get('signature_data', '')
                _save_drawn_signature(employee, sig_b64)
                leave.nurse_supt_sig_b64 = sig_b64 if sig_b64.startswith('data:image/') else (employee.signature_b64 or '')
                leave.status = LeaveRequest.STATUS_NURSE_SUPT_APPROVED
                messages.success(request, "Leave request approved. Forwarded to HR for review.")
                notify(
                    leave.employee.user,
                    'Leave Request — Nurse Superintendent Approved',
                    f'Your {leave.leave_type} request ({leave.start_date} → {leave.end_date}) '
                    f'has been approved by the Nurse Superintendent and is now awaiting HR review.',
                    notification_type='leave_manager_approved',
                    url=reverse('leaves:detail', kwargs={'pk': leave.pk}),
                )
                from accounts.models import Employee as _Emp
                for hr_emp in _Emp.objects.filter(role='hr', is_active=True).select_related('user'):
                    notify(
                        hr_emp.user,
                        f'Leave Awaiting HR Review — {leave.employee.get_full_name()}',
                        f'{leave.employee.get_full_name()} has a {leave.leave_type} request '
                        f'({leave.start_date} → {leave.end_date}, {leave.total_days} day(s)) '
                        f'pending your review. Approved by Nurse Superintendent.',
                        notification_type='leave_submitted',
                        url=reverse('leaves:hr_action', kwargs={'pk': leave.pk}),
                    )
            else:
                leave.status = LeaveRequest.STATUS_REJECTED_NURSE_SUPT
                messages.warning(request, f"Leave request #{pk} has been rejected.")
                notify(
                    leave.employee.user,
                    'Leave Request — Rejected by Nurse Superintendent',
                    f'Your {leave.leave_type} request ({leave.start_date} → {leave.end_date}) '
                    f'was rejected by the Nurse Superintendent.'
                    + (f' Remarks: {remarks}' if remarks else ''),
                    notification_type='leave_rejected',
                    url=reverse('leaves:detail', kwargs={'pk': leave.pk}),
                )
            leave.save()
            return redirect('leaves:nurse_supt_approvals')
    else:
        form = ApprovalForm()

    return render(request, 'leaves/action_form.html', {
        'leave': leave,
        'form': form,
        'action_title': 'Nurse Superintendent Review',
        'action_type': 'nurse_supt',
        'current_sig_b64': employee.signature_b64 or '',
    })


@login_required
def hr_approvals(request):
    employee = get_employee(request)
    if not employee or not employee.is_hr():
        messages.error(request, "Access denied. HR Admin only.")
        return redirect('dashboard:home')

    from django.db.models import Q
    pending = LeaveRequest.objects.filter(
        Q(status=LeaveRequest.STATUS_MANAGER_APPROVED, employee__requires_nurse_supt=False) |
        Q(status=LeaveRequest.STATUS_MANAGER_APPROVED, employee__nurse_superintendent__isnull=True) |
        Q(status=LeaveRequest.STATUS_NURSE_SUPT_APPROVED)
    ).distinct().select_related('employee__user', 'employee__department', 'leave_type', 'manager_action_by__user')

    return render(request, 'leaves/hr_approvals.html', {
        'pending_requests': pending,
    })


@login_required
def hr_action(request, pk):
    employee = get_employee(request)
    if not employee or not employee.is_hr():
        messages.error(request, "Access denied. HR Admin role required.")
        return redirect('dashboard:home')

    leave = get_object_or_404(LeaveRequest, pk=pk)

    valid_for_hr = [LeaveRequest.STATUS_MANAGER_APPROVED, LeaveRequest.STATUS_NURSE_SUPT_APPROVED]
    if leave.status not in valid_for_hr:
        messages.warning(
            request,
            f"Leave request #{pk} is not awaiting HR approval "
            f"(current status: {leave.get_status_display()}). No action taken."
        )
        return redirect('leaves:hr_approvals')

    if request.method == 'POST':
        form = ApprovalForm(request.POST)
        if form.is_valid():
            action = form.cleaned_data['action']
            remarks = form.cleaned_data['remarks']
            leave.hr_action_by = employee
            leave.hr_action_date = timezone.now()
            leave.hr_remarks = remarks
            if action == 'approve':
                sig_b64 = request.POST.get('signature_data', '')
                _save_drawn_signature(employee, sig_b64)
                hr_sig = sig_b64 if sig_b64.startswith('data:image/') else (employee.signature_b64 or '')
                leave.hr_sig_b64 = hr_sig
                if leave.employee.reports_to_hr:
                    # HR is the FINAL approver — approve directly, no Director step
                    now = timezone.now()
                    if not leave.unit_head_action_by:
                        leave.unit_head_action_by = employee
                        leave.unit_head_action_date = now
                        leave.unit_head_remarks = 'Direct report to HR — approved by HR'
                        leave.unit_head_sig_b64 = hr_sig
                    if not leave.manager_action_by:
                        leave.manager_action_by = employee
                        leave.manager_action_date = now
                        leave.manager_remarks = 'Direct report to HR — approved by HR'
                        leave.manager_sig_b64 = hr_sig
                    # Also fill director slot so PDF is complete
                    leave.director_action_by = employee
                    leave.director_action_date = now
                    leave.director_remarks = 'Final approval by HR (reports directly to HR)'
                    leave.director_sig_b64 = hr_sig
                    leave.status = LeaveRequest.STATUS_APPROVED
                    messages.success(request, "Leave request FULLY APPROVED by HR (final approver).")
                    notify(
                        leave.employee.user,
                        'Leave Request — Fully Approved by HR',
                        f'Your {leave.leave_type} request ({leave.start_date} → {leave.end_date}, '
                        f'{leave.total_days} day(s)) has been fully approved by HR.',
                        notification_type='leave_approved',
                        url=reverse('leaves:detail', kwargs={'pk': leave.pk}),
                    )
                else:
                    leave.status = LeaveRequest.STATUS_HR_APPROVED
                    messages.success(request, "Leave request approved by HR. Forwarded to Administration Director.")
                    notify(
                        leave.employee.user,
                        'Leave Request — HR Approved',
                        f'Your {leave.leave_type} request ({leave.start_date} → {leave.end_date}) '
                        f'has been approved by HR and is now awaiting the Administration Director\'s final decision.',
                        notification_type='leave_hr_approved',
                        url=reverse('leaves:detail', kwargs={'pk': leave.pk}),
                    )
                    # Notify Admin Director and Finance Director
                    from accounts.models import Employee as _Emp
                    for dir_emp in _Emp.objects.filter(role__in=('admin_director', 'finance_director'), is_active=True).select_related('user'):
                        notify(
                            dir_emp.user,
                            f'Leave Awaiting Your Approval — {leave.employee.get_full_name()}',
                            f'{leave.employee.get_full_name()} ({leave.employee.department or "No dept"}) '
                            f'has a {leave.leave_type} request ({leave.start_date} → {leave.end_date}, '
                            f'{leave.total_days} day(s)) awaiting your final decision.',
                            notification_type='leave_hr_approved',
                            url=reverse('leaves:director_action', kwargs={'pk': leave.pk}),
                        )
            else:
                leave.status = LeaveRequest.STATUS_REJECTED_HR
                messages.warning(request, f"Leave request #{pk} has been rejected by HR.")
                notify(
                    leave.employee.user,
                    'Leave Request — Rejected by HR',
                    f'Your {leave.leave_type} request ({leave.start_date} → {leave.end_date}) '
                    f'was rejected by HR.'
                    + (f' Remarks: {remarks}' if remarks else ''),
                    notification_type='leave_rejected',
                    url=reverse('leaves:detail', kwargs={'pk': leave.pk}),
                )
            leave.save()
            try:
                from dashboard.models import AuditLog
                audit_action = AuditLog.ACTION_LEAVE_APPROVE if action == 'approve' else AuditLog.ACTION_LEAVE_REJECT
                AuditLog.log(
                    request, audit_action,
                    f"HR {'approved' if action == 'approve' else 'rejected'} leave request "
                    f"for {leave.employee.get_full_name()} ({leave.leave_type}, {leave.start_date} → {leave.end_date})",
                    target_user=leave.employee.user,
                )
            except Exception:
                pass
            return redirect('leaves:hr_approvals')
    else:
        form = ApprovalForm()

    return render(request, 'leaves/action_form.html', {
        'leave': leave,
        'form': form,
        'action_title': 'HR Final Review',
        'action_type': 'hr',
        'current_sig_b64': employee.signature_b64 or '',
    })


@login_required
def director_approvals(request):
    employee = get_employee(request)
    if not employee or not employee.is_director():
        messages.error(request, "Access denied. Administration Director only.")
        return redirect('dashboard:home')

    pending = LeaveRequest.objects.filter(
        status=LeaveRequest.STATUS_HR_APPROVED
    ).exclude(
        # Admin Director, Medical Director, and reports_to_ceo go to CEO instead
        Q(employee__role__in=('admin_director', 'medical_director')) | Q(employee__reports_to_ceo=True)
    ).select_related('employee__user', 'employee__department', 'leave_type', 'manager_action_by__user', 'hr_action_by__user')

    # Check if admin director is on leave (finance director may be covering)
    from datetime import date as _date
    today = _date.today()
    admin_directors = Employee.objects.filter(role='admin_director', is_active=True)
    admin_dir_on_leave = any(
        req for ad in admin_directors
        for req in ad.leave_requests.filter(status='approved', start_date__lte=today, end_date__gte=today)
    )

    return render(request, 'leaves/director_approvals.html', {
        'pending_requests': pending,
        'employee': employee,
        'admin_dir_on_leave': admin_dir_on_leave,
    })


@login_required
def director_action(request, pk):
    employee = get_employee(request)
    if not employee or not employee.is_director():
        messages.error(request, "Access denied.")
        return redirect('dashboard:home')

    leave = get_object_or_404(LeaveRequest, pk=pk)

    if leave.status != LeaveRequest.STATUS_HR_APPROVED:
        messages.warning(
            request,
            f"Leave request #{pk} is not awaiting Director approval "
            f"(current status: {leave.get_status_display()}). No action taken."
        )
        return redirect('leaves:director_approvals')

    if request.method == 'POST':
        form = ApprovalForm(request.POST)
        if form.is_valid():
            action = form.cleaned_data['action']
            remarks = form.cleaned_data['remarks']
            leave.director_action_by = employee
            leave.director_action_date = timezone.now()
            leave.director_remarks = remarks
            if action == 'approve':
                sig_b64 = request.POST.get('signature_data', '')
                _save_drawn_signature(employee, sig_b64)
                dir_sig = sig_b64 if sig_b64.startswith('data:image/') else (employee.signature_b64 or '')
                leave.director_sig_b64 = dir_sig
                leave.status = LeaveRequest.STATUS_APPROVED
                # Auto-fill bypassed approval slots for direct reports and HR staff
                applicant = leave.employee
                now = timezone.now()
                if applicant.reports_to_director or applicant.is_hr():
                    if not leave.unit_head_action_by:
                        leave.unit_head_action_by = employee
                        leave.unit_head_action_date = now
                        leave.unit_head_remarks = 'Direct report — approved by Administration Director'
                        leave.unit_head_sig_b64 = dir_sig
                    if not leave.manager_action_by:
                        leave.manager_action_by = employee
                        leave.manager_action_date = now
                        leave.manager_remarks = 'Direct report — approved by Administration Director'
                        leave.manager_sig_b64 = dir_sig
                    if applicant.is_hr() and not leave.hr_action_by:
                        leave.hr_action_by = employee
                        leave.hr_action_date = now
                        leave.hr_remarks = 'HR applicant — approved by Administration Director'
                        leave.hr_sig_b64 = dir_sig
                messages.success(request, "Leave request FULLY APPROVED by Administration Director.")
                notify(
                    leave.employee.user,
                    'Leave Request — Fully Approved',
                    f'Great news! Your {leave.leave_type} request ({leave.start_date} → {leave.end_date}, '
                    f'{leave.total_days} day(s)) has been fully approved by the Administration Director.',
                    notification_type='leave_approved',
                    url=reverse('leaves:detail', kwargs={'pk': leave.pk}),
                )
            else:
                leave.status = LeaveRequest.STATUS_REJECTED_DIRECTOR
                messages.warning(request, f"Leave request #{pk} has been rejected by Administration Director.")
                notify(
                    leave.employee.user,
                    'Leave Request — Rejected by Director',
                    f'Your {leave.leave_type} request ({leave.start_date} → {leave.end_date}) '
                    f'was rejected by the Administration Director.'
                    + (f' Remarks: {remarks}' if remarks else ''),
                    notification_type='leave_rejected',
                    url=reverse('leaves:detail', kwargs={'pk': leave.pk}),
                )
            leave.save()
            try:
                from dashboard.models import AuditLog
                audit_action = AuditLog.ACTION_LEAVE_APPROVE if action == 'approve' else AuditLog.ACTION_LEAVE_REJECT
                AuditLog.log(
                    request, audit_action,
                    f"Director {'approved' if action == 'approve' else 'rejected'} leave request "
                    f"for {leave.employee.get_full_name()} ({leave.leave_type}, {leave.start_date} → {leave.end_date})",
                    target_user=leave.employee.user,
                )
            except Exception:
                pass
            return redirect('leaves:director_approvals')
    else:
        form = ApprovalForm()

    return render(request, 'leaves/action_form.html', {
        'leave': leave,
        'form': form,
        'action_title': 'Administration Director — Final Review',
        'action_type': 'director',
        'current_sig_b64': employee.signature_b64 or '',
    })


@login_required
def ceo_approvals(request):
    """CEO sees and approves Admin Director leave requests."""
    employee = get_employee(request)
    if not employee or not employee.is_ceo():
        messages.error(request, "Access denied. CEO only.")
        return redirect('dashboard:home')

    pending = LeaveRequest.objects.filter(
        status=LeaveRequest.STATUS_HR_APPROVED,
    ).filter(
        Q(employee__role__in=('admin_director', 'medical_director')) | Q(employee__reports_to_ceo=True)
    ).select_related('employee__user', 'employee__department', 'leave_type')

    return render(request, 'leaves/ceo_approvals.html', {
        'pending_requests': pending,
        'employee': employee,
    })


@login_required
def ceo_action(request, pk):
    """CEO approves or rejects Admin Director leave."""
    employee = get_employee(request)
    if not employee or not employee.is_ceo():
        messages.error(request, "Access denied. CEO only.")
        return redirect('dashboard:home')

    leave = get_object_or_404(LeaveRequest, pk=pk)

    applicant = leave.employee
    is_ceo_case = applicant.role in ('admin_director', 'medical_director') or applicant.reports_to_ceo
    if leave.status != LeaveRequest.STATUS_HR_APPROVED or not is_ceo_case:
        messages.warning(request, f"Leave request #{pk} is not awaiting CEO approval.")
        return redirect('leaves:ceo_approvals')

    if request.method == 'POST':
        form = ApprovalForm(request.POST)
        if form.is_valid():
            action = form.cleaned_data['action']
            remarks = form.cleaned_data['remarks']
            # Record CEO approval in the director slot (CEO is acting as final approver)
            leave.director_action_by = employee
            leave.director_action_date = timezone.now()
            leave.director_remarks = remarks
            if action == 'approve':
                sig_b64 = request.POST.get('signature_data', '')
                _save_drawn_signature(employee, sig_b64)
                ceo_sig = sig_b64 if sig_b64.startswith('data:image/') else (employee.signature_b64 or '')
                leave.director_sig_b64 = ceo_sig
                # Also fill all bypassed intermediate slots with CEO info
                now = timezone.now()
                leave.unit_head_action_by = employee
                leave.unit_head_action_date = now
                leave.unit_head_remarks = 'Approved by CEO (Administration Director leave)'
                leave.unit_head_sig_b64 = ceo_sig
                leave.manager_action_by = employee
                leave.manager_action_date = now
                leave.manager_remarks = 'Approved by CEO (Administration Director leave)'
                leave.manager_sig_b64 = ceo_sig
                leave.hr_action_by = employee
                leave.hr_action_date = now
                leave.hr_remarks = 'Approved by CEO'
                leave.hr_sig_b64 = ceo_sig
                leave.status = LeaveRequest.STATUS_APPROVED
                leave.save()
                messages.success(request, "Leave request FULLY APPROVED by CEO.")
                cover_note = ''
                if leave.employee.role == 'admin_director':
                    cover_note = ' Note: The Finance Director will cover your responsibilities during your absence.'
                elif leave.employee.role == 'medical_director':
                    cover_note = ' Please ensure your medical duties are covered during your absence.'
                notify(
                    leave.employee.user,
                    'Leave Request — Approved by CEO',
                    f'Your {leave.leave_type} request ({leave.start_date} → {leave.end_date}, '
                    f'{leave.total_days} day(s)) has been approved by the CEO.' + cover_note,
                    notification_type='leave_approved',
                    url=reverse('leaves:detail', kwargs={'pk': leave.pk}),
                )
            else:
                leave.status = LeaveRequest.STATUS_REJECTED_DIRECTOR
                leave.save()
                messages.warning(request, f"Leave request #{pk} rejected by CEO.")
                notify(
                    leave.employee.user,
                    'Leave Request — Rejected by CEO',
                    f'Your {leave.leave_type} request ({leave.start_date} → {leave.end_date}) '
                    f'was rejected by the CEO.'
                    + (f' Remarks: {remarks}' if remarks else ''),
                    notification_type='leave_rejected',
                    url=reverse('leaves:detail', kwargs={'pk': leave.pk}),
                )
            try:
                from dashboard.models import AuditLog
                audit_action = AuditLog.ACTION_LEAVE_APPROVE if action == 'approve' else AuditLog.ACTION_LEAVE_REJECT
                AuditLog.log(
                    request, audit_action,
                    f"CEO {'approved' if action == 'approve' else 'rejected'} leave request "
                    f"for {leave.employee.get_full_name()} ({leave.leave_type}, {leave.start_date} → {leave.end_date})",
                    target_user=leave.employee.user,
                )
            except Exception:
                pass
            return redirect('leaves:ceo_approvals')
    else:
        form = ApprovalForm()

    return render(request, 'leaves/action_form.html', {
        'leave': leave,
        'form': form,
        'action_title': 'CEO — Administration Director Leave Approval',
        'action_type': 'ceo',
        'current_sig_b64': employee.signature_b64 or '',
    })


@login_required
def leave_detail(request, pk):
    employee = get_employee(request)
    leave = get_object_or_404(LeaveRequest, pk=pk)

    # Only owner, their manager, HR, Director, CEO, or Superuser can view
    can_view = (
        leave.employee == employee or
        request.user.is_superuser or
        (employee and employee.is_hr()) or
        (employee and employee.is_director()) or
        (employee and employee.is_ceo()) or
        (employee and employee.is_manager() and leave.employee.supervisor == employee)
    )
    if not can_view:
        messages.error(request, "Access denied.")
        return redirect('dashboard:home')

    return render(request, 'leaves/leave_detail.html', {'leave': leave})


@login_required
def leave_edit(request, pk):
    """Employee edits their own leave request — only while status is still pending."""
    employee = get_employee(request)
    leave = get_object_or_404(LeaveRequest, pk=pk)

    if leave.employee != employee:
        messages.error(request, "You can only edit your own leave requests.")
        return redirect('dashboard:home')
    if leave.status != LeaveRequest.STATUS_PENDING:
        messages.error(request, "This request can no longer be edited — it has already been reviewed.")
        return redirect('leaves:detail', pk=pk)

    balance, _ = LeaveBalance.objects.get_or_create(
        employee=employee, year=date.today().year,
        defaults={'total_entitlement': seniority_entitlement(employee)}
    )

    dept_employees = Employee.objects.filter(
        department=employee.department, is_active=True,
    ).exclude(pk=employee.pk).select_related('user')
    on_leave_pks = set(
        LeaveRequest.objects.filter(
            status='approved',
            employee__in=dept_employees,
            start_date__lte=date.today(),
            end_date__gte=date.today(),
        ).values_list('employee_id', flat=True)
    )
    backup_choices = [
        {'id': e.pk, 'name': e.get_full_name(), 'unavailable': e.pk in on_leave_pks}
        for e in dept_employees
    ]

    form = LeaveRequestForm(request.POST or None, request.FILES or None, instance=leave)
    if request.method == 'POST' and form.is_valid():
        updated = form.save(commit=False)

        backup_id = request.POST.get('backup_employee')
        if backup_id:
            try:
                updated.backup_employee = Employee.objects.get(pk=backup_id)
            except Employee.DoesNotExist:
                pass

        if updated.leave_type.is_deductible and updated.total_days > balance.remaining_days + leave.total_days:
            messages.error(request, f"Insufficient leave balance.")
        else:
            overlapping = LeaveRequest.objects.filter(
                employee=employee,
                status__in=['pending', 'unit_head_approved', 'manager_approved', 'hr_approved', 'approved'],
                start_date__lte=updated.end_date,
                end_date__gte=updated.start_date,
            ).exclude(pk=leave.pk)
            if overlapping.exists():
                messages.error(request, "You already have a leave request for overlapping dates.")
            else:
                updated.save()
                messages.success(request, "Leave request updated successfully.")

                # Notify unit head and line manager about the change
                notif_title = f'Leave Request Edited — {employee.get_full_name()}'
                notif_msg = (
                    f'{employee.get_full_name()} has edited their {updated.leave_type} leave request '
                    f'(now: {updated.start_date} → {updated.end_date}, {updated.total_days} day(s)). '
                    f'Please review the updated request before approving.'
                )
                detail_url = reverse('leaves:detail', kwargs={'pk': pk})
                recipients = set()
                if employee.unit_head:
                    recipients.add(employee.unit_head.user)
                if employee.supervisor:
                    recipients.add(employee.supervisor.user)
                for recipient in recipients:
                    notify(recipient, notif_title, notif_msg,
                           notification_type='leave_submitted', url=detail_url)
                return redirect('leaves:detail', pk=pk)

    leave_types = LeaveType.objects.filter(is_active=True).values('id', 'is_deductible')
    deductible_map = json.dumps({str(lt['id']): lt['is_deductible'] for lt in leave_types})

    return render(request, 'leaves/edit_form.html', {
        'form': form,
        'leave': leave,
        'balance': balance,
        'deductible_map': deductible_map,
        'backup_choices': backup_choices,
        'employee': employee,
        'current_sig_b64': employee.signature_b64 or '',
    })


@login_required
def pdf_leave(request, pk):
    """Download the official leave authorisation form as a filled PDF."""
    from django.http import HttpResponse
    from .pdf_utils import generate_leave_pdf

    employee = get_employee(request)
    leave = get_object_or_404(LeaveRequest, pk=pk)

    can_view = (
        leave.employee == employee or
        request.user.is_superuser or
        (employee and employee.is_hr()) or
        (employee and employee.is_director()) or
        (employee and employee.is_ceo()) or
        (employee and employee.is_manager() and leave.employee.supervisor == employee) or
        (employee and employee.is_unit_head() and leave.employee.unit_head == employee)
    )
    if not can_view:
        messages.error(request, "Access denied.")
        return redirect('dashboard:home')

    buf = generate_leave_pdf(leave)
    filename = f"leave_authorisation_{leave.employee.user.last_name}_{leave.start_date}.pdf"
    response = HttpResponse(buf, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    return response


@login_required
def print_leave(request, pk):
    """Redirect to the official PDF — the old HTML print view is retired."""
    return redirect('leaves:pdf_leave', pk=pk)


@login_required
def all_leaves_hr(request):
    """HR/Director view of all leave requests"""
    employee = get_employee(request)
    if not employee or (not employee.is_hr() and not employee.is_director() and not employee.is_ceo()):
        messages.error(request, "Access denied.")
        return redirect('dashboard:home')

    status_filter = request.GET.get('status', '')
    dept_filter   = request.GET.get('dept', '')
    emp_filter    = request.GET.get('employee', '')
    try:
        year_filter = int(request.GET.get('year', date.today().year))
    except (ValueError, TypeError):
        year_filter = date.today().year

    qs = LeaveRequest.objects.select_related(
        'employee__user', 'employee__department', 'leave_type'
    ).filter(start_date__year=year_filter)

    if status_filter:
        qs = qs.filter(status=status_filter)
    if emp_filter:
        qs = qs.filter(employee_id=emp_filter)
    elif dept_filter:
        qs = qs.filter(employee__department_id=dept_filter)

    from accounts.models import Department
    departments = Department.objects.all()
    all_employees = Employee.objects.filter(is_active=True).select_related('user', 'department').order_by('user__last_name')

    return render(request, 'leaves/all_leaves.html', {
        'leave_requests': qs,
        'status_filter': status_filter,
        'dept_filter': dept_filter,
        'emp_filter': emp_filter,
        'year_filter': year_filter,
        'departments': departments,
        'all_employees': all_employees,
        'status_choices': LeaveRequest.STATUS_CHOICES,
        'years': range(2000, date.today().year + 11),
    })


@login_required
def employee_leave_summary(request, pk):
    """Full leave summary for a specific employee — accessible to HR, Director, and the employee themselves."""
    from accounts.models import Employee as EmpModel
    viewer = get_employee(request)
    target = get_object_or_404(EmpModel, pk=pk)

    # Access: HR, Director, superuser, or the employee viewing their own summary
    if not (request.user.is_superuser or
            (viewer and (viewer.is_hr() or viewer.is_director())) or
            (viewer and viewer.pk == target.pk)):
        messages.error(request, "Access denied.")
        return redirect('dashboard:home')

    try:
        year = int(request.GET.get('year', date.today().year))
    except (ValueError, TypeError):
        year = date.today().year
    balance, _ = LeaveBalance.objects.get_or_create(
        employee=target, year=year,
        defaults={'total_entitlement': seniority_entitlement(target, year)}
    )

    all_requests = LeaveRequest.objects.filter(
        employee=target, start_date__year=year
    ).select_related('leave_type').order_by('-start_date')

    return render(request, 'leaves/employee_leave_summary.html', {
        'target': target,
        'balance': balance,
        'all_requests': all_requests,
        'year': year,
        'years': range(2000, date.today().year + 11),
    })


@login_required
def admin_override_leave(request, pk):
    """Superuser-only: cancel or revert any leave request back to pending."""
    if not request.user.is_superuser:
        messages.error(request, "Access denied.")
        return redirect('dashboard:home')
    if request.method != 'POST':
        return redirect('leaves:detail', pk=pk)

    leave = get_object_or_404(LeaveRequest, pk=pk)
    action = request.POST.get('admin_action')
    reason = request.POST.get('admin_reason', '').strip()

    if action == 'cancel':
        leave.status = LeaveRequest.STATUS_CANCELLED
        leave.director_remarks = f"[Admin cancelled] {reason}"
        leave.save()
        messages.success(request, f"Leave request #{pk} cancelled by admin. Balance restored automatically.")

    elif action == 'revert':
        # Send back to pending — clears all approval chain
        leave.status = LeaveRequest.STATUS_PENDING
        leave.manager_action_by = None
        leave.manager_action_date = None
        leave.manager_remarks = ''
        leave.hr_action_by = None
        leave.hr_action_date = None
        leave.hr_remarks = ''
        leave.director_action_by = None
        leave.director_action_date = None
        leave.director_remarks = f"[Admin reverted to pending] {reason}"
        leave.save()
        messages.success(request, f"Leave request #{pk} reverted to Pending. It must go through approval again.")

    return redirect('leaves:detail', pk=pk)


@login_required
def admin_edit_leave(request, pk):
    """Superuser-only: correct total_days, dates, status, and add a note on any leave request."""
    if not request.user.is_superuser:
        messages.error(request, "Access denied.")
        return redirect('dashboard:home')
    if request.method != 'POST':
        return redirect('leaves:detail', pk=pk)

    leave = get_object_or_404(LeaveRequest, pk=pk)

    try:
        new_days = int(request.POST.get('corrected_days', '').strip())
        if new_days < 0:
            raise ValueError
    except (ValueError, AttributeError):
        messages.error(request, "Invalid day count. Please enter a whole number ≥ 0.")
        return redirect('leaves:detail', pk=pk)

    new_status = request.POST.get('corrected_status', '').strip()
    admin_note = request.POST.get('correction_note', '').strip()
    new_start  = request.POST.get('corrected_start', '').strip()
    new_end    = request.POST.get('corrected_end', '').strip()

    valid_statuses = dict(LeaveRequest.STATUS_CHOICES).keys()
    update_fields = {'total_days': new_days}

    if new_status and new_status in valid_statuses:
        update_fields['status'] = new_status

    # Update dates if provided — validate first
    from datetime import datetime as _dt
    parsed_start = parsed_end = None
    if new_start:
        try:
            parsed_start = _dt.strptime(new_start, '%Y-%m-%d').date()
            update_fields['start_date'] = parsed_start
        except ValueError:
            messages.error(request, "Invalid start date format.")
            return redirect('leaves:detail', pk=pk)
    if new_end:
        try:
            parsed_end = _dt.strptime(new_end, '%Y-%m-%d').date()
            update_fields['end_date'] = parsed_end
        except ValueError:
            messages.error(request, "Invalid end date format.")
            return redirect('leaves:detail', pk=pk)

    if admin_note:
        existing = leave.director_remarks or ''
        separator = '\n' if existing else ''
        update_fields['director_remarks'] = f"{existing}{separator}[Admin correction] {admin_note}"

    # Use queryset .update() to bypass save() so total_days is NOT recalculated
    LeaveRequest.objects.filter(pk=pk).update(**update_fields)

    parts = [f"{new_days} days"]
    if 'start_date' in update_fields:
        parts.append(f"start → {new_start}")
    if 'end_date' in update_fields:
        parts.append(f"end → {new_end}")
    if 'status' in update_fields:
        parts.append(f"status → {new_status}")
    messages.success(request, f"Leave #{pk} corrected: {', '.join(parts)}.")
    return redirect('leaves:detail', pk=pk)


# ── Leave Reversal (HR + Superadmin) ────────────────────────────────────────

@login_required
def leave_reversal(request, pk):
    """HR and superadmin: reverse (cancel) or modify an approved leave with full audit trail."""
    from accounts.models import Employee as _Employee
    viewer = None
    try:
        viewer = request.user.employee
    except Exception:
        pass

    is_hr_or_admin = request.user.is_superuser or (viewer and viewer.is_hr())
    if not is_hr_or_admin:
        messages.error(request, "Access denied. HR or System Admin only.")
        return redirect('dashboard:home')

    leave = get_object_or_404(
        LeaveRequest.objects.select_related('employee__user', 'employee__supervisor__user',
                                            'leave_type'),
        pk=pk
    )

    from .models import LeaveReversal as _LR
    from datetime import datetime as _dt

    if request.method == 'POST':
        action_type = request.POST.get('action_type', '').strip()
        reason      = request.POST.get('reason', '').strip()
        new_start   = request.POST.get('new_start', '').strip()
        new_end     = request.POST.get('new_end', '').strip()

        if action_type not in (_LR.ACTION_REVERSE, _LR.ACTION_MODIFIED):
            messages.error(request, "Invalid action type.")
            return redirect('leaves:leave_reversal', pk=pk)
        if not reason:
            messages.error(request, "A reason is required.")
            return redirect('leaves:leave_reversal', pk=pk)

        # Snapshot before change
        original_status = leave.status
        original_start  = leave.start_date
        original_end    = leave.end_date
        original_days   = leave.total_days

        if action_type == _LR.ACTION_REVERSE:
            leave.status = LeaveRequest.STATUS_CANCELLED
            leave.director_remarks = (
                f"[REVERSED by {request.user.get_full_name()} — {date.today()}] {reason}"
            )
            LeaveRequest.objects.filter(pk=pk).update(
                status=leave.status,
                director_remarks=leave.director_remarks,
            )
            _LR.objects.create(
                leave_request=leave,
                action_type=_LR.ACTION_REVERSE,
                reason=reason,
                reversed_by=request.user,
                original_status=original_status,
                original_start_date=original_start,
                original_end_date=original_end,
                original_total_days=original_days,
            )
            messages.success(request, f"Leave #{pk} reversed and cancelled. Audit record saved.")

        else:  # modified
            try:
                p_start = _dt.strptime(new_start, '%Y-%m-%d').date()
                p_end   = _dt.strptime(new_end,   '%Y-%m-%d').date()
            except ValueError:
                messages.error(request, "Invalid dates for modification.")
                return redirect('leaves:leave_reversal', pk=pk)
            if p_start > p_end:
                messages.error(request, "Start date must be before end date.")
                return redirect('leaves:leave_reversal', pk=pk)

            new_days = LeaveRequest._count_working_days(p_start, p_end)
            LeaveRequest.objects.filter(pk=pk).update(
                start_date=p_start,
                end_date=p_end,
                total_days=new_days,
                director_remarks=(
                    f"[MODIFIED by {request.user.get_full_name()} — {date.today()}] {reason}"
                ),
            )
            _LR.objects.create(
                leave_request=leave,
                action_type=_LR.ACTION_MODIFIED,
                reason=reason,
                reversed_by=request.user,
                original_status=original_status,
                original_start_date=original_start,
                original_end_date=original_end,
                original_total_days=original_days,
                new_start_date=p_start,
                new_end_date=p_end,
                new_total_days=new_days,
            )
            messages.success(request, f"Leave #{pk} modified: {p_start} → {p_end} ({new_days} days). Audit record saved.")

        # Notify all parties
        from notifications.utils import notify as _notify
        emp = leave.employee
        notify_users = [emp.user]
        if emp.supervisor:
            notify_users.append(emp.supervisor.user)
        if emp.unit_head:
            notify_users.append(emp.unit_head.user)
        try:
            from accounts.models import Employee as _E
            hr_users = _E.objects.filter(role='hr', is_active=True).select_related('user')
            for hr in hr_users:
                if hr.user not in notify_users:
                    notify_users.append(hr.user)
        except Exception:
            pass

        action_label = "reversed/cancelled" if action_type == _LR.ACTION_REVERSE else "dates modified"
        for u in notify_users:
            if u != request.user:
                _notify(
                    u,
                    f"Leave #{pk} {action_label} — {emp.get_full_name()}",
                    (
                        f"The leave request #{pk} for {emp.get_full_name()} "
                        f"({leave.leave_type}) has been {action_label} by HR.\n\n"
                        f"Reason: {reason}"
                    ),
                    notification_type='leave_cancelled',
                    url=reverse('leaves:detail', kwargs={'pk': pk}),
                )

        return redirect('leaves:detail', pk=pk)

    # GET — show form
    reversals = _LR.objects.filter(leave_request=leave).select_related('reversed_by')
    return render(request, 'leaves/reversal_form.html', {
        'leave': leave,
        'reversals': reversals,
    })


@login_required
def leave_reversals_report(request):
    """HR/superadmin: audit log of all leave reversals."""
    viewer = None
    try:
        viewer = request.user.employee
    except Exception:
        pass
    if not (request.user.is_superuser or (viewer and viewer.is_hr())):
        messages.error(request, "Access denied.")
        return redirect('dashboard:home')

    from .models import LeaveReversal as _LR
    reversals = _LR.objects.select_related(
        'leave_request__employee__user', 'leave_request__leave_type', 'reversed_by'
    ).all()

    # Filters
    emp_q    = request.GET.get('employee', '').strip()
    action_q = request.GET.get('action', '').strip()
    if emp_q:
        reversals = reversals.filter(
            leave_request__employee__user__first_name__icontains=emp_q
        ) | reversals.filter(
            leave_request__employee__user__last_name__icontains=emp_q
        )
    if action_q:
        reversals = reversals.filter(action_type=action_q)

    return render(request, 'leaves/reversals_report.html', {
        'reversals': reversals,
        'emp_q': emp_q,
        'action_q': action_q,
    })


# ── Leave Type Management (superuser only) ──────────────────────────────────

DEFAULT_LEAVE_TYPES = [
    {'name': 'Annual Leave',               'is_deductible': True,  'color': 'primary',   'requires_document': False},
    {'name': 'Permission',                 'is_deductible': True,  'color': 'warning',   'requires_document': False},
    {'name': 'Permission for School Leave','is_deductible': True,  'color': 'info',      'requires_document': True},
    {'name': 'Sick Leave',                 'is_deductible': False, 'color': 'danger',    'requires_document': True},
    {'name': 'Maternity Leave',            'is_deductible': False, 'color': 'success',   'requires_document': True},
    {'name': 'Paternity Leave',            'is_deductible': False, 'color': 'success',   'requires_document': True},
    {'name': 'Marriage Leave',             'is_deductible': False, 'color': 'secondary', 'requires_document': False},
    {'name': 'Compassionate Leave',        'is_deductible': False, 'color': 'dark',      'requires_document': False},
    {'name': 'Study Leave',               'is_deductible': False, 'color': 'secondary', 'requires_document': False},
]


def seed_default_leave_types():
    """Create the default leave types if they don't already exist."""
    for lt in DEFAULT_LEAVE_TYPES:
        LeaveType.objects.get_or_create(name=lt['name'], defaults={
            'is_deductible':      lt['is_deductible'],
            'color':              lt['color'],
            'requires_document':  lt['requires_document'],
            'is_active':          True,
        })


@login_required
def leave_type_list(request):
    emp = get_employee(request)
    if not (request.user.is_superuser or (emp and (emp.is_hr() or emp.is_ceo()))):
        messages.error(request, "Access denied.")
        return redirect('dashboard:home')
    leave_types = LeaveType.objects.all()
    return render(request, 'leaves/leave_type_list.html', {'leave_types': leave_types})


@login_required
def leave_type_create(request):
    if not request.user.is_superuser:
        messages.error(request, "Access denied.")
        return redirect('dashboard:home')
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        is_deductible = request.POST.get('is_deductible') == '1'
        requires_document = request.POST.get('requires_document') == '1'
        color = request.POST.get('color', 'primary').strip()
        is_active = request.POST.get('is_active') == '1'
        if not name:
            messages.error(request, "Leave type name is required.")
        elif LeaveType.objects.filter(name__iexact=name).exists():
            messages.error(request, f"A leave type named '{name}' already exists.")
        else:
            LeaveType.objects.create(
                name=name,
                is_deductible=is_deductible,
                requires_document=requires_document,
                color=color,
                is_active=is_active,
            )
            messages.success(request, f"Leave type '{name}' created.")
            return redirect('leaves:leave_type_list')
    _colors = [
        ('primary','Primary'),('secondary','Secondary'),('success','Success'),
        ('danger','Danger'),('warning','Warning'),('info','Info'),('dark','Dark'),
    ]
    return render(request, 'leaves/leave_type_form.html', {'action': 'Create', 'lt': None, 'colors': _colors})


@login_required
def leave_type_edit(request, pk):
    if not request.user.is_superuser:
        messages.error(request, "Access denied.")
        return redirect('dashboard:home')
    lt = get_object_or_404(LeaveType, pk=pk)
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        is_deductible = request.POST.get('is_deductible') == '1'
        requires_document = request.POST.get('requires_document') == '1'
        color = request.POST.get('color', 'primary').strip()
        is_active = request.POST.get('is_active') == '1'
        if not name:
            messages.error(request, "Leave type name is required.")
        elif LeaveType.objects.filter(name__iexact=name).exclude(pk=pk).exists():
            messages.error(request, f"A leave type named '{name}' already exists.")
        else:
            lt.name = name
            lt.is_deductible = is_deductible
            lt.requires_document = requires_document
            lt.color = color
            lt.is_active = is_active
            lt.save()
            messages.success(request, f"Leave type '{name}' updated.")
            return redirect('leaves:leave_type_list')
    _colors = [
        ('primary','Primary'),('secondary','Secondary'),('success','Success'),
        ('danger','Danger'),('warning','Warning'),('info','Info'),('dark','Dark'),
    ]
    return render(request, 'leaves/leave_type_form.html', {'action': 'Edit', 'lt': lt, 'colors': _colors})


@login_required
def leave_type_delete(request, pk):
    if not request.user.is_superuser:
        messages.error(request, "Access denied.")
        return redirect('dashboard:home')
    lt = get_object_or_404(LeaveType, pk=pk)
    if request.method == 'POST':
        name = lt.name
        try:
            lt.delete()
            messages.success(request, f"Leave type '{name}' deleted.")
        except Exception:
            messages.error(request, f"Cannot delete '{name}' — it is referenced by existing leave requests.")
    return redirect('leaves:leave_type_list')


@login_required
def restore_default_leave_types(request):
    if not request.user.is_superuser:
        messages.error(request, "Access denied.")
        return redirect('dashboard:home')
    if request.method == 'POST':
        seed_default_leave_types()
        messages.success(request, "Default leave types restored (existing ones were not changed).")
    return redirect('leaves:leave_type_list')


@login_required
def backfill_signatures(request):
    """
    HR/superuser: copy each employee's current signature_b64 to any historical
    leave requests where the snapshot field is still empty.
    This repairs existing leaves after the b64 signature system was introduced.
    """
    employee = get_employee(request)
    if not employee or not (employee.is_hr() or request.user.is_superuser):
        messages.error(request, "Access denied.")
        return redirect('dashboard:home')

    if request.method == 'POST':
        updated = 0
        for leave in LeaveRequest.objects.select_related(
            'employee', 'unit_head_action_by', 'manager_action_by',
            'hr_action_by', 'director_action_by'
        ):
            changed = False
            if not leave.employee_sig_b64 and leave.employee and leave.employee.signature_b64:
                leave.employee_sig_b64 = leave.employee.signature_b64
                changed = True
            if not leave.unit_head_sig_b64 and leave.unit_head_action_by and leave.unit_head_action_by.signature_b64:
                leave.unit_head_sig_b64 = leave.unit_head_action_by.signature_b64
                changed = True
            if not leave.manager_sig_b64 and leave.manager_action_by and leave.manager_action_by.signature_b64:
                leave.manager_sig_b64 = leave.manager_action_by.signature_b64
                changed = True
            if not leave.hr_sig_b64 and leave.hr_action_by and leave.hr_action_by.signature_b64:
                leave.hr_sig_b64 = leave.hr_action_by.signature_b64
                changed = True
            if not leave.director_sig_b64 and leave.director_action_by and leave.director_action_by.signature_b64:
                leave.director_sig_b64 = leave.director_action_by.signature_b64
                changed = True
            if changed:
                leave.save(update_fields=[
                    'employee_sig_b64', 'unit_head_sig_b64', 'manager_sig_b64',
                    'hr_sig_b64', 'director_sig_b64'
                ])
                updated += 1
        messages.success(request, f"Signatures synced to {updated} leave record(s).")
        return redirect('leaves:all_leaves')

    total = LeaveRequest.objects.count()
    missing = LeaveRequest.objects.filter(
        employee_sig_b64='', manager_sig_b64='', hr_sig_b64='', director_sig_b64=''
    ).count()
    return render(request, 'leaves/backfill_signatures.html', {
        'total': total,
        'missing': missing,
    })


@login_required
def set_leave_entitlement(request):
    """
    Superuser/HR: set a custom annual leave entitlement for a specific employee + year.
    Accessed via POST from the Leave Tracker modal.
    """
    employee = get_employee(request)
    if not (request.user.is_superuser or (employee and employee.is_hr())):
        messages.error(request, "Access denied.")
        return redirect('dashboard:home')

    if request.method == 'POST':
        emp_pk = request.POST.get('employee_id')
        year   = request.POST.get('year')
        days   = request.POST.get('total_entitlement')

        try:
            emp_pk = int(emp_pk)
            year   = int(year)
            days   = int(days)
            if days < 0 or days > 365:
                raise ValueError
        except (TypeError, ValueError):
            messages.error(request, "Invalid entitlement value.")
            return redirect('dashboard:tracker')

        target = get_object_or_404(Employee, pk=emp_pk)
        balance, _ = LeaveBalance.objects.get_or_create(
            employee=target, year=year,
            defaults={'total_entitlement': days}
        )
        if balance.total_entitlement != days:
            balance.total_entitlement = days
            balance.save(update_fields=['total_entitlement'])
        messages.success(
            request,
            f"Leave entitlement for {target.get_full_name()} ({year}) set to {days} days."
        )

    return redirect(request.POST.get('next') or 'dashboard:tracker')


# ── Leave Consultation (Private Seek-Guidance) ────────────────────────────────

@login_required
def seek_consultation(request, leave_pk):
    """Approver seeks private guidance from a senior colleague before deciding."""
    employee = get_employee(request)
    if not employee:
        messages.error(request, "Access denied.")
        return redirect('dashboard:home')

    leave = get_object_or_404(LeaveRequest, pk=leave_pk)

    # Only approvers who currently have authority on this leave may open a consultation
    is_approver = (
        request.user.is_superuser or
        employee.is_hr() or
        employee.is_director() or
        employee.is_ceo() or
        (employee.is_unit_head() and leave.employee.unit_head == employee) or
        (employee.is_manager() and leave.employee.supervisor == employee)
    )
    if not is_approver:
        messages.error(request, "Only approvers may seek guidance on a leave request.")
        return redirect('leaves:detail', pk=leave_pk)

    # Eligible consultees: any active employee except the leave.employee and the current user
    candidates = Employee.objects.filter(
        is_active=True
    ).exclude(pk=employee.pk).exclude(pk=leave.employee.pk).select_related('user').order_by('user__last_name')

    if request.method == 'POST':
        consulted_pk = request.POST.get('consulted_with')
        private_note = request.POST.get('private_note', '').strip()
        if not consulted_pk or not private_note:
            messages.error(request, "Please select a person and write a note.")
        else:
            consulted = get_object_or_404(Employee, pk=consulted_pk)
            consultation = LeaveConsultation.objects.create(
                leave_request=leave,
                requested_by=employee,
                consulted_with=consulted,
                private_note=private_note,
            )
            notify(
                consulted.user,
                f'Private Guidance Request — {leave.employee.get_full_name()}',
                f'{employee.get_full_name()} is seeking your private guidance on a leave request '
                f'from {leave.employee.get_full_name()} ({leave.leave_type}, '
                f'{leave.start_date} – {leave.end_date}). '
                f'Please log in to respond privately.',
                notification_type='system',
                url=reverse('leaves:respond_consultation', kwargs={'pk': consultation.pk}),
            )
            messages.success(request, f"Guidance request sent privately to {consulted.get_full_name()}.")
            return redirect('leaves:detail', pk=leave_pk)

    import json as _json
    existing = LeaveConsultation.objects.filter(leave_request=leave, requested_by=employee)
    candidates_json = _json.dumps([
        {'id': e.pk, 'name': e.get_full_name(), 'role': e.get_role_display()}
        for e in candidates
    ])
    return render(request, 'leaves/seek_consultation.html', {
        'leave': leave,
        'candidates_json': candidates_json,
        'existing': existing,
    })


@login_required
def respond_consultation(request, pk):
    """The consulted person reads the private note and responds: proceed or hold."""
    employee = get_employee(request)
    consultation = get_object_or_404(LeaveConsultation, pk=pk)

    if not employee or consultation.consulted_with != employee:
        messages.error(request, "Access denied — this consultation is not addressed to you.")
        return redirect('dashboard:home')

    if consultation.status != LeaveConsultation.STATUS_PENDING:
        messages.info(request, "You have already responded to this consultation.")
        return render(request, 'leaves/respond_consultation.html', {'consultation': consultation})

    if request.method == 'POST':
        status = request.POST.get('response_status')
        note   = request.POST.get('response_note', '').strip()
        if status not in (LeaveConsultation.STATUS_PROCEED, LeaveConsultation.STATUS_HOLD):
            messages.error(request, "Please choose Proceed or Hold.")
        else:
            consultation.status        = status
            consultation.response_note = note
            consultation.responded_at  = timezone.now()
            consultation.save()
            notify(
                consultation.requested_by.user,
                f'Guidance Response — {consultation.leave_request.employee.get_full_name()}',
                f'{employee.get_full_name()} has responded to your guidance request: '
                f'{"✅ Proceed" if status == LeaveConsultation.STATUS_PROCEED else "⏸ Hold"}. '
                + (f'Note: {note}' if note else ''),
                notification_type='system',
                url=reverse('leaves:detail', kwargs={'pk': consultation.leave_request.pk}),
            )
            messages.success(request, "Your response has been sent privately.")
            return redirect('dashboard:home')

    return render(request, 'leaves/respond_consultation.html', {'consultation': consultation})


@login_required
def my_consultations(request):
    """List consultations awaiting the current user's response."""
    employee = get_employee(request)
    if not employee:
        messages.error(request, "Access denied.")
        return redirect('dashboard:home')

    pending = LeaveConsultation.objects.filter(
        consulted_with=employee,
        status=LeaveConsultation.STATUS_PENDING,
    ).select_related('leave_request__employee__user', 'requested_by__user')

    past = LeaveConsultation.objects.filter(
        consulted_with=employee,
    ).exclude(status=LeaveConsultation.STATUS_PENDING).select_related(
        'leave_request__employee__user', 'requested_by__user'
    )[:20]

    return render(request, 'leaves/my_consultations.html', {
        'pending': pending,
        'past': past,
    })


# ==============================================================================
# TENTATIVE LEAVE PLAN
# ==============================================================================

@login_required
def plan_my_plan(request):
    from .models import TentativeLeavePlan
    emp = request.user.employee
    year = int(request.GET.get("year", timezone.now().year))
    entries = TentativeLeavePlan.objects.filter(employee=emp, year=year).select_related("leave_type")
    leave_types = LeaveType.objects.filter(is_active=True)
    years = range(timezone.now().year, timezone.now().year + 3)

    if request.method == "POST":
        action = request.POST.get("action")
        if action == "add":
            start = request.POST.get("planned_start")
            end   = request.POST.get("planned_end")
            lt_id = request.POST.get("leave_type")
            notes = request.POST.get("notes", "").strip()
            if not start or not end:
                messages.error(request, "Please provide both start and end dates.")
            elif start > end:
                messages.error(request, "End date must be after start date.")
            else:
                TentativeLeavePlan.objects.create(
                    employee=emp, year=year,
                    planned_start=start, planned_end=end,
                    leave_type_id=lt_id if lt_id else None,
                    notes=notes, status="draft",
                )
                messages.success(request, "Period added to your plan.")
        elif action == "edit":
            entry_id = request.POST.get("entry_id")
            start = request.POST.get("planned_start")
            end   = request.POST.get("planned_end")
            lt_id = request.POST.get("leave_type")
            notes = request.POST.get("notes", "").strip()
            if not start or not end:
                messages.error(request, "Please provide both start and end dates.")
            elif start > end:
                messages.error(request, "End date must be after start date.")
            else:
                TentativeLeavePlan.objects.filter(
                    pk=entry_id, employee=emp, status__in=["draft","rejected"]
                ).update(
                    planned_start=start, planned_end=end,
                    leave_type_id=lt_id if lt_id else None,
                    notes=notes,
                )
                messages.success(request, "Entry updated.")
            return redirect(request.path + "?year=" + str(year))
        elif action == "delete":
            entry_id = request.POST.get("entry_id")
            TentativeLeavePlan.objects.filter(pk=entry_id, employee=emp, status__in=["draft","rejected"]).delete()
            messages.success(request, "Entry removed.")
        elif action == "submit":
            drafts = TentativeLeavePlan.objects.filter(employee=emp, year=year, status__in=["draft","rejected"])
            if not drafts.exists():
                messages.warning(request, "No draft entries to submit.")
            else:
                count = drafts.count()
                drafts.update(status="submitted", submitted_at=timezone.now())
                messages.success(request, f"{count} plan entr{'y' if count == 1 else 'ies'} submitted to your Line Manager.")
                # Notify line manager
                manager = emp.supervisor
                if manager and hasattr(manager, 'user'):
                    from notifications.utils import notify
                    notify(
                        manager.user,
                        "Tentative Leave Plan Submitted",
                        f"{emp.get_full_name()} submitted their {year} tentative leave plan ({count} entr{'y' if count == 1 else 'ies'}) for your review.",
                        notification_type="info",
                        url="/leaves/plan/manager/?year=" + str(year),
                    )
        return redirect(request.path + "?year=" + str(year))

    edit_entry = None
    edit_id = request.GET.get("edit")
    if edit_id:
        try:
            edit_entry = entries.get(pk=edit_id, status__in=["draft","rejected"])
        except TentativeLeavePlan.DoesNotExist:
            pass

    return render(request, "leaves/plan_my_plan.html", {
        "entries": entries,
        "year": year,
        "years": years,
        "leave_types": leave_types,
        "has_draft": entries.filter(status__in=["draft","rejected"]).exists(),
        "edit_entry": edit_entry,
    })


@login_required
def plan_manager_review(request):
    from .models import TentativeLeavePlan
    emp = request.user.employee
    if not (emp.is_manager() or request.user.is_superuser):
        messages.error(request, "Access denied.")
        return redirect("dashboard:home")
    year = int(request.GET.get("year", timezone.now().year))
    team = emp.subordinates.filter(is_active=True)
    base_qs = TentativeLeavePlan.objects.filter(
        employee__in=team, year=year
    ).select_related("employee__user", "employee__department", "leave_type").order_by("employee__user__last_name", "planned_start")

    if request.method == "POST":
        action   = request.POST.get("action")
        entry_id = request.POST.get("entry_id")
        note     = request.POST.get("manager_notes", "").strip()
        try:
            entry = TentativeLeavePlan.objects.get(pk=entry_id, employee__in=team)
        except TentativeLeavePlan.DoesNotExist:
            messages.error(request, "Entry not found.")
            return redirect(request.path + "?year=" + str(year))
        if action == "confirm":
            entry.status = "confirmed"
            entry.manager_confirmed_by = request.user
            entry.manager_confirmed_at = timezone.now()
            entry.manager_notes = note
            entry.save()
            messages.success(request, f"Plan confirmed for {entry.employee.get_full_name()}.")
            from notifications.utils import notify
            notify(
                entry.employee.user,
                "Tentative Leave Plan Confirmed",
                f"Your tentative leave plan ({entry.planned_start.strftime('%d %b')}–{entry.planned_end.strftime('%d %b %Y')}, {entry.total_days} day{'s' if entry.total_days != 1 else ''}) has been confirmed by your manager."
                + (f" Note: {note}" if note else ""),
                notification_type="success",
                url="/leaves/plan/my/",
            )
        elif action == "reject":
            entry.status = "rejected"
            entry.manager_confirmed_by = request.user
            entry.manager_confirmed_at = timezone.now()
            entry.manager_notes = note
            entry.save()
            messages.warning(request, f"Plan rejected for {entry.employee.get_full_name()}.")
            from notifications.utils import notify
            notify(
                entry.employee.user,
                "Tentative Leave Plan Rejected",
                f"Your tentative leave plan ({entry.planned_start.strftime('%d %b')}–{entry.planned_end.strftime('%d %b %Y')}) was rejected by your manager."
                + (f" Reason: {note}" if note else ""),
                notification_type="warning",
                url="/leaves/plan/my/",
            )
        return redirect(request.path + "?year=" + str(year))

    years = range(timezone.now().year, timezone.now().year + 3)
    return render(request, "leaves/plan_manager.html", {
        "pending":   base_qs.filter(status="submitted"),
        "confirmed": base_qs.filter(status="confirmed"),
        "rejected":  base_qs.filter(status="rejected"),
        "year": year, "years": years,
    })


@login_required
def plan_hr_overview(request):
    from .models import TentativeLeavePlan
    from accounts.models import Department
    emp = request.user.employee
    if not (emp.is_hr() or request.user.is_superuser):
        messages.error(request, "Access denied.")
        return redirect("dashboard:home")
    year    = int(request.GET.get("year", timezone.now().year))
    dept_id = request.GET.get("dept", "")
    status_f = request.GET.get("status", "confirmed")
    qs = TentativeLeavePlan.objects.filter(year=year).select_related(
        "employee__user", "employee__department", "leave_type", "manager_confirmed_by"
    ).order_by("employee__department__name", "employee__user__last_name", "planned_start")
    if dept_id:
        qs = qs.filter(employee__department_id=dept_id)
    if status_f:
        qs = qs.filter(status=status_f)
    departments = Department.objects.all().order_by("name")
    years = range(timezone.now().year, timezone.now().year + 3)

    if request.GET.get("export") == "excel":
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse as _HR
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Leave Plan {year}"
        hfont  = Font(bold=True, color="FFFFFF", size=10)
        hfill  = PatternFill("solid", fgColor="0A4D68")
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
        thin   = Side(style="thin", color="CCCCCC")
        bdr    = Border(left=thin, right=thin, top=thin, bottom=thin)
        headers = ["Department","Employee","Leave Type","Start Date","End Date","Days","Notes","Status","Confirmed By"]
        widths  = [22, 28, 18, 14, 14, 8, 30, 14, 22]
        ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
        ws["A1"].value = f"Tentative Leave Plan {year}"
        ws["A1"].font  = Font(bold=True, size=13, color="0A4D68")
        ws["A1"].alignment = center
        ws.row_dimensions[1].height = 26
        for col, (h, w) in enumerate(zip(headers, widths), 1):
            c = ws.cell(row=2, column=col, value=h)
            c.font = hfont; c.fill = hfill; c.alignment = center; c.border = bdr
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.row_dimensions[2].height = 20
        entries_list = list(qs)
        for ri, e in enumerate(entries_list, 3):
            vals = [
                str(e.employee.department) if e.employee.department else "---",
                e.employee.get_full_name(),
                str(e.leave_type) if e.leave_type else "---",
                e.planned_start.strftime("%d/%m/%Y"),
                e.planned_end.strftime("%d/%m/%Y"),
                e.total_days,
                e.notes or "",
                e.get_status_display(),
                e.manager_confirmed_by.get_full_name() if e.manager_confirmed_by else "---",
            ]
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=ri, column=col, value=val)
                c.border = bdr
                c.alignment = center if col in (4,5,6,8) else left
            ws.row_dimensions[ri].height = 16
        ws.freeze_panes = "A3"
        resp = _HR(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp["Content-Disposition"] = f"attachment; filename=tentative_plan_{year}.xlsx"
        wb.save(resp)
        return resp

    entries_list = list(qs)
    return render(request, "leaves/plan_hr.html", {
        "entries":       entries_list,
        "year":          year,
        "years":         years,
        "departments":   departments,
        "dept_id":       dept_id,
        "status_filter": status_f,
        "total":         len(entries_list),
        "total_days":    sum(e.total_days for e in entries_list),
    })


# ── Team Leave Calendar ───────────────────────────────────────────────────────

@login_required
def team_calendar(request):
    """
    Visual month-grid calendar: who is on leave on which days.
    - Manager/Supervisor: sees direct reports only
    - HR / Director / CEO / Superuser: sees everyone (filterable by dept)
    """
    import calendar as cal_mod
    from datetime import timedelta
    from accounts.models import Department

    emp = get_employee(request)
    is_super = request.user.is_superuser
    can_see_all = is_super or (emp and (emp.is_hr() or emp.is_director() or emp.is_ceo()))
    is_mgr = emp and emp.is_manager()

    if not (can_see_all or is_mgr):
        messages.error(request, "Access denied.")
        return redirect('dashboard:home')

    today = date.today()
    try:
        year  = int(request.GET.get('year', today.year))
        month = int(request.GET.get('month', today.month))
        if not (1 <= month <= 12):
            month = today.month
    except (ValueError, TypeError):
        year, month = today.year, today.month

    dept_filter = request.GET.get('dept', '')

    # Determine employee scope
    if can_see_all:
        employees = Employee.objects.filter(is_active=True).select_related('user', 'department')
        if dept_filter:
            employees = employees.filter(department_id=dept_filter)
        employees = list(employees.order_by('user__last_name', 'user__first_name'))
    else:
        # Manager sees direct subordinates + themselves
        employees = list(
            emp.subordinates.filter(is_active=True)
            .select_related('user', 'department')
            .order_by('user__last_name', 'user__first_name')
        )

    # All days in the selected month
    num_days = cal_mod.monthrange(year, month)[1]
    days = [date(year, month, d) for d in range(1, num_days + 1)]

    # Fetch all approved leaves overlapping this month
    month_start = date(year, month, 1)
    month_end   = date(year, month, num_days)
    emp_pks = [e.pk for e in employees]

    leaves = LeaveRequest.objects.filter(
        employee_id__in=emp_pks,
        status='approved',
        start_date__lte=month_end,
        end_date__gte=month_start,
    ).select_related('leave_type')

    # Build colour palette per leave type
    PALETTE = ['#2db4c3', '#0A4D68', '#059669', '#d97706', '#7c3aed', '#e11d48', '#0891b2', '#65a30d']
    leave_type_colors = {}
    legend = {}

    # Raw lookup: emp_pk -> day_number -> {color, label}
    raw = {e.pk: {} for e in employees}

    for lr in leaves:
        lt_name = lr.leave_type.name
        if lt_name not in leave_type_colors:
            idx = len(leave_type_colors) % len(PALETTE)
            leave_type_colors[lt_name] = PALETTE[idx]
            legend[lt_name] = PALETTE[idx]
        color = leave_type_colors[lt_name]
        cur = max(lr.start_date, month_start)
        end = min(lr.end_date, month_end)
        while cur <= end:
            if cur.weekday() < 5:
                raw[lr.employee_id][cur.day] = {'color': color, 'label': lt_name}
            cur += timedelta(days=1)

    # Build template-friendly rows: list of {emp, cells}
    # Each cell: {'day': date, 'is_weekend': bool, 'is_today': bool, 'info': dict_or_None}
    rows = []
    for e in employees:
        cells = [
            {
                'day':       d,
                'is_weekend': d.weekday() >= 5,
                'is_today':   d == today,
                'info':       raw[e.pk].get(d.day),
            }
            for d in days
        ]
        rows.append({'emp': e, 'cells': cells})

    # Navigation
    if month == 1:
        prev_year, prev_month = year - 1, 12
    else:
        prev_year, prev_month = year, month - 1
    if month == 12:
        next_year, next_month = year + 1, 1
    else:
        next_year, next_month = year, month + 1

    departments = Department.objects.all() if can_see_all else []

    return render(request, 'leaves/team_calendar.html', {
        'rows':        rows,
        'days':        days,
        'legend':      legend,
        'year':        year,
        'month':       month,
        'month_name':  cal_mod.month_name[month],
        'today':       today,
        'prev_year':   prev_year,
        'prev_month':  prev_month,
        'next_year':   next_year,
        'next_month':  next_month,
        'can_see_all': can_see_all,
        'departments': departments,
        'dept_filter': dept_filter,
        'years':       range(today.year - 2, today.year + 3),
        'emp_count':   len(employees),
    })
