from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q
from django.http import HttpResponse
from .models import Trainee
from accounts.models import Employee

# PKs of designated training coordinators
TRAINING_COORDINATOR_PKS = {32, 50}  # Amogo Fouda Marie Charlotte, Nkoa Onana Nkoa Christ


def _can_access(request):
    """Training coordinators, HR, managers, directors, CEO, and superusers."""
    if request.user.is_superuser:
        return True
    emp = getattr(request.user, 'employee', None)
    if not emp:
        return False
    if emp.pk in TRAINING_COORDINATOR_PKS:
        return True
    return (emp.is_hr() or emp.is_manager() or emp.is_director() or
            emp.is_ceo() or emp.role == 'admin_director')


def _can_edit(request):
    """Only training coordinators, HR, and superusers can add/edit/delete."""
    if request.user.is_superuser:
        return True
    emp = getattr(request.user, 'employee', None)
    if not emp:
        return False
    return emp.pk in TRAINING_COORDINATOR_PKS or emp.is_hr()


@login_required
def trainee_list(request):
    if not _can_access(request):
        messages.error(request, "Access denied.")
        return redirect('dashboard:home')

    q       = request.GET.get('q', '').strip()
    program = request.GET.get('program', '').strip()
    country = request.GET.get('country', '').strip()
    status  = request.GET.get('status', '').strip()

    trainees = Trainee.objects.select_related('supervisor__user', 'registered_by__user')

    if q:
        trainees = trainees.filter(
            Q(full_name__icontains=q) | Q(email__icontains=q) |
            Q(home_institution__icontains=q) | Q(country__icontains=q)
        )
    if program:
        trainees = trainees.filter(program__icontains=program)
    if country:
        trainees = trainees.filter(country__icontains=country)

    from datetime import date
    today = date.today()
    if status == 'active':
        trainees = trainees.filter(start_date__lte=today, end_date__gte=today)
    elif status == 'upcoming':
        trainees = trainees.filter(start_date__gt=today)
    elif status == 'completed':
        trainees = trainees.filter(end_date__lt=today)

    programs  = Trainee.objects.values_list('program', flat=True).distinct().order_by('program')
    countries = Trainee.objects.values_list('country', flat=True).distinct().order_by('country')

    return render(request, 'training/trainee_list.html', {
        'trainees': trainees,
        'programs': programs,
        'countries': countries,
        'q': q, 'program': program, 'country': country, 'status': status,
        'can_edit': _can_edit(request),
        'today': today,
        'total': trainees.count(),
    })


@login_required
def trainee_add(request):
    if not _can_edit(request):
        messages.error(request, "Access denied.")
        return redirect('training:trainee_list')

    staff = Employee.objects.filter(is_active=True).select_related('user').order_by('user__last_name')

    if request.method == 'POST':
        data = request.POST
        errors = []
        if not data.get('full_name'):
            errors.append("Full name is required.")
        if not data.get('program'):
            errors.append("Program is required.")
        if not data.get('start_date'):
            errors.append("Start date is required.")
        if not data.get('end_date'):
            errors.append("End date is required.")
        if not data.get('home_institution'):
            errors.append("Home institution is required.")
        if not data.get('country'):
            errors.append("Country is required.")

        if errors:
            for e in errors:
                messages.error(request, e)
        else:
            supervisor_id = data.get('supervisor_id') or None
            supervisor = None
            if supervisor_id:
                try:
                    supervisor = Employee.objects.get(pk=supervisor_id)
                except Employee.DoesNotExist:
                    pass

            emp = getattr(request.user, 'employee', None)
            Trainee.objects.create(
                full_name=data['full_name'],
                email=data.get('email', ''),
                phone=data.get('phone', ''),
                gender=data.get('gender', ''),
                program=data['program'],
                start_date=data['start_date'],
                end_date=data['end_date'],
                home_institution=data['home_institution'],
                country=data['country'],
                supervisor=supervisor,
                certificate_issued=bool(data.get('certificate_issued')),
                remarks=data.get('remarks', ''),
                registered_by=emp,
            )
            messages.success(request, f"{data['full_name']} registered successfully.")
            return redirect('training:trainee_list')

    return render(request, 'training/trainee_form.html', {
        'staff': staff,
        'program_choices': Trainee.PROGRAM_CHOICES,
        'gender_choices': Trainee.GENDER_CHOICES,
        'action': 'Add',
    })


@login_required
def trainee_edit(request, pk):
    if not _can_edit(request):
        messages.error(request, "Access denied.")
        return redirect('training:trainee_list')

    trainee = get_object_or_404(Trainee, pk=pk)
    staff   = Employee.objects.filter(is_active=True).select_related('user').order_by('user__last_name')

    if request.method == 'POST':
        data = request.POST
        supervisor_id = data.get('supervisor_id') or None
        supervisor = None
        if supervisor_id:
            try:
                supervisor = Employee.objects.get(pk=supervisor_id)
            except Employee.DoesNotExist:
                pass

        trainee.full_name        = data.get('full_name', trainee.full_name)
        trainee.email            = data.get('email', '')
        trainee.phone            = data.get('phone', '')
        trainee.gender           = data.get('gender', '')
        trainee.program          = data.get('program', trainee.program)
        trainee.start_date       = data.get('start_date', trainee.start_date)
        trainee.end_date         = data.get('end_date', trainee.end_date)
        trainee.home_institution = data.get('home_institution', trainee.home_institution)
        trainee.country          = data.get('country', trainee.country)
        trainee.supervisor       = supervisor
        trainee.certificate_issued = bool(data.get('certificate_issued'))
        trainee.remarks          = data.get('remarks', '')
        trainee.save()
        messages.success(request, f"{trainee.full_name} updated successfully.")
        return redirect('training:trainee_list')

    return render(request, 'training/trainee_form.html', {
        'trainee': trainee,
        'staff': staff,
        'program_choices': Trainee.PROGRAM_CHOICES,
        'gender_choices': Trainee.GENDER_CHOICES,
        'action': 'Edit',
    })


@login_required
def trainee_delete(request, pk):
    if not _can_edit(request):
        messages.error(request, "Access denied.")
        return redirect('training:trainee_list')
    trainee = get_object_or_404(Trainee, pk=pk)
    if request.method == 'POST':
        name = trainee.full_name
        trainee.delete()
        messages.success(request, f"{name} removed from registry.")
    return redirect('training:trainee_list')


@login_required
def export_excel(request):
    if not _can_access(request):
        messages.error(request, "Access denied.")
        return redirect('dashboard:home')

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from datetime import date

    trainees = Trainee.objects.select_related('supervisor__user').order_by('-start_date', 'full_name')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trainees"

    header_fill = PatternFill("solid", fgColor="0A4D68")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    headers = [
        'Full Name / Nom complet', 'Email', 'Phone', 'Gender',
        'Program / Programme', 'Start Date', 'End Date', 'Duration (days)',
        'Home Institution', 'Country', 'Supervisor', 'Certificate Issued',
        'Status', 'Remarks', 'Registered By',
    ]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws.row_dimensions[1].height = 30

    today = date.today()
    for row, t in enumerate(trainees, 2):
        if today < t.start_date:
            st = 'Upcoming'
        elif today > t.end_date:
            st = 'Completed'
        else:
            st = 'Active'

        ws.append([
            t.full_name, t.email, t.phone,
            dict(Trainee.GENDER_CHOICES).get(t.gender, ''),
            t.program,
            t.start_date.strftime('%d/%m/%Y') if t.start_date else '',
            t.end_date.strftime('%d/%m/%Y') if t.end_date else '',
            t.duration_days(),
            t.home_institution, t.country,
            t.supervisor.get_full_name() if t.supervisor else '',
            'Yes' if t.certificate_issued else 'No',
            st, t.remarks,
            t.registered_by.get_full_name() if t.registered_by else '',
        ])

    col_widths = [30, 28, 16, 10, 25, 14, 14, 14, 30, 18, 25, 14, 12, 30, 25]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="trainees_{today}.xlsx"'
    wb.save(response)
    return response


@login_required
def export_pdf(request):
    if not _can_access(request):
        messages.error(request, "Access denied.")
        return redirect('dashboard:home')

    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from io import BytesIO
    from datetime import date

    trainees = Trainee.objects.select_related('supervisor__user').order_by('-start_date', 'full_name')
    today = date.today()
    buf = BytesIO()

    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)

    styles = getSampleStyleSheet()
    title_s = ParagraphStyle('t', fontSize=14, fontName='Helvetica-Bold', spaceAfter=4)
    sub_s   = ParagraphStyle('s', fontSize=9,  fontName='Helvetica', textColor=colors.grey, spaceAfter=10)
    cell_s  = ParagraphStyle('c', fontSize=7.5, fontName='Helvetica', leading=10)

    story = [
        Paragraph("AEF HRM — Trainee Registry", title_s),
        Paragraph(f"Exported: {today.strftime('%d %B %Y')} | Total: {trainees.count()} trainees", sub_s),
    ]

    headers = ['#', 'Full Name', 'Program', 'Start', 'End', 'Institution', 'Country', 'Supervisor', 'Cert.', 'Status']
    data = [[Paragraph(f'<b>{h}</b>', cell_s) for h in headers]]

    for i, t in enumerate(trainees, 1):
        if today < t.start_date:
            st = 'Upcoming'
        elif today > t.end_date:
            st = 'Completed'
        else:
            st = 'Active'
        data.append([
            Paragraph(str(i), cell_s),
            Paragraph(t.full_name, cell_s),
            Paragraph(t.program, cell_s),
            Paragraph(t.start_date.strftime('%d/%m/%y'), cell_s),
            Paragraph(t.end_date.strftime('%d/%m/%y'), cell_s),
            Paragraph(t.home_institution, cell_s),
            Paragraph(t.country, cell_s),
            Paragraph(t.supervisor.get_full_name() if t.supervisor else '—', cell_s),
            Paragraph('Yes' if t.certificate_issued else 'No', cell_s),
            Paragraph(st, cell_s),
        ])

    col_w = [0.7*cm, 5*cm, 4.5*cm, 2*cm, 2*cm, 5*cm, 3*cm, 4*cm, 1.2*cm, 2.2*cm]
    tbl = Table(data, colWidths=col_w, repeatRows=1)
    tbl.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0A4D68')),
        ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
        ('GRID',       (0,0), (-1,-1), 0.3, colors.HexColor('#cccccc')),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f5f5f5')]),
        ('TOPPADDING',    (0,0), (-1,-1), 3),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3),
        ('LEFTPADDING',   (0,0), (-1,-1), 4),
        ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
    ]))
    story.append(tbl)

    doc.build(story)
    buf.seek(0)
    response = HttpResponse(buf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="trainees_{today}.pdf"'
    return response
