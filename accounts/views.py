from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from django.http import JsonResponse
from django.db.models import Q
from django.urls import reverse
from .models import Employee, Department, HealthDependant
from .forms import LoginForm, EmployeeCreateForm, EmployeeEditForm, DepartmentForm, ChangePasswordForm, AdminResetCredentialsForm, EmployeeSelfEditForm, HealthDependantForm
from .signature_utils import process_signature


def _login_rate_key(request):
    """Cache key for tracking failed login attempts per IP."""
    x_forwarded = request.META.get('HTTP_X_FORWARDED_FOR', '')
    ip = x_forwarded.split(',')[0].strip() if x_forwarded else request.META.get('REMOTE_ADDR', '0.0.0.0')
    return f'login_fails_{ip}'


def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard:home')

    # Rate limiting: block IP after 10 failed attempts in 10 minutes
    from django.core.cache import cache
    rate_key = _login_rate_key(request)
    fail_count = cache.get(rate_key, 0)
    if fail_count >= 10:
        messages.error(request, "Too many failed login attempts. Please wait 10 minutes before trying again.")
        return render(request, 'accounts/login.html', {'form': LoginForm(), 'rate_limited': True})

    form = LoginForm(request, data=request.POST or None)
    if request.method == 'POST':
        if form.is_valid():
            user = form.get_user()
            cache.delete(rate_key)  # reset on success
            # Check if 2FA is enabled for this user
            try:
                if user.employee.totp_enabled and user.employee.totp_secret:
                    request.session['2fa_pending_user_id'] = user.pk
                    request.session['2fa_next'] = request.GET.get('next', 'dashboard:home')
                    return redirect('accounts:verify_2fa')
            except Exception:
                pass
            login(request, user)
            return redirect(request.GET.get('next', 'dashboard:home'))
        else:
            # Increment failure counter (expires after 10 minutes)
            cache.set(rate_key, fail_count + 1, timeout=600)

    return render(request, 'accounts/login.html', {'form': form})


def logout_view(request):
    logout(request)
    response = redirect('accounts:login')
    response.delete_cookie('csrftoken')
    return response


@login_required
def profile_view(request):
    from django.forms import modelformset_factory
    try:
        employee = request.user.employee
    except Employee.DoesNotExist:
        messages.error(request, "No employee profile found.")
        return redirect('dashboard:home')

    change_form   = ChangePasswordForm(request.user)
    self_edit_form = EmployeeSelfEditForm(instance=employee)

    if request.method == 'POST' and 'self_edit' in request.POST:
        self_edit_form = EmployeeSelfEditForm(request.POST, instance=employee)
        if self_edit_form.is_valid():
            self_edit_form.save()
            messages.success(request, "Profile updated successfully.")
            return redirect('accounts:profile')

    dependants = employee.health_dependants.all()
    return render(request, 'accounts/profile.html', {
        'employee': employee,
        'change_form': change_form,
        'self_edit_form': self_edit_form,
        'dependants': dependants,
    })


def _is_hr_or_superuser(user):
    """HR Admin, CEO, or Superuser — can manage employees."""
    if user.is_superuser:
        return True
    try:
        return user.employee.is_hr() or user.employee.is_ceo()
    except Exception:
        return False


def hr_required(view_func):
    @login_required
    def wrapper(request, *args, **kwargs):
        try:
            if request.user.employee.role != 'hr':
                messages.error(request, "Access denied. HR Admin only.")
                return redirect('dashboard:home')
        except Employee.DoesNotExist:
            return redirect('dashboard:home')
        return view_func(request, *args, **kwargs)
    return wrapper


def superuser_required(view_func):
    @login_required
    def wrapper(request, *args, **kwargs):
        if not request.user.is_superuser:
            messages.error(request, "Access denied. Admin only.")
            return redirect('dashboard:home')
        return view_func(request, *args, **kwargs)
    return wrapper


def get_employee(request):
    try:
        return request.user.employee
    except Exception:
        return None


def hr_or_superuser_required(view_func):
    """Decorator allowing HR Admin OR Superuser."""
    @login_required
    def wrapper(request, *args, **kwargs):
        if not _is_hr_or_superuser(request.user):
            messages.error(request, "Access denied. HR or Admin only.")
            return redirect('dashboard:home')
        return view_func(request, *args, **kwargs)
    return wrapper


def _generate_username(first_name, last_name=''):
    """Generate username as firstname.lastname (e.g. john.smith).
    If already taken, appends incrementing numbers: john.smith2, john.smith3, …
    Falls back gracefully when only first_name is provided (AJAX suggest endpoint).
    """
    from django.contrib.auth.models import User
    import re
    first = re.sub(r'[^a-z]', '', first_name.lower()) or 'user'
    last = re.sub(r'[^a-z]', '', last_name.lower())
    base = f"{first}.{last}" if last else first
    candidate = base
    counter = 2
    while User.objects.filter(username=candidate).exists():
        candidate = f"{base}{counter}"
        counter += 1
    return candidate


@hr_or_superuser_required
def employee_list(request):
    from django.db.models import Q, Value
    from django.db.models.functions import Concat
    show_former = request.GET.get('show_former', '0') == '1'
    name_q = request.GET.get('q', '').strip()
    dept_q = request.GET.get('dept', '').strip()
    role_q = request.GET.get('role', '').strip()

    employees = Employee.objects.filter(
        is_active=not show_former
    ).select_related('user', 'department', 'supervisor__user').annotate(
        full_name=Concat('user__first_name', Value(' '), 'user__last_name')
    )

    if name_q:
        employees = employees.filter(
            Q(full_name__icontains=name_q) |
            Q(user__first_name__icontains=name_q) |
            Q(user__last_name__icontains=name_q) |
            Q(employee_id__icontains=name_q) |
            Q(position__icontains=name_q)
        )
    if dept_q:
        employees = employees.filter(department_id=dept_q)
    if role_q:
        employees = employees.filter(role=role_q)

    # Alphabetical order by last name, then first name
    employees = employees.order_by('user__last_name', 'user__first_name')

    from accounts.models import Department
    departments = Department.objects.all().order_by('name')
    managers = Employee.objects.filter(
        role__in=('manager', 'unit_head', 'nurse_superintendent', 'hr', 'admin_director', 'finance_director', 'ceo')
    ).select_related('user').order_by('user__last_name', 'user__first_name')

    return render(request, 'accounts/employee_list.html', {
        'employees': employees,
        'show_former': show_former,
        'departments': departments,
        'name_q': name_q,
        'dept_q': dept_q,
        'role_q': role_q,
        'role_choices': Employee.ROLE_CHOICES,
        'managers': managers,
    })


@hr_or_superuser_required
def employee_create(request):
    # Auto-generate username suggestion on GET (or from POST first_name + last_name)
    auto_username = ''
    first_name_hint = request.POST.get('first_name', '')
    last_name_hint = request.POST.get('last_name', '')
    if first_name_hint:
        auto_username = _generate_username(first_name_hint, last_name_hint)

    form = EmployeeCreateForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        try:
            with transaction.atomic():
                employee = form.save()

                # Auto-set role based on contract type if role wasn't explicitly changed
                ct = form.cleaned_data.get('contract_type', '')
                if ct == 'INTERN' and employee.role == 'employee':
                    employee.role = 'intern'
                    employee.save(update_fields=['role'])
                elif ct == 'WACS' and employee.role == 'employee':
                    employee.role = 'wacs_resident'
                    employee.save(update_fields=['role'])

                from leaves.models import LeaveBalance
                from datetime import date
                LeaveBalance.objects.get_or_create(
                    employee=employee,
                    year=date.today().year,
                    defaults={'total_entitlement': 18}
                )

                # Create initial contract from registration data
                from contracts.models import Contract
                contract = Contract.objects.create(
                    employee=employee,
                    contract_type=form.cleaned_data['contract_type'],
                    contract_number=form.cleaned_data.get('contract_number', ''),
                    start_date=form.cleaned_data['contract_start_date'],
                    end_date=form.cleaned_data.get('contract_end_date') or None,
                    status='active',
                    created_by=request.user,
                    notes='Issued at registration.',
                )

                # Welcome notification to the new employee
                from notifications.utils import notify
                from contracts.views import _contract_issue_message
                from django.conf import settings
                _ctitle, _cmsg = _contract_issue_message(contract)
                ct = contract.contract_type
                if ct == 'INTERN':
                    welcome_intro = f"Welcome to AEF HRM, {employee.get_full_name()}! Your internship account has been created and activated."
                elif ct == 'WACS':
                    welcome_intro = f"Welcome to AEF HRM, {employee.get_full_name()}! Your WACS Residency account has been created and activated."
                else:
                    welcome_intro = f"Welcome to AEF HRM, {employee.get_full_name()}! Your account has been created and activated."
                site_base = getattr(settings, 'SITE_URL', '').rstrip('/')
                login_url = f"{site_base}/accounts/login/" if site_base else "/accounts/login/"
                _plain_password = form.cleaned_data['password']
                credentials_msg = (
                    f"\n\nYour login credentials:\n"
                    f"Username: {employee.user.username}\n"
                    f"Password: {_plain_password}\n"
                    f"Login at: {login_url}\n\n"
                    f"Please change your password after your first login."
                )
                notify(
                    employee.user,
                    title='Welcome — Your Account Is Now Active',
                    message=f"{welcome_intro}{credentials_msg}{_cmsg}",
                    notification_type='account_activated',
                    url='/contracts/my/',
                )

            messages.success(request, f"Employee {employee.get_full_name()} created and contract issued successfully.")
            try:
                from dashboard.models import AuditLog
                AuditLog.log(
                    request, AuditLog.ACTION_EMPLOYEE,
                    f"Created new employee profile: {employee.get_full_name()} ({employee.employee_id}, {employee.role})",
                    target_user=employee.user,
                )
            except Exception:
                pass
            return redirect('accounts:employee_list')
        except Exception as e:
            messages.error(request, f"Error creating employee: {e}")
    return render(request, 'accounts/employee_form.html', {
        'form': form,
        'title': 'Add New Employee',
        'auto_username': auto_username,
    })


@hr_or_superuser_required
def employee_edit(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    form = EmployeeEditForm(request.POST or None, instance=employee)
    if request.method == 'POST' and form.is_valid():
        emp_obj = form.save()
        # Handle signature upload separately from the form
        sig_file = request.FILES.get('signature')
        if sig_file:
            from django.core.files.base import ContentFile
            import os
            processed = process_signature(sig_file)
            if processed:
                # Delete old signature file if exists
                if emp_obj.signature:
                    try:
                        emp_obj.signature.delete(save=False)
                    except Exception:
                        pass
                fname = os.path.splitext(sig_file.name)[0] + '_sig.png'
                raw_bytes = processed.read()
                emp_obj.signature.save(fname, ContentFile(raw_bytes), save=False)
                import base64 as _b64
                emp_obj.signature_b64 = 'data:image/png;base64,' + _b64.b64encode(raw_bytes).decode('utf-8')
                emp_obj.save(update_fields=['signature', 'signature_b64'])
        messages.success(request, "Employee updated successfully.")
        return redirect('accounts:employee_list')

    from contracts.models import Contract
    active_contract = employee.contracts.filter(status='active').first()
    contract_history = employee.contracts.exclude(pk=active_contract.pk).order_by('-start_date') if active_contract else employee.contracts.order_by('-start_date')

    return render(request, 'accounts/employee_form.html', {
        'form': form,
        'title': 'Edit Employee',
        'employee': employee,
        'active_contract': active_contract,
        'contract_history': contract_history,
    })


@hr_or_superuser_required
def employee_delete(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    if request.method == 'POST':
        name = employee.get_full_name()
        employee.user.delete()  # cascades to employee record
        messages.success(request, f"Employee {name} deleted successfully.")
        return redirect('accounts:employee_list')
    return render(request, 'accounts/employee_confirm_delete.html', {'employee': employee})


@login_required
def toggle_account_lock(request, pk):
    """HR or superuser: lock/unlock an employee's login account."""
    if not _is_hr_or_superuser(request.user):
        messages.error(request, "Access denied.")
        return redirect('dashboard:home')

    employee = get_object_or_404(Employee, pk=pk)
    if request.method == 'POST':
        user = employee.user
        user.is_active = not user.is_active
        user.save(update_fields=['is_active'])
        state = "unlocked" if user.is_active else "locked"
        messages.success(request, f"Account for {employee.get_full_name()} has been {state}.")
        try:
            from dashboard.models import AuditLog
            AuditLog.log(
                request, AuditLog.ACTION_EMPLOYEE,
                f"Account {state} for {employee.get_full_name()} ({employee.employee_id})",
                target_user=user,
            )
        except Exception:
            pass
    return redirect(request.POST.get('next', 'accounts:employee_list'))


@superuser_required
def department_list(request):
    departments = Department.objects.all()
    form = DepartmentForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, "Department added.")
        return redirect('accounts:department_list')
    return render(request, 'accounts/department_list.html', {'departments': departments, 'form': form})


@superuser_required
def department_delete(request, pk):
    dept = get_object_or_404(Department, pk=pk)
    if request.method == 'POST':
        dept.delete()
        messages.success(request, "Department deleted.")
    return redirect('accounts:department_list')


@login_required
def set_employee_signature(request, pk):
    """Superuser-only: draw and save a signature on behalf of a key approver."""
    if not request.user.is_superuser:
        messages.error(request, "Access denied.")
        return redirect('accounts:employee_list')
    if request.method != 'POST':
        return redirect('accounts:employee_edit', pk=pk)

    employee = get_object_or_404(Employee, pk=pk)
    b64_data = request.POST.get('signature_data', '')

    if not b64_data or not b64_data.startswith('data:image/png;base64,'):
        messages.error(request, "No signature drawn.")
        return redirect('accounts:employee_edit', pk=pk)

    import base64
    from django.core.files.base import ContentFile
    try:
        raw = base64.b64decode(b64_data.split(',', 1)[1])
        if employee.signature:
            try:
                employee.signature.delete(save=False)
            except Exception:
                pass
        fname = f"{employee.employee_id}_sig.png"
        employee.signature.save(fname, ContentFile(raw), save=False)
        employee.signature_b64 = b64_data
        employee.save(update_fields=['signature', 'signature_b64'])
        messages.success(request, f"Signature saved for {employee.get_full_name()}.")
    except Exception:
        messages.error(request, "Could not save signature. Please try again.")
    return redirect('accounts:employee_edit', pk=pk)


@login_required
def profile_save_signature(request):
    """Any logged-in employee can save/update their own signature from the profile page."""
    from django.http import JsonResponse as _JsonResponse
    if request.method != 'POST':
        return _JsonResponse({'error': 'Method not allowed'}, status=405)

    try:
        employee = request.user.employee
    except Exception:
        return _JsonResponse({'error': 'Employee profile not found'}, status=400)

    b64_data = request.POST.get('signature_data', '')
    if not b64_data or not b64_data.startswith('data:image/png;base64,'):
        return _JsonResponse({'error': 'No signature drawn'}, status=400)

    import base64
    from django.core.files.base import ContentFile
    employee.signature_b64 = b64_data
    save_fields = ['signature_b64']
    try:
        raw = base64.b64decode(b64_data.split(',', 1)[1])
        if employee.signature:
            try:
                employee.signature.delete(save=False)
            except Exception:
                pass
        fname = f"{employee.employee_id}_sig.png"
        employee.signature.save(fname, ContentFile(raw), save=False)
        save_fields.append('signature')
    except Exception:
        pass
    employee.save(update_fields=save_fields)
    return _JsonResponse({'ok': True})


def _mask_email(email):
    """Return a masked email: john.doe@company.com → j***e@company.com"""
    try:
        local, domain = email.split('@', 1)
        if len(local) <= 2:
            masked_local = local[0] + '***'
        elif len(local) <= 4:
            masked_local = local[0] + '***' + local[-1]
        else:
            masked_local = local[0] + local[1] + '***' + local[-2] + local[-1]
        return f"{masked_local}@{domain}"
    except Exception:
        return '***@***.***'


def password_reset_view(request):
    """
    2-step secure password reset:
      Step 1 — enter username
      Step 2 — system shows masked email; user must type the full email to confirm
    Wrong email at step 2 → hard fail.
    """
    from django.contrib.auth import get_user_model
    from django.contrib.auth.forms import PasswordResetForm

    User = get_user_model()
    step = request.POST.get('step', '1')

    if request.method == 'POST':
        if step == '1':
            username = request.POST.get('username', '').strip()
            try:
                user = User.objects.get(username=username)
                if not user.email:
                    return render(request, 'accounts/password_reset_form.html', {
                        'step': '1',
                        'error': 'This account has no email address on file. Please contact HR.',
                    })
                return render(request, 'accounts/password_reset_form.html', {
                    'step': '2',
                    'username': username,
                    'masked_email': _mask_email(user.email),
                })
            except User.DoesNotExist:
                return render(request, 'accounts/password_reset_form.html', {
                    'step': '1',
                    'error': 'No account found with that username.',
                })

        elif step == '2':
            username = request.POST.get('username', '').strip()
            email_input = request.POST.get('email', '').strip().lower()
            try:
                user = User.objects.get(username=username)
                if user.email.lower() != email_input:
                    return render(request, 'accounts/password_reset_form.html', {
                        'step': '2',
                        'username': username,
                        'masked_email': _mask_email(user.email),
                        'error': 'The email you entered does not match the email on this account. Please try again.',
                    })
                # Email confirmed — send the reset link via Django's built-in form
                form = PasswordResetForm({'email': email_input})
                if form.is_valid():
                    form.save(
                        request=request,
                        use_https=request.is_secure(),
                        email_template_name='accounts/password_reset_email.html',
                        subject_template_name='accounts/password_reset_subject.txt',
                    )
                return redirect('accounts:password_reset_done')
            except User.DoesNotExist:
                return render(request, 'accounts/password_reset_form.html', {
                    'step': '1',
                    'error': 'No account found with that username.',
                })

    return render(request, 'accounts/password_reset_form.html', {'step': '1'})


@login_required
def change_password(request):
    """Allows any logged-in user to change their own password."""
    from django.contrib.auth import update_session_auth_hash
    form = ChangePasswordForm(request.user, request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        update_session_auth_hash(request, request.user)  # keep user logged in
        messages.success(request, "Password changed successfully.")
        return redirect('accounts:profile')
    return render(request, 'accounts/change_password.html', {'form': form})


@superuser_required
def admin_reset_credentials(request, pk):
    """Allows superuser to reset any employee's username and/or password."""
    employee = get_object_or_404(Employee, pk=pk)
    form = AdminResetCredentialsForm(employee.user, request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, f"Credentials updated for {employee.get_full_name()}.")
        return redirect('accounts:employee_list')
    return render(request, 'accounts/admin_reset_credentials.html', {
        'form': form,
        'employee': employee,
    })


@login_required
def employee_history(request, pk):
    """Comprehensive employee record: profile, contracts, leaves, discipline.
    Access: HR, Admin Director, Finance Director, CEO, Superuser only.
    """
    viewer = None
    try:
        viewer = request.user.employee
    except Employee.DoesNotExist:
        pass

    is_privileged = (
        request.user.is_superuser
        or (viewer and (viewer.is_hr() or viewer.is_director() or viewer.is_ceo()))
    )
    if not is_privileged:
        messages.error(request, "Access denied.")
        return redirect('dashboard:home')

    employee = get_object_or_404(
        Employee.objects.select_related('user', 'department', 'supervisor__user'),
        pk=pk
    )

    from contracts.models import Contract
    from leaves.models import LeaveRequest, LeaveBalance
    from discipline.models import DisciplineRecord

    contracts = Contract.objects.filter(employee=employee).order_by('start_date', 'created_at')
    leave_requests = LeaveRequest.objects.filter(employee=employee).select_related(
        'leave_type', 'manager_action_by__user', 'hr_action_by__user'
    ).order_by('-created_at')
    discipline_records = DisciplineRecord.objects.filter(employee=employee).select_related(
        'issued_by'
    ).order_by('-date_issued')

    from datetime import date
    today = date.today()
    balance = LeaveBalance.objects.filter(employee=employee, year=today.year).first()

    from accounts.models import EmployeeDocument
    documents = EmployeeDocument.objects.filter(employee=employee).select_related("uploaded_by")

    from appraisals.models import AppraisalRecord
    appraisals = AppraisalRecord.objects.filter(employee=employee).select_related('cycle').order_by('-cycle__year', '-cycle__trimester')

    dependants = employee.health_dependants.all()

    return render(request, 'accounts/employee_history.html', {
        'employee': employee,
        'contracts': contracts,
        'leave_requests': leave_requests,
        'discipline_records': discipline_records,
        'balance': balance,
        'today': today,
        'documents': documents,
        'is_privileged': is_privileged,
        'appraisals': appraisals,
        'dependants': dependants,
    })


@login_required
def username_suggest(request):
    """AJAX endpoint: suggest next available username for a given first_name + last_name."""
    from django.http import JsonResponse
    first_name = request.GET.get('first_name', '').strip()
    last_name = request.GET.get('last_name', '').strip()
    if not first_name:
        return JsonResponse({'username': ''})
    username = _generate_username(first_name, last_name)
    return JsonResponse({'username': username})


@hr_or_superuser_required
def bulk_assign_manager(request):
    """Assign a line manager or unit head to multiple selected employees at once."""
    if request.method != 'POST':
        return redirect('accounts:employee_list')

    employee_ids = request.POST.getlist('employee_ids')
    assignment_type = request.POST.get('assignment_type')
    manager_id = request.POST.get('manager_id')

    if not employee_ids:
        messages.error(request, "No employees selected.")
        return redirect('accounts:employee_list')
    if not manager_id:
        messages.error(request, "Please select a manager to assign.")
        return redirect('accounts:employee_list')

    try:
        manager = Employee.objects.get(pk=manager_id)
    except Employee.DoesNotExist:
        messages.error(request, "Selected manager/unit head not found.")
        return redirect('accounts:employee_list')

    qs = Employee.objects.filter(pk__in=employee_ids)
    count = qs.count()

    if assignment_type == 'department':
        try:
            dept = Department.objects.get(pk=manager_id)
        except Department.DoesNotExist:
            messages.error(request, "Selected department not found.")
            return redirect('accounts:employee_list')
        qs.update(department=dept)
        messages.success(request, f"Department set to {dept.name} for {count} employee(s).")
    elif assignment_type == 'unit_head':
        qs.update(unit_head=manager)
        messages.success(request, f"Unit Head set to {manager.get_full_name()} for {count} employee(s).")
    elif assignment_type == 'nurse_superintendent':
        qs.update(nurse_superintendent=manager, requires_nurse_supt=True)
        messages.success(request, f"Nurse Superintendent set to {manager.get_full_name()} for {count} employee(s).")
    else:
        qs.update(supervisor=manager)
        messages.success(request, f"Line Manager set to {manager.get_full_name()} for {count} employee(s).")

    return redirect('accounts:employee_list')


@login_required
def employee_import(request):
    if not request.user.is_superuser:
        messages.error(request, "Access denied. System Admin only.")
        return redirect('dashboard:home')
    """Bulk-import employees from an Excel (.xlsx) file."""
    from django.http import HttpResponse
    import openpyxl
    from contracts.models import Contract
    from datetime import date as _date, datetime as _datetime

    COLUMN_MAP = [
        'first_name', 'last_name', 'email', 'employee_id', 'department',
        'role', 'position', 'phone', 'date_of_birth', 'date_joined_company',
        'sex', 'nationality', 'qualifications', 'staff_category',
        'contract_number', 'contract_type', 'contract_start_date', 'contract_end_date',
    ]

    if request.method == 'GET' and request.GET.get('template') == '1':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Employees'
        for col_idx, h in enumerate(COLUMN_MAP, 1):
            ws.cell(row=1, column=col_idx, value=h)
        example = [
            'Jane', 'Doe', 'jane.doe@example.com', 'EMP001', 'Nursing',
            'employee', 'Staff Nurse', '+237600000000', '1990-05-20', '2020-01-15',
            'F', 'Cameroonian', 'BSc Nursing', 'A',
            'CTR-2024-001', 'CDI', '2020-01-15', '',
        ]
        for col_idx, val in enumerate(example, 1):
            ws.cell(row=2, column=col_idx, value=val)
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="employee_import_template.xlsx"'
        wb.save(response)
        return response

    results = []
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, "Please select an Excel file to upload.")
            return redirect('accounts:employee_import')
        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            ws = wb.active
        except Exception as e:
            messages.error(request, f"Could not read Excel file: {e}")
            return redirect('accounts:employee_import')

        headers_raw = [
            str(ws.cell(row=1, column=c).value or '').strip().lower().replace(' ', '_')
            for c in range(1, ws.max_column + 1)
        ]

        def _get(row_vals, key):
            return str(row_vals.get(key, '') or '').strip()

        def _parse_date(val):
            if not val:
                return None
            for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y'):
                try:
                    return _datetime.strptime(val, fmt).date()
                except ValueError:
                    pass
            return None

        created_count = 0
        error_count = 0

        for row_idx in range(2, ws.max_row + 1):
            row_vals = {}
            for col_idx, h in enumerate(headers_raw, 1):
                row_vals[h] = ws.cell(row=row_idx, column=col_idx).value

            if not any(v for v in row_vals.values() if v is not None):
                continue

            first_name = _get(row_vals, 'first_name')
            last_name  = _get(row_vals, 'last_name')
            emp_id     = _get(row_vals, 'employee_id')

            if not (first_name and last_name and emp_id):
                results.append({'row': row_idx, 'status': 'error',
                    'name': f"{first_name} {last_name}".strip() or f"Row {row_idx}",
                    'message': "Missing first_name, last_name or employee_id."})
                error_count += 1
                continue

            if Employee.objects.filter(employee_id=emp_id).exists():
                results.append({'row': row_idx, 'status': 'skipped',
                    'name': f"{first_name} {last_name}",
                    'message': f"Employee ID {emp_id} already exists."})
                error_count += 1
                continue

            try:
                with transaction.atomic():
                    from django.contrib.auth.models import User as _User

                    dept_name = _get(row_vals, 'department')
                    dept = Department.objects.filter(name__iexact=dept_name).first() if dept_name else None

                    role = _get(row_vals, 'role') or 'employee'
                    ct   = (_get(row_vals, 'contract_type') or 'CDI').upper()
                    if ct not in ('CDI', 'CDD', 'INTERN', 'WACS'):
                        ct = 'CDI'
                    if ct == 'INTERN' and role == 'employee':
                        role = 'intern'
                    elif ct == 'WACS' and role == 'employee':
                        role = 'wacs_resident'

                    username = _generate_username(first_name)
                    email    = _get(row_vals, 'email') or f"{username}@hospital.local"

                    user = _User.objects.create_user(
                        username=username, first_name=first_name,
                        last_name=last_name, email=email, password='Micei2021',
                    )
                    employee = Employee.objects.create(
                        user=user, employee_id=emp_id, department=dept, role=role,
                        position=_get(row_vals, 'position'),
                        phone=_get(row_vals, 'phone'),
                        date_of_birth=_parse_date(_get(row_vals, 'date_of_birth')),
                        date_joined_company=_parse_date(_get(row_vals, 'date_joined_company')),
                        sex=_get(row_vals, 'sex'),
                        nationality=_get(row_vals, 'nationality'),
                        qualifications=_get(row_vals, 'qualifications'),
                        staff_category=_get(row_vals, 'staff_category'),
                        contract_number=_get(row_vals, 'contract_number'),
                    )

                    from leaves.models import LeaveBalance
                    LeaveBalance.objects.get_or_create(
                        employee=employee, year=_date.today().year,
                        defaults={'total_entitlement': 18},
                    )
                    Contract.objects.create(
                        employee=employee, contract_type=ct,
                        contract_number=_get(row_vals, 'contract_number'),
                        start_date=_parse_date(_get(row_vals, 'contract_start_date')) or _date.today(),
                        end_date=_parse_date(_get(row_vals, 'contract_end_date')),
                        status='active', created_by=request.user,
                        notes='Imported via Excel.',
                    )

                results.append({'row': row_idx, 'status': 'created',
                    'name': f"{first_name} {last_name}",
                    'message': f"Created (username: {username})."})
                created_count += 1

            except Exception as exc:
                results.append({'row': row_idx, 'status': 'error',
                    'name': f"{first_name} {last_name}",
                    'message': str(exc)})
                error_count += 1

        messages.success(request, f"Import complete: {created_count} created, {error_count} errors/skipped.")

    return render(request, 'accounts/employee_import.html', {
        'results': results,
        'column_map': COLUMN_MAP,
    })


# ─────────────────────────────────────────────────────────────
#  EXCEL BULK UPLOAD
# ─────────────────────────────────────────────────────────────

# Column name → (Employee field, type)
# Required columns are marked True; optional are False.
EXCEL_COLUMNS = {
    # Personal
    'first_name':         ('first_name',          'str',  True),
    'last_name':          ('last_name',           'str',  True),
    'email':              ('email',               'str',  True),
    'employee_id':        ('employee_id',         'str',  True),
    'date_of_birth':      ('date_of_birth',       'date', False),
    'sex':                ('sex',                 'str',  False),
    'nationality':        ('nationality',         'str',  False),
    'phone':              ('phone',               'str',  False),
    # Employment
    'position':           ('position',            'str',  False),
    'department':         ('department',          'dept', False),
    'role':               ('role',               'str',  False),
    'staff_category':     ('staff_category',     'str',  False),
    'date_joined_hospital': ('date_joined_company', 'date', False),
    'qualifications':     ('qualifications',      'str',  False),
    'contract_number':    ('contract_number',     'str',  False),
    # Contract (REQUIRED at creation)
    'contract_type':      ('contract_type',       'str',  True),
    'contract_start_date': ('contract_start_date', 'date', True),
    'contract_end_date':  ('contract_end_date',   'date', False),
}

VALID_ROLES = {v for v, _ in Employee.ROLE_CHOICES}
VALID_CONTRACT_TYPES = {'CDI', 'CDD', 'INTERN', 'WACS'}
DEFAULT_PASSWORD = 'Micei2021'


def _parse_cell(value, cell_type):
    """Parse a raw Excel cell value to the target Python type."""
    from datetime import date as _date
    import datetime as _dt

    if value is None or str(value).strip() == '':
        return None

    if cell_type == 'str':
        return str(value).strip()
    if cell_type == 'date':
        if isinstance(value, (_date, _dt.datetime)):
            return value.date() if hasattr(value, 'date') else value
        # Try ISO string
        try:
            return _dt.date.fromisoformat(str(value).strip()[:10])
        except ValueError:
            return None
    if cell_type == 'dept':
        return str(value).strip()
    return value


def _normalize_col(val):
    """Normalise a column header: lowercase, collapse non-alphanumeric runs to underscore."""
    import re
    return re.sub(r'[^a-z0-9]+', '_', str(val or '').strip().lower()).strip('_')


# Maps normalised column names from any source (staff list, custom template) to internal keys.
_COL_ALIAS = {
    'matricule':            'employee_id',
    'employee_id':          'employee_id',
    'name':                 'full_name',
    'full_name':            'full_name',
    'first_name':           'first_name',
    'last_name':            'last_name',
    'category':             'staff_category',
    'staff_category':       'staff_category',
    'date_of_birth':        'date_of_birth',
    'sex':                  'sex',
    'nat':                  'nationality',
    'nationality':          'nationality',
    'date_employment':      'date_joined_company',
    'date_joined_hospital': 'date_joined_company',
    'date_joined_company':  'date_joined_company',
    'contrat_n':            'contract_number',
    'contract_n':           'contract_number',
    'contract_number':      'contract_number',
    'status':               'contract_type_raw',
    'contract_type':        'contract_type_raw',
    'position':             'position',
    'department':           'department',
    'qualifications':       'qualifications',
    'certifications':       'certifications',
    'email':                'email',
    'phone':                'phone',
    'contract_start_date':  'contract_start_date',
    'contract_end_date':    'contract_end_date',
    'role':                 'role',
}

_RECOGNIZED = set(_COL_ALIAS.keys())


def _map_contract_type(raw):
    """Map any Status/contract_type string to a valid contract type code."""
    raw = str(raw or '').strip().upper()
    for code in ('CDI', 'CDD', 'INTERN', 'WACS'):
        if code in raw:
            return code
    return 'CDI'  # default


@hr_or_superuser_required
def employee_excel_upload(request):
    """
    GET  → show upload form / template download
    POST → parse Excel (supports the actual staff list format), create employees
    """
    from datetime import date as _today_dt

    if request.method == 'GET':
        if request.GET.get('download_template') == '1':
            return _excel_template_download()
        return render(request, 'accounts/employee_excel_upload.html', {
            'column_info': EXCEL_COLUMNS,
        })

    # ── POST ────────────────────────────────────────────────
    uploaded = request.FILES.get('excel_file')
    if not uploaded:
        messages.error(request, "No file uploaded.")
        return redirect('accounts:employee_excel_upload')

    try:
        import openpyxl
        wb = openpyxl.load_workbook(uploaded, data_only=True)
        ws = wb.active
    except Exception as e:
        messages.error(request, f"Could not open Excel file: {e}")
        return redirect('accounts:employee_excel_upload')

    # ── Detect header row (row 1 may be a title row) ────────
    def _read_headers(row_idx):
        return [_normalize_col(ws.cell(row=row_idx, column=c).value)
                for c in range(1, ws.max_column + 1)]

    headers_r1 = _read_headers(1)
    headers_r2 = _read_headers(2)
    recognized_r1 = sum(1 for h in headers_r1 if h in _RECOGNIZED)
    recognized_r2 = sum(1 for h in headers_r2 if h in _RECOGNIZED)

    if recognized_r2 > recognized_r1:
        # Row 1 is a title; real headers are on row 2
        headers = headers_r2
        data_start_row = 3
    else:
        headers = headers_r1
        data_start_row = 2

    # Map positional headers → internal keys
    mapped_headers = [_COL_ALIAS.get(h, h) for h in headers]

    created_rows = []
    skipped_rows = []
    today = _today_dt.today()

    for row_idx, row in enumerate(ws.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row):
        if all(v is None or str(v).strip() == '' for v in row):
            continue  # blank

        row_data = {mapped_headers[i]: row[i] for i in range(min(len(mapped_headers), len(row)))}
        errors = []

        # ── Name resolution ──────────────────────────────────
        if row_data.get('full_name'):
            full = _parse_cell(row_data['full_name'], 'str') or ''
            parts = full.split(None, 1)
            last_name  = parts[0] if parts else ''
            first_name = parts[1] if len(parts) > 1 else last_name
        else:
            first_name = _parse_cell(row_data.get('first_name'), 'str') or ''
            last_name  = _parse_cell(row_data.get('last_name'),  'str') or ''

        emp_id = _parse_cell(row_data.get('employee_id'), 'str') or ''

        if not first_name:
            errors.append("Missing name / first_name")
        if not emp_id:
            errors.append("Missing Matricule / employee_id")

        # ── Contract type ───────────────────────────────────
        contract_type = _map_contract_type(
            _parse_cell(row_data.get('contract_type_raw'), 'str') or ''
        )

        # ── Dates ───────────────────────────────────────────
        date_joined = _parse_cell(row_data.get('date_joined_company'), 'date')
        contract_start = _parse_cell(row_data.get('contract_start_date'), 'date') or date_joined
        contract_end   = _parse_cell(row_data.get('contract_end_date'), 'date')
        if not contract_start:
            contract_start = today  # last resort default

        # ── Department ──────────────────────────────────────
        dept_name = _parse_cell(row_data.get('department'), 'dept')
        dept_obj = None
        if dept_name:
            dept_obj = Department.objects.filter(name__iexact=dept_name).first()
            if not dept_obj:
                errors.append(f"Department '{dept_name}' not found")

        # ── Role ────────────────────────────────────────────
        role = _parse_cell(row_data.get('role'), 'str') or ''
        if not role:
            if contract_type == 'INTERN':
                role = 'intern'
            elif contract_type == 'WACS':
                role = 'wacs_resident'
            else:
                role = 'employee'
        if role not in VALID_ROLES:
            role = 'employee'

        # ── Duplicate check ─────────────────────────────────
        if emp_id and Employee.objects.filter(employee_id=emp_id).exists():
            errors.append(f"Employee ID '{emp_id}' already exists")

        # ── Qualifications (merge qualifications + certifications) ──
        qualif = _parse_cell(row_data.get('qualifications'), 'str') or ''
        certif = _parse_cell(row_data.get('certifications'), 'str') or ''
        combined_qualif = ' | '.join(filter(None, [qualif, certif]))

        if errors:
            skipped_rows.append({'row': row_idx, 'name': f"{first_name} {last_name}".strip(), 'errors': errors})
            continue

        # ── Create user + employee + contract ───────────────
        try:
            with transaction.atomic():
                from django.contrib.auth.models import User as _User
                username = _generate_username(first_name, last_name)
                email = _parse_cell(row_data.get('email'), 'str') or ''
                if not email:
                    email = f"{username}@hospital.local"

                user = _User(username=username, first_name=first_name,
                             last_name=last_name, email=email)
                user.set_password(DEFAULT_PASSWORD)
                user.save()

                emp = Employee(
                    user=user,
                    employee_id=emp_id,
                    department=dept_obj,
                    role=role,
                    staff_category=_parse_cell(row_data.get('staff_category'), 'str') or '',
                    position=_parse_cell(row_data.get('position'), 'str') or '',
                    phone=_parse_cell(row_data.get('phone'), 'str') or '',
                    date_of_birth=_parse_cell(row_data.get('date_of_birth'), 'date'),
                    date_joined_company=date_joined,
                    sex=(_parse_cell(row_data.get('sex'), 'str') or '')[:1].upper(),
                    nationality=_parse_cell(row_data.get('nationality'), 'str') or '',
                    contract_number=_parse_cell(row_data.get('contract_number'), 'str') or '',
                    qualifications=combined_qualif,
                )
                emp.save()

                from leaves.models import LeaveBalance
                LeaveBalance.objects.get_or_create(
                    employee=emp, year=today.year,
                    defaults={'total_entitlement': 18},
                )

                from contracts.models import Contract
                Contract.objects.create(
                    employee=emp,
                    contract_type=contract_type,
                    start_date=contract_start,
                    end_date=contract_end,
                    status='active',
                    created_by=request.user,
                    notes='Imported via Excel upload.',
                )

                from notifications.utils import notify
                from django.conf import settings as _settings
                _site_base = getattr(_settings, 'SITE_URL', '').rstrip('/')
                _login_url = f"{_site_base}/accounts/login/" if _site_base else "/accounts/login/"
                notify(
                    user,
                    title='Welcome — Your Account Is Now Active',
                    message=(
                        f"Welcome to AEF HRM, {emp.get_full_name()}! Your account has been created. "
                        f"A {contract_type} contract starting {contract_start.strftime('%d %b %Y')} has been issued.\n\n"
                        f"Username: {username}\nPassword: {DEFAULT_PASSWORD}\nLogin at: {_login_url}\n\n"
                        f"Please change your password after first login."
                    ),
                    notification_type='account_activated',
                    url='/contracts/my/',
                )

            created_rows.append({'row': row_idx, 'name': emp.get_full_name(), 'emp_id': emp_id, 'contract': contract_type})
        except Exception as exc:
            skipped_rows.append({'row': row_idx, 'name': f"{first_name} {last_name}".strip(), 'errors': [str(exc)]})

    return render(request, 'accounts/employee_excel_results.html', {
        'created': created_rows,
        'skipped': skipped_rows,
        'total_processed': len(created_rows) + len(skipped_rows),
    })


def _excel_template_download():
    """Return a .xlsx template that mirrors the staff list column layout."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employees"

    # Columns match the actual staff list format
    headers = [
        'Matricule', 'Name', 'Category', 'Date of Birth', 'Sex',
        'Nat.', 'Date Employment', 'Contrat N°', 'Status',
        'Position', 'Department', 'Qualifications', 'Certifications',
        'Email', 'Contract End Date',
    ]
    required_cols = {'Matricule', 'Name', 'Status'}

    header_fill   = PatternFill(start_color="0A4D68", end_color="0A4D68", fill_type="solid")
    required_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
    header_font   = Font(color="FFFFFF", bold=True)

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.fill = required_fill if h in required_cols else header_fill
        ws.column_dimensions[cell.column_letter].width = max(18, len(h) + 4)

    # Sample row
    ws.append([
        'AEF001', 'Doe Jane', '12B-J', '1990-05-15', 'F',
        'CMR', '2020-01-15', 'AEF00120200115', 'CDI',
        'Staff Nurse', 'Nursing', 'BSc Nursing', 'WACS',
        'jane.doe@hospital.cm', '',
    ])

    # Notes sheet
    notes_ws = wb.create_sheet("Notes")
    notes_ws['A1'] = "Field Notes"
    notes_ws['A1'].font = Font(bold=True)
    notes = [
        ('Matricule',         'Unique employee ID (required)'),
        ('Name',              'Full name — Family name first, e.g. "Doe Jane". Required.'),
        ('Category',          'Staff category — any format (e.g. 5, 12B-J, A, III)'),
        ('Date of Birth',     'Format: YYYY-MM-DD or DD/MM/YYYY'),
        ('Sex',               'M or F'),
        ('Nat.',              'Nationality (e.g. CMR, Cameroonian)'),
        ('Date Employment',   'Employment start date — used as contract start if Contract Start Date is absent'),
        ('Contrat N°',        'Official contract/personnel number'),
        ('Status',            'Contract type: CDI | CDD | INTERN | WACS  (default: CDI)'),
        ('Position',          'Job title / position'),
        ('Department',        'Must match an existing department name'),
        ('Qualifications',    'Academic qualifications'),
        ('Certifications',    'Professional certifications (merged with Qualifications)'),
        ('Email',             'Work email — auto-generated if blank'),
        ('Contract End Date', 'Required for CDD, INTERN, WACS. Leave blank for CDI.'),
    ]
    for r, (field, note) in enumerate(notes, start=3):
        notes_ws.cell(row=r, column=1, value=field).font = Font(bold=True)
        notes_ws.cell(row=r, column=2, value=note)
    notes_ws.column_dimensions['A'].width = 22
    notes_ws.column_dimensions['B'].width = 65

    buf = __import__('io').BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="staff_import_template.xlsx"'
    return response



@hr_or_superuser_required
def export_credentials(request):
    """Download a CSV of all employee usernames (passwords cannot be recovered - shows default)."""
    import csv
    from django.http import HttpResponse
    from datetime import date as _date

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="staff_credentials_' + str(_date.today()) + '.csv"'

    writer = csv.writer(response)
    writer.writerow(['Employee ID', 'Full Name', 'Username', 'Email', 'Role', 'Department', 'Default Password'])

    for emp in Employee.objects.select_related('user', 'department').order_by('user__last_name', 'user__first_name'):
        writer.writerow([
            emp.employee_id,
            emp.get_full_name(),
            emp.user.username,
            emp.user.email,
            emp.get_role_display(),
            emp.department.name if emp.department else '',
            'Micei2021 (default - may have been changed by employee)',
        ])
    return response


@login_required
def document_upload(request, employee_pk):
    """HR/superuser uploads a document to an employee's file."""
    from accounts.models import EmployeeDocument
    employee = get_object_or_404(Employee, pk=employee_pk)
    viewer = get_employee(request)
    is_privileged = (
        request.user.is_superuser
        or (viewer and (viewer.is_hr() or viewer.is_director() or viewer.is_ceo()))
    )
    if not is_privileged:
        return redirect('dashboard:home')
    if request.method == 'POST':
        from leave_system.file_utils import validate_upload
        title = request.POST.get('title', '').strip()
        category = request.POST.get('category', 'other')
        file = request.FILES.get('file')
        expiry_date = request.POST.get('expiry_date') or None
        expiry_note = request.POST.get('expiry_note', '').strip()
        if title and file:
            ok, err = validate_upload(file)
            if not ok:
                messages.error(request, err)
            else:
                EmployeeDocument.objects.create(
                    employee=employee,
                    title=title,
                    category=category,
                    file=file,
                    expiry_date=expiry_date,
                    expiry_note=expiry_note,
                    uploaded_by=request.user,
                )
                messages.success(request, f'Document "{title}" uploaded successfully.')
        else:
            messages.error(request, 'Title and file are required.')
    return redirect(reverse('accounts:employee_history', args=[employee_pk]) + '#documents')


@login_required
def my_documents(request):
    """Any employee uploads and views their own documents."""
    from accounts.models import EmployeeDocument
    try:
        employee = request.user.employee
    except Employee.DoesNotExist:
        messages.error(request, "No employee profile found.")
        return redirect('dashboard:home')

    if request.method == 'POST':
        from leave_system.file_utils import validate_upload
        title = request.POST.get('title', '').strip()
        category = request.POST.get('category', 'other')
        file = request.FILES.get('file')
        expiry_date = request.POST.get('expiry_date') or None
        expiry_note = request.POST.get('expiry_note', '').strip()
        if title and file:
            ok, err = validate_upload(file)
            if not ok:
                messages.error(request, err)
            else:
                EmployeeDocument.objects.create(
                    employee=employee,
                    title=title,
                    category=category,
                    file=file,
                    expiry_date=expiry_date,
                    expiry_note=expiry_note,
                    uploaded_by=request.user,
                )
                messages.success(request, f'Document "{title}" uploaded successfully.')
        else:
            messages.error(request, 'Title and file are required.')
        return redirect('accounts:my_documents')

    documents = EmployeeDocument.objects.filter(employee=employee).select_related('uploaded_by')
    from accounts.models import EmployeeDocument as _ED
    return render(request, 'accounts/my_documents.html', {
        'employee': employee,
        'documents': documents,
        'categories': _ED.CATEGORY_CHOICES,
    })


@login_required
def my_document_delete(request, doc_pk):
    """Employee deletes one of their own documents."""
    from accounts.models import EmployeeDocument
    doc = get_object_or_404(EmployeeDocument, pk=doc_pk)
    try:
        employee = request.user.employee
    except Employee.DoesNotExist:
        return redirect('dashboard:home')
    if doc.employee != employee and not request.user.is_superuser:
        messages.error(request, "You can only delete your own documents.")
        return redirect('accounts:my_documents')
    if request.method == 'POST':
        doc.file.delete(save=False)
        doc.delete()
        messages.success(request, 'Document deleted.')
    return redirect('accounts:my_documents')


@login_required
def document_delete(request, doc_pk):
    """HR/superuser deletes an employee document."""
    from accounts.models import EmployeeDocument
    doc = get_object_or_404(EmployeeDocument, pk=doc_pk)
    viewer = get_employee(request)
    is_privileged = (
        request.user.is_superuser
        or (viewer and (viewer.is_hr() or viewer.is_director() or viewer.is_ceo()))
    )
    if not is_privileged:
        return redirect('dashboard:home')
    employee_pk = doc.employee.pk
    if request.method == 'POST':
        doc.file.delete(save=False)
        doc.delete()
        messages.success(request, 'Document deleted.')
    return redirect(reverse('accounts:employee_history', args=[employee_pk]) + '#documents')


@login_required
def setup_2fa(request):
    """Allow any user to set up TOTP 2FA on their account."""
    try:
        employee = request.user.employee
    except Exception:
        messages.error(request, "No employee profile found.")
        return redirect('dashboard:home')

    try:
        import pyotp, qrcode, io, base64
    except ImportError:
        messages.error(request, "2FA library not installed. Contact admin.")
        return redirect('accounts:profile')

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'enable':
            token = request.POST.get('token', '').strip()
            secret = request.session.get('pending_totp_secret', '')
            if not secret:
                messages.error(request, "Session expired. Please start again.")
                return redirect('accounts:setup_2fa')
            totp = pyotp.TOTP(secret)
            if totp.verify(token, valid_window=1):
                employee.totp_secret = secret
                employee.totp_enabled = True
                employee.save(update_fields=['totp_secret', 'totp_enabled'])
                request.session.pop('pending_totp_secret', None)
                messages.success(request, "Two-factor authentication enabled successfully.")
                return redirect('accounts:profile')
            else:
                messages.error(request, "Invalid code. Please try again.")

        elif action == 'disable':
            employee.totp_enabled = False
            employee.totp_secret = ''
            employee.save(update_fields=['totp_enabled', 'totp_secret'])
            messages.success(request, "Two-factor authentication disabled.")
            return redirect('accounts:profile')

    # Generate new secret for setup
    secret = pyotp.random_base32()
    request.session['pending_totp_secret'] = secret
    totp = pyotp.TOTP(secret)
    uri = totp.provisioning_uri(
        name=request.user.email or request.user.username,
        issuer_name='AEF HRM'
    )
    img = qrcode.make(uri)
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    qr_b64 = base64.b64encode(buf.getvalue()).decode()

    return render(request, 'accounts/setup_2fa.html', {
        'employee': employee,
        'secret': secret,
        'qr_b64': qr_b64,
    })


def verify_2fa(request):
    """Step 2 of login: verify TOTP token."""
    pending_user_id = request.session.get('2fa_pending_user_id')
    if not pending_user_id:
        return redirect('accounts:login')

    from django.contrib.auth.models import User
    try:
        user = User.objects.get(pk=pending_user_id)
        employee = user.employee
    except Exception:
        return redirect('accounts:login')

    if request.method == 'POST':
        import pyotp
        token = request.POST.get('token', '').strip()
        totp = pyotp.TOTP(employee.totp_secret)
        if totp.verify(token, valid_window=1):
            request.session.pop('2fa_pending_user_id', None)
            next_url = request.session.pop('2fa_next', 'dashboard:home')
            login(request, user, backend='django.contrib.auth.backends.ModelBackend')
            return redirect(next_url)
        else:
            messages.error(request, "Invalid or expired code. Try again.")

    return render(request, 'accounts/verify_2fa.html', {
        'username': user.get_full_name() or user.username
    })


@login_required
def expiring_documents(request):
    """HR/superuser: list all employee documents expiring within 90 days."""
    viewer = get_employee(request)
    is_privileged = (
        request.user.is_superuser
        or (viewer and (viewer.is_hr() or viewer.is_director() or viewer.is_ceo()))
    )
    if not is_privileged:
        return redirect('dashboard:home')

    from accounts.models import EmployeeDocument
    from datetime import date, timedelta
    today = date.today()
    in_90 = today + timedelta(days=90)

    docs = (
        EmployeeDocument.objects
        .filter(expiry_date__isnull=False, expiry_date__lte=in_90)
        .select_related('employee__user', 'employee__department')
        .order_by('expiry_date')
    )
    return render(request, 'accounts/expiring_documents.html', {
        'docs': docs,
        'today': today,
    })


@login_required
def onboarding_list(request):
    """HR view: list all employees with their onboarding status."""
    if not (request.user.is_superuser or (hasattr(request.user, 'employee') and request.user.employee.is_hr())):
        messages.error(request, "Access denied.")
        return redirect('dashboard:home')

    from .models import OnboardingChecklist
    # Ensure checklists exist for all employees
    from django.db.models import Prefetch
    employees = Employee.objects.filter(is_active=True).select_related('user', 'department')
    for emp in employees:
        OnboardingChecklist.objects.get_or_create(employee=emp)

    checklists = OnboardingChecklist.objects.select_related(
        'employee__user', 'employee__department'
    ).filter(employee__is_active=True).order_by('employee__user__last_name')

    return render(request, 'accounts/onboarding_list.html', {'checklists': checklists})


@login_required
def onboarding_update(request, pk):
    """HR: toggle onboarding checklist items for an employee."""
    if not (request.user.is_superuser or (hasattr(request.user, 'employee') and request.user.employee.is_hr())):
        messages.error(request, "Access denied.")
        return redirect('dashboard:home')

    from .models import OnboardingChecklist
    checklist = get_object_or_404(OnboardingChecklist, pk=pk)

    if request.method == 'POST':
        fields = ['issue_contract', 'set_leave_balance', 'assign_manager',
                  'profile_photo', 'signature_captured', 'credentials_sent', 'id_document_uploaded']
        for field in fields:
            setattr(checklist, field, field in request.POST)
        checklist.notes = request.POST.get('notes', '')
        checklist.save()
        messages.success(request, f"Onboarding updated for {checklist.employee.get_full_name()}.")
        return redirect('accounts:onboarding_list')

    return render(request, 'accounts/onboarding_detail.html', {'checklist': checklist})


# ── Health Insurance ──────────────────────────────────────────────────────────

@login_required
def health_insurance_edit(request, pk):
    """HR / superuser only — manage health insurance data for an employee."""
    from datetime import date as _date
    target = get_object_or_404(Employee, pk=pk)
    try:
        viewer = request.user.employee
    except Employee.DoesNotExist:
        viewer = None
    is_hr = request.user.is_superuser or (viewer and (viewer.is_hr() or viewer.is_director() or viewer.is_ceo()))
    if not is_hr:
        messages.error(request, "Access denied. HR only.")
        return redirect('dashboard:home')

    if request.method == 'POST':
        p = request.POST
        marital = p.get('marital_status', 'single')
        if marital not in ('single', 'married'):
            marital = 'single'
        target.marital_status = marital
        target.save(update_fields=['marital_status'])

        # Clear existing dependants and rebuild from POST
        target.health_dependants.all().delete()

        today = _date.today()

        # Spouse (married only)
        if marital == 'married':
            sp_name = p.get('spouse_name', '').strip()
            sp_dob_str = p.get('spouse_dob', '').strip()
            if sp_name:
                sp_dob = None
                if sp_dob_str:
                    try:
                        from datetime import datetime
                        sp_dob = datetime.strptime(sp_dob_str, '%Y-%m-%d').date()
                    except ValueError:
                        pass
                HealthDependant.objects.create(
                    employee=target, relation=HealthDependant.SPOUSE,
                    full_name=sp_name, date_of_birth=sp_dob,
                )

        # Beneficiary children (married + single female, max 3)
        is_female = (target.sex == 'F')
        if marital == 'married' or is_female:
            for i in range(1, 4):
                name = p.get(f'ben_child_{i}_name', '').strip()
                dob_str = p.get(f'ben_child_{i}_dob', '').strip()
                if not name:
                    continue
                dob = None
                if dob_str:
                    try:
                        from datetime import datetime
                        dob = datetime.strptime(dob_str, '%Y-%m-%d').date()
                    except ValueError:
                        pass
                HealthDependant.objects.create(
                    employee=target, relation=HealthDependant.CHILD_BEN,
                    full_name=name, date_of_birth=dob,
                )

        # Non-beneficiary children (all statuses)
        nb_count_str = p.get('non_ben_count', '0')
        try:
            nb_count = max(0, min(20, int(nb_count_str)))
        except ValueError:
            nb_count = 0
        for i in range(1, nb_count + 1):
            name = p.get(f'non_ben_{i}_name', '').strip()
            dob_str = p.get(f'non_ben_{i}_dob', '').strip()
            if not name:
                continue
            dob = None
            if dob_str:
                try:
                    from datetime import datetime
                    dob = datetime.strptime(dob_str, '%Y-%m-%d').date()
                except ValueError:
                    pass
            HealthDependant.objects.create(
                employee=target, relation=HealthDependant.CHILD_OTHER,
                full_name=name, date_of_birth=dob,
            )

        messages.success(request, "Health insurance information saved.")
        return redirect('accounts:health_insurance_edit', pk=pk)

    dependants = target.health_dependants.all()
    spouse = dependants.filter(relation=HealthDependant.SPOUSE).first()
    ben_children = list(dependants.filter(relation=HealthDependant.CHILD_BEN))
    non_ben_children = list(dependants.filter(relation=HealthDependant.CHILD_OTHER))

    # Flag beneficiary children who are 18+
    from datetime import date as _date2
    today = _date2.today()
    for child in ben_children:
        child.is_overage = (child.age is not None and child.age >= 18)

    return render(request, 'accounts/health_insurance_edit.html', {
        'target': target,
        'is_hr': is_hr,
        'spouse': spouse,
        'ben_children': ben_children,
        'non_ben_children': non_ben_children,
    })


@login_required
def health_insurance_pdf(request, pk):
    """PDF report for a single employee's health insurance dependants."""
    from io import BytesIO
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer, Paragraph
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from django.http import HttpResponse
    from datetime import date

    emp = get_object_or_404(Employee, pk=pk)
    # Access: HR/superuser or the employee
    try:
        viewer = request.user.employee
    except Employee.DoesNotExist:
        viewer = None
    is_hr = request.user.is_superuser or (viewer and (viewer.is_hr() or viewer.is_director() or viewer.is_ceo()))
    if not is_hr and (viewer is None or viewer != emp):
        messages.error(request, "Access denied.")
        return redirect('dashboard:home')

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=18*mm, bottomMargin=18*mm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('title', parent=styles['Heading1'],
                                 fontSize=14, textColor=colors.HexColor('#0A4D68'), spaceAfter=4)
    sub_style = ParagraphStyle('sub', parent=styles['Normal'], fontSize=9,
                               textColor=colors.grey, spaceAfter=10)
    story = []

    story.append(Paragraph("AEF Health Insurance — Dependant Record", title_style))
    story.append(Paragraph(
        f"Employee: {emp.get_full_name()} | ID: {emp.employee_id} | "
        f"Department: {emp.department or '—'} | "
        f"Marital Status: {emp.get_marital_status_display()} | "
        f"Generated: {date.today().strftime('%d %B %Y')}",
        sub_style))
    story.append(Spacer(1, 4*mm))

    dependants = emp.health_dependants.all()
    if dependants:
        headers = ['#', 'Relation', 'Full Name', 'Date of Birth', 'Age', 'Status']
        rows = [headers]
        for i, d in enumerate(dependants, 1):
            dob_str = d.date_of_birth.strftime('%d/%m/%Y') if d.date_of_birth else '—'
            age_str = str(d.age) if d.age is not None else '—'
            if d.relation == HealthDependant.SPOUSE:
                status = 'Active'
            elif d.relation == HealthDependant.CHILD_BEN:
                status = 'Eligible' if d.insurance_active else 'Aged out (>18)'
            else:
                status = 'Non-beneficiary'
            rows.append([str(i), d.get_relation_display(), d.full_name, dob_str, age_str, status])

        col_w = [8*mm, 40*mm, 65*mm, 28*mm, 14*mm, 30*mm]
        t = Table(rows, colWidths=col_w, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND',  (0, 0), (-1, 0), colors.HexColor('#0A4D68')),
            ('TEXTCOLOR',   (0, 0), (-1, 0), colors.white),
            ('FONTSIZE',    (0, 0), (-1, 0), 9),
            ('FONTSIZE',    (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f9ff')]),
            ('GRID',        (0, 0), (-1, -1), 0.4, colors.HexColor('#cde8ef')),
            ('VALIGN',      (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING',  (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        story.append(t)
    else:
        story.append(Paragraph("No dependants recorded.", styles['Normal']))

    doc.build(story)
    buf.seek(0)
    resp = HttpResponse(buf, content_type='application/pdf')
    resp['Content-Disposition'] = f'inline; filename="health_insurance_{emp.employee_id}.pdf"'
    return resp


@login_required
def health_insurance_pdf_bulk(request):
    """PDF report listing ALL employees with their health insurance dependants."""
    from io import BytesIO
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer, Paragraph, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from django.http import HttpResponse
    from datetime import date

    if not (_is_hr_or_superuser(request.user)):
        messages.error(request, "Access denied.")
        return redirect('dashboard:home')

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=18*mm, bottomMargin=18*mm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('title', parent=styles['Heading1'],
                                 fontSize=13, textColor=colors.HexColor('#0A4D68'), spaceAfter=3)
    sub_style = ParagraphStyle('sub', parent=styles['Normal'], fontSize=8,
                               textColor=colors.grey, spaceAfter=8)
    emp_style = ParagraphStyle('emp', parent=styles['Heading2'],
                               fontSize=10, textColor=colors.HexColor('#0A4D68'),
                               spaceBefore=6, spaceAfter=2)
    story = []

    story.append(Paragraph("AEF Health Insurance — All Employees Report", title_style))
    story.append(Paragraph(f"Generated: {date.today().strftime('%d %B %Y')}", sub_style))
    story.append(Spacer(1, 4*mm))

    employees = Employee.objects.filter(is_active=True).prefetch_related('health_dependants')
    headers = ['Relation', 'Full Name', 'Date of Birth', 'Age', 'Status']

    for emp in employees:
        story.append(Paragraph(
            f"{emp.get_full_name()} ({emp.employee_id}) — {emp.department or '—'} — "
            f"{emp.get_marital_status_display()}",
            emp_style))
        dependants = emp.health_dependants.all()
        if dependants:
            rows = [headers]
            for d in dependants:
                dob_str = d.date_of_birth.strftime('%d/%m/%Y') if d.date_of_birth else '—'
                age_str = str(d.age) if d.age is not None else '—'
                if d.relation == HealthDependant.SPOUSE:
                    status = 'Active'
                elif d.relation == HealthDependant.CHILD_BEN:
                    status = 'Eligible' if d.insurance_active else 'Aged out'
                else:
                    status = 'Non-beneficiary'
                rows.append([d.get_relation_display(), d.full_name, dob_str, age_str, status])
            col_w = [38*mm, 62*mm, 28*mm, 14*mm, 30*mm]
            t = Table(rows, colWidths=col_w, repeatRows=1)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#31b8cf')),
                ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
                ('FONTSIZE',   (0, 0), (-1, 0), 8),
                ('FONTSIZE',   (0, 1), (-1, -1), 7.5),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f9ff')]),
                ('GRID',       (0, 0), (-1, -1), 0.4, colors.HexColor('#cde8ef')),
                ('TOPPADDING', (0, 0), (-1, -1), 2),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ]))
            story.append(t)
        else:
            story.append(Paragraph("No dependants recorded.", ParagraphStyle('nd', parent=styles['Normal'],
                                                                              fontSize=8, textColor=colors.grey,
                                                                              spaceAfter=4)))
        story.append(Spacer(1, 3*mm))

    doc.build(story)
    buf.seek(0)
    resp = HttpResponse(buf, content_type='application/pdf')
    resp['Content-Disposition'] = 'inline; filename="health_insurance_all_employees.pdf"'
    return resp


@login_required
def employee_search_api(request):
    """AJAX endpoint: search employees by name or employee_id. Returns JSON."""
    q = request.GET.get('q', '').strip()
    results = []
    if len(q) >= 2:
        qs = Employee.objects.filter(
            Q(user__first_name__icontains=q) |
            Q(user__last_name__icontains=q) |
            Q(employee_id__icontains=q),
            is_active=True,
        ).exclude(user=request.user).select_related('user', 'department')[:12]
        results = [
            {
                'id': e.pk,
                'label': e.get_full_name() + (' (' + e.employee_id + ')' if e.employee_id else ''),
                'name': e.get_full_name(),
                'emp_id': e.employee_id,
                'dept': str(e.department) if e.department else '',
                'role': e.get_role_display(),
            }
            for e in qs
        ]
    return JsonResponse({'results': results})


# ── Audit Log View ────────────────────────────────────────────────────────────
@login_required
def audit_log_view(request):
    from dashboard.models import AuditLog
    emp = request.user.employee
    if not (emp.is_hr() or emp.is_ceo() or request.user.is_superuser):
        from django.core.exceptions import PermissionDenied
        raise PermissionDenied

    qs = AuditLog.objects.select_related('user', 'target_user').all()

    # Filters
    action_filter = request.GET.get('action', '')
    user_filter   = request.GET.get('user', '')
    date_from     = request.GET.get('date_from', '')
    date_to       = request.GET.get('date_to', '')
    search        = request.GET.get('q', '')

    if action_filter:
        qs = qs.filter(action=action_filter)
    if user_filter:
        qs = qs.filter(
            Q(user__first_name__icontains=user_filter) |
            Q(user__last_name__icontains=user_filter) |
            Q(user__username__icontains=user_filter)
        )
    if search:
        qs = qs.filter(description__icontains=search)
    if date_from:
        try:
            from datetime import datetime
            qs = qs.filter(timestamp__date__gte=datetime.strptime(date_from, '%Y-%m-%d').date())
        except ValueError:
            pass
    if date_to:
        try:
            from datetime import datetime
            qs = qs.filter(timestamp__date__lte=datetime.strptime(date_to, '%Y-%m-%d').date())
        except ValueError:
            pass

    from django.core.paginator import Paginator
    paginator = Paginator(qs, 50)
    page = request.GET.get('page', 1)
    logs = paginator.get_page(page)

    return render(request, 'accounts/audit_log.html', {
        'logs': logs,
        'action_choices': AuditLog.ACTION_CHOICES,
        'filters': {
            'action': action_filter,
            'user': user_filter,
            'date_from': date_from,
            'date_to': date_to,
            'q': search,
        },
    })


# ─────────────────────────────────────────────
#  ACTING ROLE PANEL  (HR-only)
# ─────────────────────────────────────────────

ACTING_ROLE_CHOICES = [
    ('admin_director',      'Acting Administration Director'),
    ('finance_director',    'Acting Finance Director'),
    ('medical_director',    'Acting Medical Director'),
    ('manager',             'Acting Line Manager'),
    ('unit_head',           'Acting Unit Head'),
    ('nurse_superintendent','Acting Nurse Superintendent'),
    ('hr',                  'Acting HR'),
]

ACTING_ROLE_LABELS = dict(ACTING_ROLE_CHOICES)


@login_required
def acting_roles_list(request):
    """HR-only panel: view current acting assignments and assign new ones."""
    emp = getattr(request.user, 'employee', None)
    if not emp or not (emp.is_hr() or request.user.is_superuser):
        messages.error(request, "Access denied. HR only.")
        return redirect('dashboard:home')

    from datetime import date
    today = date.today()

    # Current active acting assignments
    active = Employee.objects.filter(
        acting_role__gt='',
        is_active=True,
    ).select_related('user', 'acting_for__user').order_by('user__last_name')

    # All employees for the select dropdown
    all_employees = Employee.objects.filter(is_active=True).select_related('user').order_by('user__last_name', 'user__first_name')

    return render(request, 'accounts/acting_roles.html', {
        'active_acting': active,
        'all_employees': all_employees,
        'role_choices': ACTING_ROLE_CHOICES,
        'today': today,
    })


@login_required
def acting_role_assign(request):
    """HR-only: assign an acting role to an employee."""
    emp = getattr(request.user, 'employee', None)
    if not emp or not (emp.is_hr() or request.user.is_superuser):
        messages.error(request, "Access denied. HR only.")
        return redirect('dashboard:home')

    if request.method != 'POST':
        return redirect('accounts:acting_roles_list')

    from datetime import date
    from notifications.utils import notify

    employee_id  = request.POST.get('employee_id')
    acting_role  = request.POST.get('acting_role')
    acting_for_id = request.POST.get('acting_for_id') or None
    start_date   = request.POST.get('start_date') or str(date.today())
    end_date     = request.POST.get('end_date')

    if not employee_id or not acting_role or not end_date:
        messages.error(request, "Employee, acting role, and end date are required.")
        return redirect('accounts:acting_roles_list')

    if acting_role not in dict(ACTING_ROLE_CHOICES):
        messages.error(request, "Invalid acting role selected.")
        return redirect('accounts:acting_roles_list')

    try:
        target = Employee.objects.get(pk=employee_id)
    except Employee.DoesNotExist:
        messages.error(request, "Employee not found.")
        return redirect('accounts:acting_roles_list')

    covering = None
    if acting_for_id:
        try:
            covering = Employee.objects.get(pk=acting_for_id)
        except Employee.DoesNotExist:
            pass

    target.acting_role  = acting_role
    target.acting_for   = covering
    target.acting_since = start_date
    target.acting_until = end_date
    target.save(update_fields=['acting_role', 'acting_for', 'acting_since', 'acting_until'])

    role_label = ACTING_ROLE_LABELS.get(acting_role, acting_role)
    covering_line = f" covering for {covering.get_full_name()}" if covering else ""

    notify(
        target.user,
        f'Acting Role Assigned — {role_label}',
        f'You have been designated as {role_label}{covering_line} '
        f'effective {start_date} until {end_date}. '
        f'During this period you will receive and be able to act on all responsibilities '
        f'associated with this role. Please log in to review any pending items.',
        notification_type='leave_approved',
    )

    messages.success(request, f"{target.get_full_name()} assigned as {role_label} until {end_date}.")
    return redirect('accounts:acting_roles_list')


@login_required
def acting_role_remove(request, pk):
    """HR-only: remove an acting role assignment immediately."""
    emp = getattr(request.user, 'employee', None)
    if not emp or not (emp.is_hr() or request.user.is_superuser):
        messages.error(request, "Access denied. HR only.")
        return redirect('dashboard:home')

    if request.method != 'POST':
        return redirect('accounts:acting_roles_list')

    from notifications.utils import notify

    try:
        target = Employee.objects.get(pk=pk)
    except Employee.DoesNotExist:
        messages.error(request, "Employee not found.")
        return redirect('accounts:acting_roles_list')

    role_label = ACTING_ROLE_LABELS.get(target.acting_role, target.acting_role)
    old_role = role_label

    target.acting_role  = ''
    target.acting_for   = None
    target.acting_since = None
    target.acting_until = None
    target.save(update_fields=['acting_role', 'acting_for', 'acting_since', 'acting_until'])

    notify(
        target.user,
        f'Acting Role Ended — {old_role}',
        f'Your acting assignment as {old_role} has been ended by HR. '
        f'Your regular role and responsibilities have been restored.',
        notification_type='leave_approved',
    )

    messages.success(request, f"Acting role removed for {target.get_full_name()}.")
    return redirect('accounts:acting_roles_list')


# ── System Backup ─────────────────────────────────────────────────────────────

BACKUP_DIR = '/root/backups'
BACKUP_SCRIPT = '/usr/local/bin/hrm_backup_now.sh'


def _superuser_only(request):
    return request.user.is_authenticated and request.user.is_superuser


@login_required
def backup_panel(request):
    if not _superuser_only(request):
        messages.error(request, "Superuser access only.")
        return redirect('dashboard:home')

    import os
    backups = []
    if os.path.isdir(BACKUP_DIR):
        for folder in sorted(os.listdir(BACKUP_DIR), reverse=True):
            folder_path = os.path.join(BACKUP_DIR, folder)
            if not os.path.isdir(folder_path):
                continue
            files = []
            total_bytes = 0
            for fname in os.listdir(folder_path):
                fpath = os.path.join(folder_path, fname)
                size = os.path.getsize(fpath)
                total_bytes += size
                files.append({'name': fname, 'size_mb': round(size / 1024 / 1024, 1)})
            backups.append({
                'folder': folder,
                'files': sorted(files, key=lambda x: x['name']),
                'total_mb': round(total_bytes / 1024 / 1024, 1),
            })

    return render(request, 'accounts/backup_panel.html', {'backups': backups})


@login_required
def backup_run(request):
    if not _superuser_only(request):
        messages.error(request, "Superuser access only.")
        return redirect('dashboard:home')
    if request.method != 'POST':
        return redirect('accounts:backup_panel')

    import subprocess, os
    if not os.path.isfile(BACKUP_SCRIPT):
        messages.error(request, f"Backup script not found at {BACKUP_SCRIPT}.")
        return redirect('accounts:backup_panel')

    try:
        result = subprocess.run(
            [BACKUP_SCRIPT],
            capture_output=True, text=True, timeout=300
        )
        if result.returncode == 0:
            messages.success(request, "Backup completed successfully.")
        else:
            messages.error(request, f"Backup failed: {result.stderr[:300]}")
    except subprocess.TimeoutExpired:
        messages.error(request, "Backup timed out after 5 minutes.")
    except Exception as e:
        messages.error(request, f"Backup error: {e}")

    return redirect('accounts:backup_panel')


@login_required
def backup_download(request, filename):
    if not _superuser_only(request):
        messages.error(request, "Superuser access only.")
        return redirect('dashboard:home')

    import os
    from django.http import FileResponse, Http404
    # filename is like "hrm_2026-08-18_082919/database_....dump"
    full_path = os.path.join(BACKUP_DIR, filename)
    # Security: must stay within BACKUP_DIR
    real_path = os.path.realpath(full_path)
    real_base = os.path.realpath(BACKUP_DIR)
    if not real_path.startswith(real_base + os.sep):
        raise Http404
    if not os.path.isfile(real_path):
        raise Http404

    return FileResponse(
        open(real_path, 'rb'),
        as_attachment=True,
        filename=os.path.basename(real_path),
    )


@login_required
def backup_delete(request, folder):
    if not _superuser_only(request):
        messages.error(request, "Superuser access only.")
        return redirect('dashboard:home')
    if request.method != 'POST':
        return redirect('accounts:backup_panel')

    import os, shutil
    folder_path = os.path.join(BACKUP_DIR, folder)
    real_path = os.path.realpath(folder_path)
    real_base = os.path.realpath(BACKUP_DIR)
    if not real_path.startswith(real_base + os.sep):
        messages.error(request, "Invalid path.")
        return redirect('accounts:backup_panel')
    if os.path.isdir(real_path):
        shutil.rmtree(real_path)
        messages.success(request, f"Backup '{folder}' deleted.")
    else:
        messages.error(request, "Backup not found.")
    return redirect('accounts:backup_panel')
